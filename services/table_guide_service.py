"""Table Guide Builder 서비스 레이어.

Phase 2: Base Definition + Net/Recode
Phase 3: Banner Management + Sort + SubBanner
Phase 4: Special Instructions + Full Compile + Export
"""

import io
import json as _json
import logging
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from typing import Callable, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models.survey import (
    Banner, BannerPoint, SurveyDocument, SurveyQuestion, TableGuideDocument,
)
from services.llm_client import (
    DEFAULT_MODEL, MODEL_TITLE_GENERATOR, call_llm, call_llm_json,
)

logger = logging.getLogger(__name__)


def _extract_json_from_text(text: str) -> dict:
    """텍스트에서 JSON 객체를 추출.

    마크다운 코드 펜스, 설명 텍스트, 후행 텍스트 등을 제거하고
    첫 번째 `{` ~ 마지막 `}`를 JSON으로 파싱.
    """


    cleaned = text.strip()

    # 1) 마크다운 코드 펜스 제거
    if cleaned.startswith("```"):
        lines = cleaned.split("\n")
        lines = lines[1:]  # opening fence
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        cleaned = "\n".join(lines).strip()

    # 2) 직접 파싱 시도
    try:
        return _json.loads(cleaned)
    except (ValueError, _json.JSONDecodeError):
        pass

    # 3) 첫 `{` ~ 마지막 `}` 추출
    first_brace = cleaned.find("{")
    last_brace = cleaned.rfind("}")
    if first_brace != -1 and last_brace > first_brace:
        candidate = cleaned[first_brace:last_brace + 1]
        return _json.loads(candidate)

    raise ValueError(f"No JSON object found in text (length={len(text)})")


def _call_llm_json_with_fallback(system_prompt: str, user_prompt: str,
                                  model: str, **kwargs) -> dict:
    """call_llm_json + text-mode 폴백 (2회 재시도).

    일부 모델/프록시에서 response_format=json_object가 빈 응답을 반환하는 경우
    call_llm(text mode)로 재시도 후 JSON 파싱.
    """
    # 1차: JSON 모드
    try:
        return call_llm_json(system_prompt, user_prompt, model, **kwargs)
    except Exception as e:
        logger.warning(f"call_llm_json failed ({e}), retrying with text mode...")

    # 2차/3차: 텍스트 모드 (최대 2회)
    full_prompt = (
        f"{system_prompt}\n\n"
        "IMPORTANT: Respond with valid JSON only. No markdown code fences, no explanation.\n\n"
        f"{user_prompt}"
    )
    max_tokens = kwargs.get("max_tokens", 8192)
    temperature = kwargs.get("temperature", 0.2)
    top_p = kwargs.get("top_p", 0.8)

    last_error = None
    for attempt in range(2):
        try:
            raw = call_llm(full_prompt, model, max_tokens=max_tokens,
                           temperature=temperature, top_p=top_p)
            return _extract_json_from_text(raw)
        except Exception as e:
            last_error = e
            if attempt == 0:
                logger.warning(f"Text-mode attempt {attempt + 1} failed ({e}), retrying...")

    raise last_error

# ── 모델 할당 ────────────────────────────────────────────────────
MODEL_INTELLIGENCE = MODEL_TITLE_GENERATOR          # GPT-5 — 깊은 이해력 필요
MODEL_BASE_GENERATOR = DEFAULT_MODEL               # GPT-4.1-mini
MODEL_NET_GENERATOR = DEFAULT_MODEL                # GPT-4.1-mini
MODEL_BANNER_SUGGESTER = MODEL_TITLE_GENERATOR     # GPT-5 — 전문가 수준 배너 설계
MODEL_SUBBANNER_SUGGESTER = DEFAULT_MODEL          # GPT-4.1-mini
MODEL_SPECIAL_INSTRUCTIONS = DEFAULT_MODEL         # GPT-4.1-mini

BATCH_SIZE = 20

_PRIORITY_MAP = {"critical": "high", "important": "high", "supplementary": "medium"}


# ── 공통 유틸 ──────────────────────────────────────────────────────

def _format_questions_compact(questions: List[SurveyQuestion],
                               include_options: bool = False,
                               max_option_len: int = 150) -> str:
    """문항 리스트를 LLM 프롬프트용 compact 텍스트로 변환.

    Args:
        questions: 문항 리스트
        include_options: 보기를 포함할지 여부
        max_option_len: 보기 텍스트 최대 길이

    Returns:
        포맷된 텍스트 문자열
    """
    seen = set()
    lines = []
    for q in questions:
        if q.question_number in seen:
            continue
        seen.add(q.question_number)

        text = (q.question_text or "").replace("\n", " ")[:100]
        qtype = q.question_type or ""
        filt = f" [Filter: {(q.filter_condition or '')[:50]}]" if q.filter_condition else ""
        line = f"[{q.question_number}] {text} ({qtype})"
        if include_options and q.answer_options:
            opts = q.answer_options_compact()
            if len(opts) > max_option_len:
                opts = opts[:max_option_len] + "..."
            line += f"\n  Options: {opts}"
        if filt:
            line += filt
        lines.append(line)
    return "\n".join(lines)


def _format_questions_full(questions: List[SurveyQuestion]) -> str:
    """선정된 후보 문항의 FULL 상세 정보를 LLM 프롬프트용으로 변환.

    배너 설계 단계에서 보기 코드, 필터, 스킵 로직 등 모든 정보 포함.
    """
    seen = set()
    lines = []
    for q in questions:
        if q.question_number in seen:
            continue
        seen.add(q.question_number)

        lines.append(f"[{q.question_number}]")
        lines.append(f"  Text: {q.question_text}")
        lines.append(f"  Type: {q.question_type or 'SA'}")
        if q.answer_options:
            lines.append(f"  Options: {q.answer_options_compact()}")
        if q.filter_condition:
            lines.append(f"  Filter: {q.filter_condition}")
        if q.skip_logic:
            lines.append(f"  Skip: {q.skip_logic_display()}")
        lines.append("")
    return "\n".join(lines)


def _build_code_map(questions: List[SurveyQuestion]) -> dict:
    """문항번호 → 유효 코드 리스트 맵 생성 (검증용).

    Returns:
        dict: {"SQ1": ["1", "2", "3"], "SQ2": ["1", "2", "3", "4"]}
    """
    code_map = {}
    for q in questions:
        if q.question_number not in code_map and q.answer_options:
            code_map[q.question_number] = [opt.code for opt in q.answer_options]
    return code_map


# ── Structured Study Parameters (Change 1) ────────────────────────────

def _build_structured_study_params(survey_context: str,
                                    intelligence: dict | None) -> str:
    """intelligence에서 핵심 파라미터를 추출하여 구조화된 블록 생성.

    LLM이 freeform markdown에서 파싱하지 않아도 되도록 명시적 블록을 생성.
    """
    if not intelligence:
        return ""

    parts = ["## STUDY PARAMETERS (use these to guide ALL decisions)"]
    client = intelligence.get("client_name", "")
    if client:
        parts.append(f"- **Client/Brand**: {client}")
    study_type = intelligence.get("study_type", "")
    if study_type:
        parts.append(f"- **Study Type**: {study_type}")
    objectives = intelligence.get("research_objectives", [])
    if objectives:
        parts.append("- **Research Objectives**:")
        for obj in objectives[:5]:
            parts.append(f"  - {obj}")
    segments = intelligence.get("key_segments", [])
    if segments:
        parts.append("- **Key Segments**:")
        for seg in segments[:6]:
            seg_name = seg.get("name", "") if isinstance(seg, dict) else str(seg)
            seg_q = seg.get("question", "") if isinstance(seg, dict) else ""
            seg_type = seg.get("type", "") if isinstance(seg, dict) else ""
            label = seg_name
            if seg_q:
                label += f" ({seg_q})"
            if seg_type:
                label += f" [{seg_type}]"
            parts.append(f"  - {label}")
    banner_recs = intelligence.get("banner_recommendations", [])
    if banner_recs:
        parts.append("- **Banner Recommendations**:")
        for rec in banner_recs[:4]:
            rec_name = rec.get("name", "") if isinstance(rec, dict) else str(rec)
            parts.append(f"  - {rec_name}")

    if len(parts) <= 1:
        return ""
    parts.append("")
    return "\n".join(parts)


# ── Domain Category Hints (Change 4) ─────────────────────────────────

_DOMAIN_CATEGORY_HINTS: dict[str, dict] = {
    "brand tracking": {
        "description": "Brand health & equity tracking study",
        "recommended_categories": [
            "Brand Funnel (awareness → consideration → trial → usage → loyalty)",
            "Competitive Landscape (client vs key competitors, switching dynamics)",
            "Media & Touchpoints (ad recall, channel effectiveness, media mix)",
            "Attitudinal Segments (brand image, emotional connection, NPS tiers)",
        ],
    },
    "u&a": {
        "description": "Usage & Attitude study",
        "recommended_categories": [
            "Category Engagement (usage frequency, occasion, repertoire breadth)",
            "Brand Repertoire (primary brand, brand set, switching patterns)",
            "Need States (functional vs emotional needs, unmet needs)",
            "User Typology (heavy/medium/light, loyal/switcher/lapsed)",
        ],
    },
    "satisfaction": {
        "description": "Customer satisfaction / NPS study",
        "recommended_categories": [
            "Touchpoint Experience (channel satisfaction, journey stage)",
            "CSAT-Loyalty Nexus (satisfaction × NPS × retention intent)",
            "Problem Resolution (complaint type, resolution satisfaction)",
            "Customer Lifetime Value (tenure × spend × advocacy)",
        ],
    },
    "ad test": {
        "description": "Advertising / Creative testing study",
        "recommended_categories": [
            "Ad Impact (recall, persuasion, brand linkage)",
            "Creative Diagnostics (message clarity, emotional response, uniqueness)",
            "Target Receptivity (by segment, by media exposure, by brand relationship)",
            "Competitive Context (category ad clutter, share of voice)",
        ],
    },
    "concept test": {
        "description": "Concept / Product testing study",
        "recommended_categories": [
            "Concept Appeal (purchase intent, uniqueness, relevance)",
            "Need Fit (problem-solution fit, unmet need addressal)",
            "Price Sensitivity (willingness to pay, value perception)",
            "Target Segments (early adopter, mainstream, skeptic)",
        ],
    },
    "product test": {
        "description": "Product testing / sensory evaluation study",
        "recommended_categories": [
            "Product Experience (overall liking, attribute ratings)",
            "Sensory Profile (taste, texture, appearance, aroma)",
            "Usage Context (occasion, preparation method, pairing)",
            "Preference Segments (by product variant, by user type)",
        ],
    },
    "segmentation": {
        "description": "Market segmentation study",
        "recommended_categories": [
            "Behavioral Segments (usage patterns, purchase behavior)",
            "Attitudinal Segments (values, motivations, lifestyle)",
            "Needs-Based Segments (primary need, occasion-driven needs)",
            "Value Segments (price sensitivity × quality expectation)",
        ],
    },
}


def _get_domain_guidance(intelligence: dict | None) -> str:
    """study_type partial matching으로 해당 도메인 힌트를 반환.

    매칭 실패 시 빈 문자열 (graceful degradation).
    """
    if not intelligence:
        return ""
    study_type = (intelligence.get("study_type", "") or "").lower()
    if not study_type:
        return ""

    for domain_key, hints in _DOMAIN_CATEGORY_HINTS.items():
        if domain_key in study_type or study_type in domain_key:
            lines = [f"## DOMAIN GUIDANCE — {hints['description']}",
                     "The following category themes are typical for this study type. "
                     "Use them as inspiration (adapt to this specific study):"]
            for cat in hints["recommended_categories"]:
                lines.append(f"- {cat}")
            lines.append("")
            return "\n".join(lines)
    return ""


# ── Domain Composite Examples (Change 9) ─────────────────────────────

_DOMAIN_COMPOSITE_EXAMPLES: dict[str, str] = {
    "brand tracking": """## DOMAIN COMPOSITE EXAMPLES (Brand Tracking)
- **Funnel Stage**: awareness Q × consideration Q × trial Q → "Loyal Advocate", "Aware Non-Considerer", "Unaware"
- **Brand Equity Segment**: overall opinion Q × recommendation Q → "Brand Champion", "Passive Positive", "Detractor"
- **Media-Influenced**: ad recall Q × brand consideration Q → "Ad-Driven Considerer", "Organic Considerer", "Exposed Non-Considerer"
- **Competitive Vulnerability**: client brand satisfaction Q × competitor consideration Q → "Secure Loyal", "At-Risk", "Already Lost"
""",
    "u&a": """## DOMAIN COMPOSITE EXAMPLES (U&A)
- **Category Engagement**: usage frequency Q × number of brands used Q → "Heavy Loyalist", "Heavy Switcher", "Light User"
- **Need-Based Segment**: primary need Q × satisfaction Q → "Satisfied Core Need", "Unmet Need Seeker", "Indifferent"
- **Brand Relationship**: brand usage Q × purchase intent Q → "Committed User", "Habitual User", "Trial Seeker"
- **Occasion Typology**: usage occasion Q × usage frequency Q → "Daily Routine", "Special Occasion", "Impulse"
""",
    "satisfaction": """## DOMAIN COMPOSITE EXAMPLES (Satisfaction)
- **Loyalty Risk**: overall satisfaction Q × NPS Q → "Secure Promoter", "Vulnerable Passive", "Active Detractor"
- **Service Recovery**: problem experience Q × resolution satisfaction Q → "Recovered", "Unresolved Complainer", "Silent Sufferer"
- **Value Segment**: satisfaction Q × price sensitivity Q → "Value Advocate", "Price-Trapped", "Premium Loyalist"
- **Engagement Level**: contact frequency Q × satisfaction Q × NPS Q → "Engaged Promoter", "Disengaged Passive", "Frequent Detractor"
""",
    "ad test": """## DOMAIN COMPOSITE EXAMPLES (Ad Test)
- **Ad Effectiveness**: ad recall Q × brand linkage Q → "Strong Brander", "Generic Recall", "No Impact"
- **Persuasion Segment**: purchase intent shift Q × ad liking Q → "Persuaded Liker", "Liked Not Persuaded", "Resistant"
- **Creative Resonance**: message clarity Q × emotional response Q → "Head & Heart", "Rational Only", "Emotional Only"
""",
    "concept test": """## DOMAIN COMPOSITE EXAMPLES (Concept Test)
- **Concept Viability**: purchase intent Q × uniqueness Q → "Must-Have", "Nice-to-Have", "Me-Too"
- **Target Fit**: relevance Q × unmet need Q → "Perfect Fit", "Partial Fit", "No Fit"
- **Adoption Readiness**: interest Q × price acceptance Q → "Early Adopter", "Wait-and-See", "Price Barrier"
""",
}


def _get_domain_composite_examples(intelligence: dict | None) -> str:
    """도메인별 composite 패턴 예시를 반환.

    매칭 실패 시 빈 문자열 (graceful degradation).
    """
    if not intelligence:
        return ""
    study_type = (intelligence.get("study_type", "") or "").lower()
    if not study_type:
        return ""

    for domain_key, examples in _DOMAIN_COMPOSITE_EXAMPLES.items():
        if domain_key in study_type or study_type in domain_key:
            return examples
    return ""


# ── Role-Banner Relevance (Change 8 — semantic assignment) ───────────

_ROLE_BANNER_RELEVANCE: dict[str, dict[str, float]] = {
    "awareness": {
        "brand": 1.0, "funnel": 1.0, "media": 0.9, "competitive": 0.9,
        "demographic": 0.8, "attitude": 0.7, "segment": 1.0,
    },
    "usage_experience": {
        "usage": 1.0, "behavior": 1.0, "brand": 0.9, "satisfaction": 0.8,
        "demographic": 0.8, "segment": 1.0, "need": 0.9,
    },
    "evaluation": {
        "satisfaction": 1.0, "attitude": 1.0, "brand": 0.9, "loyalty": 0.9,
        "demographic": 0.8, "segment": 1.0, "competitive": 0.8,
    },
    "intent_loyalty": {
        "loyalty": 1.0, "brand": 1.0, "funnel": 0.9, "satisfaction": 0.9,
        "demographic": 0.8, "segment": 1.0, "competitive": 0.9,
    },
    "other": {
        "demographic": 0.8, "segment": 1.0, "brand": 0.8,
        "attitude": 0.8, "behavior": 0.8,
    },
}

_MIN_RELEVANCE_THRESHOLD = 0.0  # 초기값: 필터링 없이 정렬만 (안전)


def _score_banner_relevance(question: SurveyQuestion,
                             banner: Banner) -> float:
    """문항-배너 의미적 적합도 점수를 반환.

    - Composite 배너 → 항상 1.0 (전략 세그먼트)
    - role 미설정 → 0.8 (기본 포함)
    - role + category keyword 매칭 → 가중치 적용
    """
    if banner.banner_type == "composite":
        return 1.0

    role = (question.role or "").lower()
    if not role:
        return 0.8

    role_weights = _ROLE_BANNER_RELEVANCE.get(role, _ROLE_BANNER_RELEVANCE.get("other", {}))
    if not role_weights:
        return 0.8

    # 배너 카테고리 + 이름에서 keyword 매칭
    banner_text = f"{(banner.category or '')} {(banner.name or '')}".lower()
    best_weight = 0.8  # 기본값
    for keyword, weight in role_weights.items():
        if keyword in banner_text:
            best_weight = max(best_weight, weight)
    return best_weight


# ======================================================================
# Survey Intelligence Analysis
# ======================================================================

_INTELLIGENCE_SYSTEM_PROMPT = """You are a senior marketing research strategist analyzing a complete survey questionnaire.
Your task is to deeply understand the survey's purpose, client, research framework, and recommend optimal analysis dimensions.

## User-Provided Context
If the user has specified a Client Brand or Study Objective, use these as authoritative ground truth:
- **Client Brand**: The brand commissioning the study. All brand-based segmentation should center on this brand (e.g., "Client Brand Users" vs "Competitor Users"). If not provided, infer from question texts and answer options.
- **Study Objective**: The stated research purpose. Use this to prioritize which analytical dimensions and segments are most relevant. If not provided, infer from the survey flow.

## Analysis Tasks
1. **Client/Brand Identification**: Use the user-provided Client Brand if available. Otherwise extract brand/company names from question texts, answer options, and instructions.
2. **Study Type Classification**: Classify into one of: Brand Tracking, U&A (Usage & Attitude), Satisfaction/NPS, Ad/Creative Test, Concept Test, Product Test, Segmentation, Omnibus, Custom. Use the most specific applicable type.
3. **Research Objectives**: If user provided a Study Objective, incorporate it as the primary objective and infer 2-4 additional objectives from the survey flow. Otherwise infer 3-5 core research questions.
4. **Analysis Framework**: Map each question number to its role in the study: screening, demographics, awareness, usage_experience, evaluation, intent_loyalty, or other.
5. **Key Segment Variables**: Identify the most important cross-analysis variables — demographics (age, gender, region), behavioral (usage, purchase), attitudinal (satisfaction, preference), brand-based (client brand users vs competitors).
6. **Banner Recommendations**: Suggest 2-4 banner groupings that would yield the most actionable cross-tabulation insights for the client brand.

## JSON Output Format
{
  "client_name": "string or empty if not identifiable",
  "study_type": "string",
  "research_objectives": ["objective1", "objective2", "objective3"],
  "analysis_framework": {
    "screening": ["S1", "S2"],
    "demographics": ["DQ1", "DQ2"],
    "awareness": ["Q1", "Q2"],
    "usage_experience": ["Q3", "Q4"],
    "evaluation": ["Q5", "Q6"],
    "intent_loyalty": ["Q7", "Q8"],
    "other": ["Q9"]
  },
  "key_segments": [
    {"question": "S2", "name": "Age Group", "type": "demographic"},
    {"question": "Q3", "name": "Brand Users", "type": "behavioral"}
  ],
  "banner_recommendations": [
    {
      "name": "Demographics",
      "rationale": "Standard demographic cuts for profiling",
      "points": ["S1(Gender)", "S2(Age Group)"]
    }
  ]
}"""


def analyze_survey_intelligence(questions: List[SurveyQuestion],
                                language: str = "ko",
                                client_brand: str = "",
                                study_objective: str = "") -> dict:
    """설문지 전체를 분석하여 구조화된 Survey Intelligence를 반환.

    단일 LLM 호출로 클라이언트, 조사유형, 목적, 프레임워크, 세그먼트를 추출.

    Args:
        questions: 전체 문항 리스트
        language: 설문지 언어
        client_brand: 사용자가 입력한 클라이언트 브랜드명 (e.g. "Hyundai")
        study_objective: 사용자가 입력한 조사 목적

    Returns:
        dict: intelligence 결과 (client_name, study_type, research_objectives, ...)
    """
    if not questions:
        return {}

    # 중복 문항번호 제거
    seen = set()
    unique_qs = []
    for q in questions:
        if q.question_number not in seen:
            seen.add(q.question_number)
            unique_qs.append(q)

    # 문항 요약 생성 (토큰 효율을 위해 compact)
    lines = []
    if client_brand:
        lines.append(f"Client Brand: {client_brand}")
    if study_objective:
        lines.append(f"Study Objective: {study_objective}")
    if lines:
        lines.append("")
    lines.append(f"Survey questionnaire with {len(unique_qs)} questions (language: {language}):\n")
    for q in unique_qs:
        text = (q.question_text or "").replace("\n", " ")[:100]
        qtype = q.question_type or ""
        opts = q.answer_options_compact()
        if len(opts) > 150:
            opts = opts[:150] + "..."
        filt = f" [Filter: {(q.filter_condition or '')[:50]}]" if q.filter_condition else ""
        line = f"[{q.question_number}] {text} ({qtype})"
        if opts:
            line += f"\n  Options: {opts}"
        if filt:
            line += filt
        lines.append(line)

    user_prompt = "\n".join(lines)

    try:
        result = call_llm_json(_INTELLIGENCE_SYSTEM_PROMPT, user_prompt,
                               MODEL_INTELLIGENCE, max_tokens=8192)
        # 기본값 보장
        result.setdefault("client_name", "")
        result.setdefault("study_type", "")
        result.setdefault("research_objectives", [])
        result.setdefault("analysis_framework", {})
        result.setdefault("key_segments", [])
        result.setdefault("banner_recommendations", [])
        # 사용자 입력이 있으면 LLM 결과 보강
        if client_brand and not result["client_name"]:
            result["client_name"] = client_brand
        return result
    except Exception as e:
        logger.error(f"Survey intelligence analysis failed: {e}")
        # 실패해도 사용자 입력 정보는 보존
        fallback = {"client_name": client_brand, "study_type": "",
                     "research_objectives": [], "analysis_framework": {},
                     "key_segments": [], "banner_recommendations": []}
        if study_objective:
            fallback["research_objectives"] = [study_objective]
        return fallback


# ======================================================================
# Phase 2: Base Definition
# ======================================================================

_BASE_SYSTEM_PROMPT = """You are a DP specialist for marketing research cross-tabulation.
Your task is to generate a human-readable Base description for each survey question.

## Rules
1. If a question has a filter_condition, describe WHO is included in the base.
   - Convert raw filter codes like "Q2=1,2" into readable labels using the referenced question's text and answer options.
   - Example: Q2=1,2 where Q2 is "Brand Awareness" with options 1=Aware, 2=Used → "Q2(Brand Awareness) 'Aware' or 'Used' respondents"
2. If no filter_condition exists, the base is "All Respondents".
3. Keep descriptions concise (one line).
4. Respond in the same language as the question text.
5. When Survey Context is provided, use the overall questionnaire flow to better interpret filter references and produce more accurate base descriptions.

## JSON Output Format
{
  "results": [
    {"question_number": "Q1", "base": "All Respondents"},
    {"question_number": "Q3", "base": "Q2(Brand Awareness) 'Aware' or 'Used' respondents"}
  ]
}"""


def generate_bases(questions: List[SurveyQuestion], language: str = "ko",
                   progress_callback=None, survey_context: str = "") -> dict:
    """Base 설명 생성 — 필터 있으면 LLM 보강, 없으면 'All Respondents'.

    Returns:
        dict: {question_number: base_string}
    """
    result = {}
    needs_llm = []

    # 문항별 컨텍스트 맵 (필터 해석에 사용)
    qn_context = {}
    for q in questions:
        qn_context[q.question_number] = {
            "text": q.question_text,
            "options": q.answer_options_compact(),
        }

    for q in questions:
        if q.filter_condition and q.filter_condition.strip():
            needs_llm.append(q)
        else:
            result[q.question_number] = "All Respondents"

    if not needs_llm:
        return result

    # 배치 분할
    batches = [needs_llm[i:i + BATCH_SIZE] for i in range(0, len(needs_llm), BATCH_SIZE)]
    total_batches = len(batches)

    for batch_idx, batch in enumerate(batches):
        if progress_callback:
            progress_callback("base_batch_start", {
                "batch_index": batch_idx, "total_batches": total_batches,
                "question_count": len(batch),
            })

        # 프롬프트에 필터 문항의 컨텍스트 포함
        lines = []
        if survey_context:
            lines.append(survey_context)
            lines.append("")
        referenced_qns = set()
        for q in batch:
            # 필터에서 참조하는 문항번호 추출
            refs = re.findall(r'([A-Z]+\d+[a-zA-Z]*)', q.filter_condition or "")
            for ref in refs:
                if ref in qn_context:
                    referenced_qns.add(ref)

        # 참조 문항 컨텍스트
        if referenced_qns:
            lines.append("## Referenced Questions Context")
            for ref_qn in sorted(referenced_qns):
                ctx = qn_context[ref_qn]
                lines.append(f"[{ref_qn}] {ctx['text']}")
                if ctx['options']:
                    lines.append(f"  Options: {ctx['options']}")
            lines.append("")

        lines.append("## Questions to process")
        for q in batch:
            lines.append(f"[{q.question_number}] Filter: {q.filter_condition}")
            lines.append(f"  Text: {q.question_text}")

        user_prompt = "\n".join(lines)

        try:
            raw = call_llm_json(_BASE_SYSTEM_PROMPT, user_prompt, MODEL_BASE_GENERATOR)
            for r in raw.get("results", []):
                qn = str(r.get("question_number", "")).strip()
                base = str(r.get("base", "")).strip()
                if qn and base:
                    result[qn] = base
        except Exception as e:
            logger.error(f"Base generation batch {batch_idx} failed: {e}")

        # LLM이 누락한 문항은 필터 텍스트 그대로 사용
        for q in batch:
            if q.question_number not in result:
                result[q.question_number] = q.filter_condition or "All Respondents"

        if progress_callback:
            progress_callback("base_batch_done", {
                "batch_index": batch_idx, "total_batches": total_batches,
            })

    return result


# ======================================================================
# Phase 2: Net/Recode
# ======================================================================

_NET_SYSTEM_PROMPT = """You are a DP specialist for marketing research cross-tabulation.
Your task is to suggest meaningful Net/Recode groupings for survey questions.

## Rules
1. For SA/MA questions with answer options, suggest logical groupings:
   - Income ranges → Low/Mid/High
   - Age groups → Young/Middle/Senior
   - Likert scales → Top2/Bot2
   - Frequency → Frequent/Infrequent
2. Only suggest nets when grouping adds analytical value. If no meaningful grouping exists, return empty string.
3. Format: "NetName(codes): NetName(codes)" e.g. "Top2(4+5) / Bot2(1+2)"
4. Keep it concise.
5. When Survey Context is provided, consider the question's role in the overall study (e.g., screening, demographics, evaluation) to decide whether a net adds analytical value and what groupings are most meaningful.

## JSON Output Format
{
  "results": [
    {"question_number": "S2", "net_recode": "Young(1+2) / Middle(3+4) / Senior(5+6)"},
    {"question_number": "Q5", "net_recode": ""}
  ]
}"""


def _generate_scale_net(summary_types: list,
                        answer_options: list | None = None) -> str:
    """SCALE 문항의 Net/Recode를 알고리즘으로 생성.

    answer_options에서 실제 스케일 범위를 감지하여 정확한 Top2/Bot2 코드 생성.
    SummaryType에 이미 Top2/Bot2/Mean 정보가 있으면 해당 정보 참고.
    """
    # 실제 숫자 코드 추출 → 스케일 범위 결정
    numeric_codes = []
    if answer_options:
        for opt in answer_options:
            code = getattr(opt, "code", str(opt)) if not isinstance(opt, str) else opt
            try:
                numeric_codes.append(int(code))
            except (ValueError, TypeError):
                pass

    if numeric_codes:
        numeric_codes.sort()
        lo1, lo2 = numeric_codes[0], numeric_codes[1] if len(numeric_codes) > 1 else numeric_codes[0]
        hi1, hi2 = numeric_codes[-1], numeric_codes[-2] if len(numeric_codes) > 1 else numeric_codes[-1]
        top2_str = f"Top2({hi2}+{hi1})"
        bot2_str = f"Bot2({lo1}+{lo2})"
    else:
        # 코드 정보 없으면 일반적인 5점 척도 가정
        top2_str = "Top2(4+5)"
        bot2_str = "Bot2(1+2)"

    has_top2 = any("top2" in st.lower() for st in summary_types if st)
    has_bot2 = any("bot2" in st.lower() or "bottom2" in st.lower() for st in summary_types if st)
    has_mean = any("mean" in st.lower() for st in summary_types if st)

    parts = []
    if has_top2:
        parts.append(top2_str)
    if has_bot2:
        parts.append(bot2_str)
    if has_mean:
        parts.append("Mean")

    if parts:
        return " / ".join(parts)

    # 기본: Top2 / Bot2 / Mean
    return f"{top2_str} / {bot2_str} / Mean"


def generate_net_recodes(questions: List[SurveyQuestion], language: str = "ko",
                         progress_callback=None, survey_context: str = "") -> dict:
    """Net/Recode 제안 — SCALE은 알고리즘, SA/MA는 LLM.

    Returns:
        dict: {question_number: net_recode_string}
    """
    result = {}
    needs_llm = []

    # 문항별 SummaryType 수집
    qn_summary_types = {}
    for q in questions:
        if q.question_number not in qn_summary_types:
            qn_summary_types[q.question_number] = []
        if q.summary_type:
            qn_summary_types[q.question_number].append(q.summary_type)

    # 이미 처리된 문항번호 추적 (중복 방지)
    seen_qn = set()
    for q in questions:
        if q.question_number in seen_qn:
            continue
        seen_qn.add(q.question_number)

        qtype = (q.question_type or "").strip().upper()
        sts = qn_summary_types.get(q.question_number, [])

        # SCALE/매트릭스 → 알고리즘
        if "SCALE" in qtype or re.match(r'\d+\s*PT\s*X\s*\d+', qtype):
            result[q.question_number] = _generate_scale_net(sts, q.answer_options)
        elif q.answer_options and len(q.answer_options) >= 4:
            # SA/MA with enough options → LLM 제안 대상
            needs_llm.append(q)
        else:
            result[q.question_number] = ""

    if not needs_llm:
        return result

    batches = [needs_llm[i:i + BATCH_SIZE] for i in range(0, len(needs_llm), BATCH_SIZE)]
    total_batches = len(batches)

    for batch_idx, batch in enumerate(batches):
        if progress_callback:
            progress_callback("net_batch_start", {
                "batch_index": batch_idx, "total_batches": total_batches,
                "question_count": len(batch),
            })

        lines = []
        if survey_context:
            lines.append(survey_context)
            lines.append("")
        lines.append("Generate Net/Recode suggestions for these questions:\n")
        for q in batch:
            lines.append(f"[{q.question_number}] {q.question_text}")
            lines.append(f"  Type: {q.question_type or 'SA'}")
            lines.append(f"  Options: {q.answer_options_compact()}")
            lines.append("")

        user_prompt = "\n".join(lines)

        try:
            raw = call_llm_json(_NET_SYSTEM_PROMPT, user_prompt, MODEL_NET_GENERATOR)
            for r in raw.get("results", []):
                qn = str(r.get("question_number", "")).strip()
                net = str(r.get("net_recode", "")).strip()
                if qn:
                    result[qn] = net
        except Exception as e:
            logger.error(f"Net/Recode batch {batch_idx} failed: {e}")

        for q in batch:
            if q.question_number not in result:
                result[q.question_number] = ""

        if progress_callback:
            progress_callback("net_batch_done", {
                "batch_index": batch_idx, "total_batches": total_batches,
            })

    return result


# ======================================================================
# Phase 3: Banner Management — 3-Step CoT Pipeline
# ======================================================================

# ── Step 1: Analysis Plan ────────────────────────────────────────────

_ANALYSIS_PLAN_SYSTEM_PROMPT = """You are a research director at a top-tier marketing research firm (Ipsos/Kantar/Nielsen level) designing a cross-tabulation analysis framework.

## Your Mindset
You are NOT listing questions that could become banners. You are designing the **analytical lenses** through which the client's brand team will read every table in the report. Each lens must answer a specific strategic question. The output of this step directly determines the quality of the entire Table Guide.

## Banner Convention in Marketing Research
In professional cross-tabulation, banners are organized as multi-dimensional column headers grouped into **3-5 banner groups** (Banner A, B, C, D…). Each group covers a distinct analytical theme derived from the study's objectives and questionnaire structure.

**You must determine the banner groups yourself** by analyzing the research objectives, the industry/category, and the specific questions in the questionnaire. Do NOT use generic predefined categories. Instead, think as both a **marketing research expert** and a **domain specialist for the industry being studied** (e.g., automotive, FMCG, healthcare, finance, technology).

For example:
- An automotive brand tracking study might produce: "Demographics", "Vehicle Ownership Profile", "Brand Relationship & Funnel", "Purchase Journey & Intent"
- A healthcare patient satisfaction study might produce: "Patient Demographics", "Treatment & Care Experience", "Provider Trust & Loyalty", "Composite Health Segments"
- An FMCG U&A study might produce: "Shopper Profile", "Category Usage & Habits", "Brand Repertoire & Switching", "Attitudinal Segments"

**Key principle**: Demographics alone are NEVER sufficient. Any question — whether it measures attitudes, behaviors, ownership, satisfaction, or intent — can be a banner dimension if it helps explain variance in the study's core metrics. Think about which respondent characteristics would produce the most meaningful differences when cross-tabulated against other questions.

## Your CoT Process (follow these steps exactly)

### Step 1: Study Comprehension
- What is the study type? (Brand Tracking / U&A / Satisfaction / Ad Test / Concept Test / Segmentation / etc.)
- Who is the client? What brand/category?
- What are the 3-5 core research questions this study must answer?
- What decisions will the client make based on this data?

### Step 2: Analytical Perspective Design
Based on Step 1, determine what analytical lenses (categories) are needed.
Each perspective must answer a specific strategic question for the client.

**Derive your categories from the questionnaire and research objectives.** Each category should represent a distinct analytical lens that a research director would present to the client as a reason to look at the data differently.

Guidelines:
- Minimum 4, maximum 7 perspectives (categories)
- Each perspective must have 2-5 banner dimensions
- At least 30% of total dimensions must be composite (combining 2+ questions)
- Every perspective must pass the "So What?" test — if removed, would the client miss critical insight?
- **Demographic-only dimensions should be ≤ 30% of total dimensions** — the majority should come from questions measuring behavior, attitudes, experience, or strategic segments
- Categories must reflect the study's **industry context and research objectives**, not generic textbook labels

### Step 3: Dimension Specification
For each perspective, define concrete dimensions with:
- Candidate questions (exact question numbers)
- Grouping strategy (exact codes — be specific, not vague)
- Whether it's composite or simple
- Each dimension should have **3-6 meaningful breakpoints** (not just binary splits)

## How to Think About Category Design
Ask yourself these questions for each potential category:
1. **What respondent characteristic** does this capture? (who they are, what they do, what they think, how they relate to the brand)
2. **What strategic question** does cross-tabulating by this category answer?
3. **Would the client pay for a separate analysis** using this lens? If yes, it's a valid category.

Consider ALL question types as potential banner sources:
- **Profiling questions** (demographics, firmographics) — baseline segmentation
- **Behavioral questions** (usage, purchase, channel, frequency) — what respondents DO
- **Attitudinal questions** (satisfaction, intent, preference, NPS, agreement scales) — what respondents THINK
- **Experiential questions** (touchpoints, journey stages, service interactions) — what respondents EXPERIENCE
- **Composite segments** (combining 2-4 questions) — strategic groups that don't exist in any single question

The specific categories you create should be **tailored to the industry and study objectives**. A luxury brand tracker needs different lenses than a healthcare patient survey or a B2B SaaS satisfaction study. Use your knowledge of the specific industry to identify the most insightful analytical dimensions.

## Rules for Dimension Design
1. **At least 30% of dimensions MUST be composite** (is_composite: true) — combining 2+ questions
2. **Grouping strategies must be specific**: exact code groupings, not vague descriptions
3. **Every dimension must answer a strategic question** — "So what?" test
4. **EXCLUDE**: OE questions, screening termination questions, country/market filters
5. **Think in terms of SEGMENTS, not questions**: "High-Intent EV Switchers" not "Q15 responses"
6. **Each dimension should produce 3-6 breakpoints**, not binary splits. Binary splits (Yes/No) are acceptable only for naturally binary questions (e.g., Gender)
7. **Attitudinal/behavioral questions are prime banner candidates** when their responses help explain variance in other study questions

## JSON Output Format
{
  "cot_reasoning": {
    "study_type": "Identified study type and rationale",
    "client_brand": "Identified or inferred client brand",
    "core_research_questions": ["Q1", "Q2", "Q3"],
    "client_decisions": "What decisions will the client make from this data?",
    "perspective_rationale": "Why these specific perspectives were chosen over alternatives"
  },
  "analysis_strategy": "3-4 sentences: What is the core analytical story? What strategic questions will this banner framework answer?",
  "categories": [
    {
      "category_name": "Descriptive name for this analytical perspective",
      "business_rationale": "What strategic question does this perspective answer?",
      "priority": "critical | important | supplementary",
      "banner_dimensions": [
        {
          "dimension_name": "Human-readable segment name (how it appears in report)",
          "analytical_question": "What strategic question does this dimension reveal?",
          "candidate_questions": ["SQ1", "SQ2"],
          "grouping_strategy": "Specific grouping: Code 1,2 = Group A, Code 3,4 = Group B",
          "is_composite": false
        }
      ]
    }
  ]
}"""


def _create_analysis_plan(questions: List[SurveyQuestion],
                          language: str,
                          survey_context: str,
                          intelligence: dict | None) -> dict | None:
    """Step 1: 분석 계획 생성 — 어떤 교차분석이 필요한지 전략적으로 분석.

    Args:
        questions: 전체 문항 리스트
        language: 설문지 언어
        survey_context: Study Brief + Intelligence + Question Flow
        intelligence: Survey Intelligence 결과 dict

    Returns:
        분석 계획 dict 또는 None (실패 시)
    """
    lines = []
    if survey_context:
        lines.append(survey_context)
        lines.append("")

    # Intelligence가 있으면 배너 추천 정보 포함
    if intelligence:
        recs = intelligence.get("banner_recommendations", [])
        if recs:
            lines.append("## Prior Banner Recommendations (from intelligence)")
            for rec in recs:
                lines.append(f"- {rec.get('name', '')}: {rec.get('rationale', '')}")
                pts = rec.get("points", [])
                if pts:
                    lines.append(f"  Variables: {', '.join(pts)}")
            lines.append("")

    # Step 1은 전략적 분석 단계 — 상세 보기 코드 불필요 (Step 2에서 사용)
    # 옵션 제외로 프롬프트 크기 ~50% 감소 → GPT-5 빈 응답 문제 방지
    lines.append(f"## Complete Question List ({len(questions)} questions, language: {language})")
    lines.append("")
    lines.append(_format_questions_compact(questions, include_options=False))

    user_prompt = "\n".join(lines)

    try:
        result = _call_llm_json_with_fallback(
            _ANALYSIS_PLAN_SYSTEM_PROMPT, user_prompt,
            MODEL_INTELLIGENCE, temperature=0.3, max_tokens=16384,
        )
        result.setdefault("analysis_strategy", "")
        result.setdefault("categories", [])
        result.setdefault("cot_reasoning", {})
        # Log CoT reasoning for debugging
        cot = result.get("cot_reasoning", {})
        if cot:
            logger.info(f"Analysis plan CoT: study_type={cot.get('study_type', '')}, "
                        f"client={cot.get('client_brand', '')}, "
                        f"perspectives={len(result.get('categories', []))}")
        # Flatten banner_dimensions from categories for downstream compatibility
        all_dims = []
        for cat in result.get("categories", []):
            cat_name = cat.get("category_name", "")
            cat_priority = _PRIORITY_MAP.get(cat.get("priority", ""), "high")
            for dim in cat.get("banner_dimensions", []):
                dim["category"] = cat_name
                # Map is_composite to variable_type for compatibility
                if dim.get("is_composite"):
                    dim.setdefault("variable_type", "composite")
                dim.setdefault("priority", cat_priority)
                all_dims.append(dim)
        result["banner_dimensions"] = all_dims
        # Extract composite opportunities from is_composite dims
        composites = []
        for dim in all_dims:
            if dim.get("is_composite"):
                composites.append({
                    "name": dim.get("dimension_name", ""),
                    "component_questions": dim.get("candidate_questions", []),
                    "logic": dim.get("grouping_strategy", ""),
                    "analytical_value": dim.get("analytical_question", ""),
                })
        result["composite_opportunities"] = composites
        total_dims = len(all_dims)
        cat_count = len(result["categories"])
        logger.info(f"Analysis plan: {cat_count} categories, {total_dims} dimensions, "
                    f"{len(composites)} composite opportunities")
        return result
    except Exception as e:
        logger.error(f"Analysis plan creation failed: {e}")
        return None


# ======================================================================
# Expert Consensus Pipeline — Research Plan + Expert Panel + Synthesis
# ======================================================================

# ── Step 0.5: Research Plan ─────────────────────────────────────────

_RESEARCH_PLAN_SYSTEM_PROMPT = """You are a Senior Research Planner creating a structured research brief from a survey questionnaire.

## Your Task
Analyze the questionnaire to produce a structured research brief that will guide a panel of MR experts in designing cross-tabulation banners. This brief must clearly articulate the study's purpose, research objectives, and the analytical dimensions needed to address each objective.

## Process
1. **Study Comprehension**: Understand the study type, client, category, and target audience from the questionnaire flow.
2. **Objective Extraction**: Identify 3-7 research objectives, each linked to specific questions. Classify each as "primary" (must-have analysis) or "secondary" (nice-to-have).
3. **Dimension Mapping**: For each objective, identify the analytical dimensions (cross-tabulation variables) needed. Each dimension should be tagged as "simple" (single question) or "composite" (combining 2+ questions).

## Rules
- Every primary objective MUST have at least one composite dimension proposed.
- Dimensions must reference exact question numbers from the questionnaire.
- The brief should be concise but complete — experts will rely on it exclusively.
- Consider the study's industry context when identifying objectives and dimensions.

## JSON Output Format
{
  "study_brief": "2-3 sentence summary for expert consumption",
  "research_objectives": [
    {
      "id": "RO1",
      "description": "Research objective description",
      "priority": "primary or secondary",
      "related_questions": ["Q1", "Q2", "Q3"],
      "analytical_need": "What cross-analysis is needed to address this objective"
    }
  ],
  "objective_dimension_map": [
    {
      "objective_id": "RO1",
      "dimensions": [
        {
          "name": "Dimension name",
          "candidate_questions": ["Q1"],
          "type": "simple or composite",
          "rationale": "Why this dimension is needed for the objective"
        }
      ]
    }
  ]
}"""


def _create_research_plan(
    questions: List[SurveyQuestion],
    language: str,
    survey_context: str,
    intelligence: dict | None,
) -> dict | None:
    """Step 0.5: 구조화된 연구 기획서 생성.

    설문지에서 연구 목적, 분석 차원을 추출하여 전문가 패널의 입력으로 사용.

    Returns:
        연구 기획서 dict 또는 None (실패 시)
    """
    lines = []
    study_params = _build_structured_study_params(survey_context, intelligence)
    if study_params:
        lines.append(study_params)
    if survey_context:
        lines.append(survey_context)
        lines.append("")

    if intelligence:
        objectives = intelligence.get("research_objectives", [])
        if objectives:
            lines.append("## Prior Intelligence — Research Objectives")
            for obj in objectives:
                lines.append(f"- {obj}")
            lines.append("")
        segments = intelligence.get("key_segments", [])
        if segments:
            lines.append("## Prior Intelligence — Key Segments")
            for seg in segments:
                lines.append(f"- {seg.get('name', '')} ({seg.get('question', '')}, {seg.get('type', '')})")
            lines.append("")

    lines.append(f"## Complete Question List ({len(questions)} questions, language: {language})")
    lines.append("")
    lines.append(_format_questions_compact(questions, include_options=False))

    user_prompt = "\n".join(lines)

    try:
        result = _call_llm_json_with_fallback(
            _RESEARCH_PLAN_SYSTEM_PROMPT, user_prompt,
            MODEL_INTELLIGENCE, temperature=0.2, max_tokens=8192,
        )
        result.setdefault("study_brief", "")
        result.setdefault("research_objectives", [])
        result.setdefault("objective_dimension_map", [])

        # 품질 게이트: objectives >= 3, dimensions >= 6, primary마다 composite >= 1
        objectives = result.get("research_objectives", [])
        dim_map = result.get("objective_dimension_map", [])
        all_dims = [d for m in dim_map for d in m.get("dimensions", [])]

        if len(objectives) < 3:
            logger.warning(f"Research plan has only {len(objectives)} objectives (min 3)")
        if len(all_dims) < 6:
            logger.warning(f"Research plan has only {len(all_dims)} dimensions (min 6)")

        logger.info(f"Research plan: {len(objectives)} objectives, {len(all_dims)} dimensions")
        return result
    except Exception as e:
        logger.error(f"Research plan creation failed: {e}")
        return None


# ── Step 1 (Expert Panel): 3 Expert Personas ────────────────────────

_EXPERT_COMMON_PREAMBLE = """You are part of a 3-person expert panel analyzing a survey questionnaire to design cross-tabulation banners.

## Your Input
You receive:
1. A structured Research Plan with research objectives and dimension mapping
2. Survey context and question list
3. Your specific expert role and evaluation criteria

## Your Task
Independently analyze the Research Plan and questionnaire, then propose banner categories and dimensions from YOUR expert perspective. Other experts will provide different perspectives; a synthesis step will merge all views.

## Output Schema (MUST follow exactly)
{
  "expert_name": "your_role_id",
  "categories": [
    {
      "category_name": "Category name",
      "business_rationale": "Why this category matters",
      "priority": "critical|important|supplementary",
      "banner_dimensions": [
        {
          "dimension_name": "Human-readable dimension name",
          "candidate_questions": ["Q1", "Q2"],
          "grouping_strategy": "Code 1,2 = Group A; Code 3,4 = Group B",
          "is_composite": false,
          "analytical_question": "What strategic question does this answer?",
          "objective_ids": ["RO1"]
        }
      ]
    }
  ],
  "priority_rankings": [
    {"dimension_name": "...", "score": 8, "rationale": "Why this score"}
  ],
  "concerns": ["Issues or gaps identified"],
  "composite_proposals": [
    {
      "name": "Composite segment name",
      "questions": ["Q1", "Q2"],
      "logic": "Combination logic description",
      "analytical_value": "Why this composite is valuable",
      "quality_scores": {
        "feasibility": 8,
        "uniqueness": 7,
        "business_impact": 9
      }
    }
  ]
}

## Composite Quality Scoring
For EVERY composite proposal, rate these 3 dimensions (1-10):
- **Feasibility** (DP perspective): Can the combination actually be computed? Are filter conditions compatible?
- **Uniqueness** (Research perspective): Does this reveal insights NOT visible in any single question?
- **Business Impact** (Client perspective): Would this segment change the client's strategy?
Only propose composites with average score >= 6."""


_EXPERT_RESEARCH_DIRECTOR_SYSTEM = _EXPERT_COMMON_PREAMBLE + """

## Your Role: Research Director
You are a senior Research Director at a top-tier MR firm (Ipsos/Kantar level).

## Your Evaluation Criteria
1. **Research Objective Coverage**: Every primary objective in the Research Plan must be addressable through at least one banner dimension.
2. **Analysis Story Arc**: Categories should tell a coherent analytical story (e.g., "Who are they?" → "What do they do?" → "What do they think?" → "What will they do?").
3. **Category Composition**: 4-6 categories, each with a clear strategic purpose. Category names should reflect the study's industry and objectives.
4. **Composite Segments**: Propose at least 3 composite dimensions that combine questions to create strategic segments not visible in any single question.

## Your Priorities
- Favor categories that directly map to research objectives
- Ensure every primary objective has at least one "critical" priority dimension
- Think in terms of the client's decision-making: what segments would change their strategy?
- Name categories strategically (e.g., "Brand Relationship Journey" not "Brand Questions")"""


_EXPERT_DP_MANAGER_SYSTEM = _EXPERT_COMMON_PREAMBLE + """

## Your Role: DP Manager
You are a senior Data Processing Manager responsible for technical feasibility of cross-tabulation banners.

## Your Evaluation Criteria
1. **Code Existence**: Every grouping strategy must reference actual answer codes that exist in the questionnaire. Verify code ranges.
2. **Mutual Exclusivity**: Banner values within a dimension must be mutually exclusive — no respondent should fall into two values.
3. **Composite Feasibility**: For composite banners (combining 2+ questions), verify that:
   - All referenced questions exist
   - Filter conditions are compatible (respondents who answer Q1 also answer Q2)
   - The AND combination produces meaningful, non-empty segments
4. **Practical Constraints**: Flag dimensions where expected cell sizes may be too small for analysis.

## Your Priorities
- Be precise with grouping strategies: use exact codes (e.g., "Code 1,2 = Group A")
- Flag any dimension where codes may not exist or overlap
- For composites, specify the exact AND conditions needed
- Reject or modify proposals that are technically infeasible
- You have access to full answer options — USE them to verify code references"""


_EXPERT_CLIENT_INSIGHTS_SYSTEM = _EXPERT_COMMON_PREAMBLE + """

## Your Role: Client Insights Manager
You are a senior Insights Manager who presents research findings to client brand teams.

## Your Evaluation Criteria
1. **Decision Usefulness**: Each banner dimension must help the client make a specific business decision. Ask "So what?" for every dimension.
2. **Competitive Benchmarking**: Include dimensions that enable comparison between the client's brand and competitors.
3. **Presentation Value**: Dimensions should produce interesting, story-worthy cross-tabulations that a VP of Marketing would find compelling.
4. **Business Priority Scoring**: Score each dimension 1-10 on business value. Reserve 9-10 for truly differentiating insights.

## Your Priorities
- Prioritize dimensions that reveal competitive advantages/disadvantages
- Include at least one "headline-worthy" composite segment (e.g., "Loyal Advocates" vs "At-Risk Users")
- Ensure demographic cuts are balanced with behavioral/attitudinal dimensions
- Think about what would make a compelling chart or headline in the final presentation"""


def _expert_research_director(
    research_plan: dict,
    questions: List[SurveyQuestion],
    language: str,
    survey_context: str,
    intelligence: dict | None = None,
) -> dict:
    """Research Director 전문가 분석."""

    lines = []
    study_params = _build_structured_study_params(survey_context, intelligence)
    if study_params:
        lines.append(study_params)
    domain_guide = _get_domain_guidance(intelligence)
    if domain_guide:
        lines.append(domain_guide)
    if survey_context:
        lines.append(survey_context)
        lines.append("")
    lines.append("## Research Plan")
    lines.append(_json.dumps(research_plan, ensure_ascii=False, indent=2))
    lines.append("")
    lines.append(f"## Question List ({len(questions)} questions, language: {language})")
    lines.append(_format_questions_compact(questions, include_options=False))

    result = _call_llm_json_with_fallback(
        _EXPERT_RESEARCH_DIRECTOR_SYSTEM, "\n".join(lines),
        MODEL_INTELLIGENCE, temperature=0.4, max_tokens=12288,
    )
    result.setdefault("expert_name", "research_director")
    return result


def _expert_dp_manager(
    research_plan: dict,
    questions: List[SurveyQuestion],
    language: str,
    survey_context: str,
    intelligence: dict | None = None,
) -> dict:
    """DP Manager 전문가 분석 — full answer options 포함."""

    lines = []
    study_params = _build_structured_study_params(survey_context, intelligence)
    if study_params:
        lines.append(study_params)
    if survey_context:
        lines.append(survey_context)
        lines.append("")
    lines.append("## Research Plan")
    lines.append(_json.dumps(research_plan, ensure_ascii=False, indent=2))
    lines.append("")
    # DP Manager는 코드 검증 필요 → full options 포함
    lines.append(f"## Question List with Answer Options ({len(questions)} questions, language: {language})")
    lines.append(_format_questions_compact(questions, include_options=True, max_option_len=300))

    result = _call_llm_json_with_fallback(
        _EXPERT_DP_MANAGER_SYSTEM, "\n".join(lines),
        MODEL_INTELLIGENCE, temperature=0.25, max_tokens=12288,
    )
    result.setdefault("expert_name", "dp_manager")
    return result


def _expert_client_insights(
    research_plan: dict,
    questions: List[SurveyQuestion],
    language: str,
    survey_context: str,
    intelligence: dict | None = None,
) -> dict:
    """Client Insights Manager 전문가 분석."""

    lines = []
    study_params = _build_structured_study_params(survey_context, intelligence)
    if study_params:
        lines.append(study_params)
    domain_guide = _get_domain_guidance(intelligence)
    if domain_guide:
        lines.append(domain_guide)
    if survey_context:
        lines.append(survey_context)
        lines.append("")
    lines.append("## Research Plan")
    lines.append(_json.dumps(research_plan, ensure_ascii=False, indent=2))
    lines.append("")
    lines.append(f"## Question List ({len(questions)} questions, language: {language})")
    lines.append(_format_questions_compact(questions, include_options=False))

    result = _call_llm_json_with_fallback(
        _EXPERT_CLIENT_INSIGHTS_SYSTEM, "\n".join(lines),
        MODEL_INTELLIGENCE, temperature=0.3, max_tokens=12288,
    )
    result.setdefault("expert_name", "client_insights")
    return result


def _run_expert_panel(
    research_plan: dict,
    questions: List[SurveyQuestion],
    language: str,
    survey_context: str,
    progress_callback: Callable | None = None,
    intelligence: dict | None = None,
) -> List[dict]:
    """3명의 전문가 패널 병렬 실행.

    ThreadPoolExecutor로 3개 LLM 호출을 동시 실행.

    Returns:
        전문가 출력 리스트 (최대 3개, 실패 시 해당 전문가 제외)
    """
    expert_fns = [
        ("Research Director", _expert_research_director),
        ("DP Manager", _expert_dp_manager),
        ("Client Insights", _expert_client_insights),
    ]
    expert_outputs: List[dict] = []

    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = {}
        for name, fn in expert_fns:
            future = executor.submit(
                fn, research_plan, questions, language, survey_context,
                intelligence=intelligence,
            )
            futures[future] = name

        done_count = 0
        for future in as_completed(futures):
            name = futures[future]
            done_count += 1
            try:
                result = future.result()
                expert_outputs.append(result)
                if progress_callback:
                    progress_callback("expert_done", {
                        "name": name, "index": done_count, "total": 3,
                    })
                logger.info(f"Expert '{name}' completed: "
                            f"{len(result.get('categories', []))} categories, "
                            f"{len(result.get('composite_proposals', []))} composites")
            except Exception as e:
                logger.error(f"Expert '{name}' failed: {e}")
                if progress_callback:
                    progress_callback("expert_done", {
                        "name": f"{name} (failed)", "index": done_count, "total": 3,
                    })

    return expert_outputs


# ── Step 1.5: Synthesis ─────────────────────────────────────────────

_SYNTHESIS_SYSTEM_PROMPT = """You are a senior Research Director synthesizing independent analyses from 3 MR experts into a unified banner analysis plan.

## Expert Roles
- **Research Director**: Strategic analysis framework, category composition, objective coverage
- **DP Manager**: Technical feasibility, code verification, composite realizability
- **Client Insights Manager**: Business value scoring, competitive benchmarking, presentation impact

## Mediation Rules (follow strictly)
1. **Majority Rule**: A dimension proposed by 2/3 experts → automatically included.
2. **Minority Override**: A dimension proposed by only 1 expert with priority >= "important" → included.
3. **Grouping Conflicts**: Use DP Manager's grouping strategy (technical accuracy priority).
4. **Category Naming**: Use Research Director's category names (strategic framing priority).
5. **Priority Rankings**: Use Client Insights Manager's scores (business value priority).
6. **Composite Banners**: Only include composites that DP Manager has technically validated.
7. **Concerns**: Address all concerns raised by any expert; document resolution.

## Quality Requirements
- Minimum 4 categories
- Minimum 10 total dimensions across all categories
- At least **35%** composite dimensions (combining 2+ questions)
- **At least 4** distinct composite dimensions
- **At least 1** "deep" composite combining 3+ questions
- Every primary research objective must have at least 1 dimension
- Demographics should be <= 30% of total dimensions

## Output Format
Produce an analysis plan in the EXACT format below. This will be passed directly to the banner design step.

{
  "analysis_strategy": "3-4 sentence summary of the unified analysis approach",
  "categories": [
    {
      "category_name": "Category name (from Research Director)",
      "business_rationale": "Strategic purpose",
      "priority": "critical|important|supplementary",
      "banner_dimensions": [
        {
          "dimension_name": "Human-readable name",
          "analytical_question": "What does this dimension reveal?",
          "candidate_questions": ["Q1", "Q2"],
          "grouping_strategy": "From DP Manager: exact code groupings",
          "is_composite": false
        }
      ]
    }
  ],
  "composite_opportunities": [
    {
      "name": "Composite name",
      "component_questions": ["Q1", "Q2"],
      "logic": "Combination logic from DP Manager",
      "analytical_value": "Business value from Client Insights"
    }
  ],
  "consensus_notes": "Summary of how expert disagreements were resolved",
  "agreement_score": 0.85,
  "expert_contributions": {
    "research_director": ["Key contributions from RD"],
    "dp_manager": ["Key contributions from DP"],
    "client_insights": ["Key contributions from CI"]
  }
}

The agreement_score should reflect how much the experts agreed (0.0 = complete disagreement, 1.0 = perfect agreement). Consider: overlap in proposed dimensions, consistency in priority ratings, and alignment on composite proposals."""


def _synthesize_expert_consensus(
    expert_outputs: List[dict],
    research_plan: dict,
    questions: List[SurveyQuestion],
    language: str,
    survey_context: str = "",
    intelligence: dict | None = None,
) -> dict | None:
    """3명의 전문가 출력을 중재 규칙으로 통합하여 합의 분석 계획 생성.

    Returns:
        합의 분석 계획 dict (기존 _create_analysis_plan 출력 호환) 또는 None
    """

    if not expert_outputs:
        logger.warning("No expert outputs to synthesize")
        return None

    lines = []
    study_params = _build_structured_study_params(survey_context, intelligence)
    if study_params:
        lines.append(study_params)
    domain_guide = _get_domain_guidance(intelligence)
    if domain_guide:
        lines.append(domain_guide)
    lines.append(f"## Research Plan (language: {language})")
    lines.append(_json.dumps(research_plan, ensure_ascii=False, indent=2))
    lines.append("")

    for i, expert in enumerate(expert_outputs):
        name = expert.get("expert_name", f"expert_{i}")
        lines.append(f"## Expert {i+1}: {name}")
        lines.append(_json.dumps(expert, ensure_ascii=False, indent=2))
        lines.append("")

    user_prompt = "\n".join(lines)

    try:
        result = _call_llm_json_with_fallback(
            _SYNTHESIS_SYSTEM_PROMPT, user_prompt,
            MODEL_INTELLIGENCE, temperature=0.15, max_tokens=16384,
        )
        result.setdefault("analysis_strategy", "")
        result.setdefault("categories", [])
        result.setdefault("composite_opportunities", [])
        result.setdefault("consensus_notes", "")
        result.setdefault("agreement_score", 0.0)
        result.setdefault("expert_contributions", {})

        # 기존 파이프라인 호환: banner_dimensions 플랫 리스트 생성
        all_dims = []
        for cat in result.get("categories", []):
            cat_name = cat.get("category_name", "")
            cat_priority = _PRIORITY_MAP.get(cat.get("priority", ""), "high")
            for dim in cat.get("banner_dimensions", []):
                dim["category"] = cat_name
                if dim.get("is_composite"):
                    dim.setdefault("variable_type", "composite")
                dim.setdefault("priority", cat_priority)
                all_dims.append(dim)
        result["banner_dimensions"] = all_dims

        # CoT reasoning placeholder (UI 호환)
        result.setdefault("cot_reasoning", {
            "study_type": research_plan.get("study_brief", ""),
            "client_brand": "",
            "core_research_questions": [
                obj.get("description", "")
                for obj in research_plan.get("research_objectives", [])
                if obj.get("priority") == "primary"
            ],
            "perspective_rationale": result.get("consensus_notes", ""),
        })

        # _research_plan 원본 참조 보존
        result["_research_plan"] = research_plan

        score = result.get("agreement_score", 0)
        logger.info(f"Expert consensus: {len(result['categories'])} categories, "
                    f"{len(all_dims)} dimensions, agreement={score:.2f}")
        return result
    except Exception as e:
        logger.error(f"Expert synthesis failed: {e}")
        # 폴백: Research Director 단독 출력 사용
        rd_output = next((e for e in expert_outputs
                          if e.get("expert_name") == "research_director"), None)
        if rd_output:
            logger.warning("Falling back to Research Director output only")
            # RD 출력을 analysis plan 형태로 변환
            # composite_proposals → composite_opportunities 필드명 변환
            raw_composites = rd_output.get("composite_proposals", [])
            composites = []
            for comp in raw_composites:
                composites.append({
                    "name": comp.get("name", ""),
                    "component_questions": comp.get("questions", comp.get("component_questions", [])),
                    "logic": comp.get("logic", ""),
                    "analytical_value": comp.get("analytical_value", ""),
                })
            fallback = {
                "analysis_strategy": "Single-expert fallback (Research Director only)",
                "categories": rd_output.get("categories", []),
                "composite_opportunities": composites,
                "consensus_notes": "Synthesis failed — using Research Director output only",
                "agreement_score": 0.0,
                "expert_contributions": {"research_director": ["Sole contributor (fallback)"]},
                "_research_plan": research_plan,
            }
            # banner_dimensions 플랫 리스트 생성
            all_dims = []
            for cat in fallback.get("categories", []):
                for dim in cat.get("banner_dimensions", []):
                    dim["category"] = cat.get("category_name", "")
                    if dim.get("is_composite"):
                        dim.setdefault("variable_type", "composite")
                    all_dims.append(dim)
            fallback["banner_dimensions"] = all_dims
            fallback["cot_reasoning"] = {
                "study_type": research_plan.get("study_brief", ""),
                "client_brand": "",
                "core_research_questions": [
                    obj.get("description", "")
                    for obj in research_plan.get("research_objectives", [])
                    if obj.get("priority") == "primary"
                ],
                "perspective_rationale": "Fallback: Research Director only",
            }
            return fallback
        return None


# ── Step 2: Banner Design ────────────────────────────────────────────

_BANNER_DESIGN_SYSTEM_PROMPT = """You are the head of DP at a top-tier research firm, implementing a cross-tabulation banner framework from an analysis plan.

## Your Task
Convert EVERY dimension from the analysis plan into production-ready banner specifications with exact conditions. The quality of the final report depends entirely on the precision of your banner definitions.

## Professional Banner Structure
In marketing research, banners are organized into thematic groups (Banner A, B, C, D…). The categories from the analysis plan define these groups. Each group should contain 2-6 banners covering a distinct analytical perspective.

**The categories were determined by the analysis plan** based on the study's research objectives and industry context. Your job is to implement EVERY dimension from the plan as a production-ready banner specification. Do NOT add generic categories — faithfully implement the plan's categories.

Banner dimensions can come from ANY question type — profiling, behavioral, attitudinal, experiential, or composite combinations. The key criterion is whether cross-tabulating by that dimension reveals actionable insights for the client's business decisions.

## Quality Standard
Each banner must pass the "VP of Insights" test: if a VP sees this banner in a cross-tab, they should immediately understand what segment they're looking at and why it matters for the brand strategy.

## Survey Flow Context
You may receive a list of "Other Questions" that are NOT candidates for banners. Use them ONLY to understand the survey flow, question ordering, and thematic structure. Do NOT create banners from non-candidate questions.

## Implementation Rules

### Condition Format
- Single code: `SQ1=1`
- Multiple codes (OR): `SQ1=1,2,3`
- Cross-question AND: `Q3=1&Q5=1,2`
- **NEVER** use negative conditions (`!=`, `NOT`, `≠`)
- Every value must have an explicit, executable condition
- Values within a banner must be **mutually exclusive**

### Simple Banners (single question)
- **Demographics**: Gender (Male/Female), Age (3-4 bands), Region (2-4 clusters)
- **Behavioral**: Group codes into NET segments — never list individual codes as banner values. E.g., Usage frequency: Heavy (daily+weekly) / Medium (monthly) / Light (less than monthly)
- **Attitudinal**: Satisfaction or intent scales → Top2 / Mid / Bot2 or Top2 / Bot3 depending on scale length
- **Ownership/Brand**: Group by strategic segments — Client Brand / Key Competitors / Other / Non-Owner

### Composite Banners (multiple questions combined with "&")
These are the MOST VALUABLE banners. They create strategic segments that don't exist in any single question.

**2-question composites:**
- **Brand Loyalty**: ownership × repurchase intent
  - "Loyal Owner" = `SQ10=1&SQ17=1` (owns client brand AND intends to repurchase)
  - "At-Risk Owner" = `SQ10=1&SQ17=2,3,4,5` (owns but considering switch)
  - "Conquest Target" = `SQ10=2,3,4&D4=1` (owns competitor but considers client)

**3-question composites (higher analytical value):**
- **Brand Funnel Stage**: awareness × consideration × ownership
  - "Loyal Advocate" = `A3=1&D4=1&SQ10=1` (aware + considers + owns client brand)
  - "Aware Non-Considerer" = `A3=1&D4=2,3,4,5&SQ10=2,3,4` (aware but doesn't consider)
  - "Unaware" = `A3=2,3,4,5` (not aware of client brand)
- **Engaged Satisfaction**: satisfaction × NPS × usage frequency
  - "Promoter Power User" = `Q15=1,2&Q16=9,10&Q8=1,2` (satisfied + promoter + heavy user)
  - "Detractor at Risk" = `Q15=4,5&Q16=0,1,2,3,4,5,6&Q8=1,2` (dissatisfied + detractor + still using)
  - "Passive Drifter" = `Q15=3&Q16=7,8&Q8=3,4,5` (neutral + passive + light user)

### Value Label Guidelines
- Labels must be short (2-4 words), descriptive, and meaningful
- BAD: "Code 1-3", "Q5=1,2", "Group A"
- GOOD: "Loyal Owner", "Active Researcher", "Price-Sensitive", "EV Considerer"
- Each banner should have **3-6 values**. Avoid 2-value binary splits unless the question is naturally binary (e.g., Gender). Scales should typically have 3 values (Top2/Mid/Bot2).

## Output Requirements
- **Minimum 12 banners** across all categories
- **At least 4 composite banners** (banner_type: "composite")
- **At least 2 composite banners must combine 3+ questions**
- **category field MUST match** the analysis plan's category_name exactly
- Every category must have at least 2 banners
- **Pure demographic dimensions should be ≤ 30% of total banners**
- **Average values per banner should be ≥ 3** across the full set
- Do NOT skip any dimension from the analysis plan

## JSON Output Format
{
  "banners": [
    {
      "category": "Brand Relationship",
      "name": "Brand Funnel Stage",
      "rationale": "Identifies where in the awareness→consideration→purchase funnel the client is losing prospects to competitors",
      "banner_type": "composite",
      "source_questions": ["A3", "D4", "SQ10"],
      "values": [
        {
          "label": "Loyal Advocate",
          "condition": "A3=1&D4=1&SQ10=1",
          "reasoning": "Aware + considers + owns client brand — core loyal base"
        },
        {
          "label": "At-Risk Owner",
          "condition": "SQ10=1&D4=2,3,4,5",
          "reasoning": "Owns client brand but not considering repurchase — retention priority"
        },
        {
          "label": "Conquest Target",
          "condition": "A3=1&D4=1&SQ10=2,3,4",
          "reasoning": "Aware + considers client but owns competitor — acquisition opportunity"
        },
        {
          "label": "Aware Non-Considerer",
          "condition": "A3=1&D4=2,3,4,5&SQ10=2,3,4",
          "reasoning": "Aware but not considering client — image/positioning gap"
        },
        {
          "label": "Unaware",
          "condition": "A3=2,3,4,5",
          "reasoning": "Not aware of client brand — awareness-building target"
        }
      ]
    }
  ]
}"""


def _design_banners_from_plan(analysis_plan: dict,
                               questions: List[SurveyQuestion],
                               language: str,
                               survey_context: str,
                               intelligence: dict | None = None) -> dict | None:
    """Step 2: 분석 계획 기반 배너 설계.

    Args:
        analysis_plan: Step 1의 분석 계획
        questions: 전체 문항 리스트
        language: 설문지 언어
        survey_context: Study Brief + Intelligence
        intelligence: Survey Intelligence 결과 dict

    Returns:
        배너 스펙 dict ({"banners": [...]}) 또는 None
    """
    # 분석 계획에서 후보 문항번호 수집
    candidate_qns = set()
    for dim in analysis_plan.get("banner_dimensions", []):
        for qn in dim.get("candidate_questions", []):
            candidate_qns.add(qn.strip())
    for comp in analysis_plan.get("composite_opportunities", []):
        for qn in comp.get("component_questions", []):
            candidate_qns.add(qn.strip())

    if not candidate_qns:
        logger.warning("Analysis plan has no candidate questions")
        return None

    # 후보 문항의 full 상세 정보 수집
    qn_map = {}
    for q in questions:
        if q.question_number not in qn_map:
            qn_map[q.question_number] = q
    candidate_qs = [qn_map[qn] for qn in candidate_qns if qn in qn_map]

    if not candidate_qs:
        logger.warning("No matching questions found for analysis plan candidates")
        return None

    # 프롬프트 구성

    lines = []
    study_params = _build_structured_study_params(survey_context, intelligence)
    if study_params:
        lines.append(study_params)
    domain_guide = _get_domain_guidance(intelligence)
    if domain_guide:
        lines.append(domain_guide)
    if survey_context:
        lines.append(survey_context)
        lines.append("")

    lines.append("## Analysis Plan")
    lines.append(f"Strategy: {analysis_plan.get('analysis_strategy', '') or analysis_plan.get('analysis_reasoning', '')}")
    lines.append("")

    # Category-based dimension listing
    categories = analysis_plan.get("categories", [])
    if categories:
        lines.append("### Categories & Dimensions")
        for cat in categories:
            cat_name = cat.get("category_name", "")
            lines.append(f"\n#### {cat_name}")
            lines.append(f"Rationale: {cat.get('business_rationale', '')}")
            for dim in cat.get("banner_dimensions", []):
                composite_tag = " [COMPOSITE]" if dim.get("is_composite") else ""
                lines.append(f"- **{dim.get('dimension_name', '')}**{composite_tag}")
                lines.append(f"  Question: {dim.get('analytical_question', '')}")
                lines.append(f"  Candidates: {', '.join(dim.get('candidate_questions', []))}")
                lines.append(f"  Grouping: {dim.get('grouping_strategy', '')}")
    else:
        # Fallback for old-style plans without categories
        lines.append("### Banner Dimensions")
        for dim in analysis_plan.get("banner_dimensions", []):
            lines.append(f"- **{dim.get('dimension_name', '')}** ({dim.get('variable_type', '')})")
            lines.append(f"  Question: {dim.get('analytical_question', '')}")
            lines.append(f"  Candidates: {', '.join(dim.get('candidate_questions', []))}")
            lines.append(f"  Grouping: {dim.get('grouping_strategy', '')}")

    composites = analysis_plan.get("composite_opportunities", [])
    if composites:
        lines.append("")
        lines.append("### Composite Opportunities")
        for comp in composites:
            lines.append(f"- **{comp.get('name', '')}**: {comp.get('logic', '')}")
            lines.append(f"  Value: {comp.get('analytical_value', '')}")

    # Domain-specific composite examples (Change 9)
    domain_composites = _get_domain_composite_examples(intelligence)
    if domain_composites:
        lines.append("")
        lines.append(domain_composites)

    lines.append("")
    lines.append(f"## Candidate Question Details ({len(candidate_qs)} questions)")
    lines.append("")
    lines.append(_format_questions_full(candidate_qs))

    # Non-candidate questions for flow context (Change 2)
    non_candidate_qs = [q for q in questions if q.question_number not in candidate_qns]
    if non_candidate_qs:
        lines.append("")
        lines.append(f"## Other Questions (for flow context only) ({len(non_candidate_qs)} questions)")
        lines.append("NOTE: You may reference these for understanding the survey flow, "
                      "but ONLY create banners from candidate questions above.")
        lines.append(_format_questions_compact(non_candidate_qs, include_options=False))

    user_prompt = "\n".join(lines)

    try:
        result = _call_llm_json_with_fallback(
            _BANNER_DESIGN_SYSTEM_PROMPT, user_prompt,
            MODEL_INTELLIGENCE, temperature=0.2, max_tokens=16384,
        )
        banners = result.get("banners", [])
        logger.info(f"Banner design: {len(banners)} banners generated")
        return result
    except Exception as e:
        logger.error(f"Banner design failed: {e}")
        return None


# ── Step 3: Validation ───────────────────────────────────────────────

_BANNER_VALIDATION_SYSTEM_PROMPT = """You are a DP quality checker validating cross-tabulation banner specifications.

## Your Task
Check each banner value's condition against the actual answer codes and fix any issues.

## Validation Rules
1. **CODE_EXISTS**: Every code in a condition must exist in that question's answer options. Remove or replace codes that don't exist.
2. **MUTUAL_EXCLUSIVITY**: Within the same banner, values should not overlap (same respondent shouldn't fall into two values). Flag but do not remove if overlap exists — just add a warning.
3. **NO_NEGATIVE**: Conditions must NOT use "!=", "NOT", "≠", or "<>". Convert to positive conditions using actual existing codes.
4. **COMPOSITE_CONSISTENCY**: For "&" conditions, all referenced questions must be present in the provided code map.
5. **COMPLETENESS**: If a major segment seems missing (e.g., banner has "Male" but no "Female"), add a warning.

## Input Format
You will receive:
- A list of banners with their conditions
- A code_map showing each question's valid codes

## JSON Output Format
Return the CORRECTED banners in the same format, with an additional "warnings" field:
{
  "banners": [
    {
      "category": "Demographics",
      "name": "...",
      "rationale": "...",
      "banner_type": "simple or composite",
      "source_questions": ["SQ1"],
      "values": [
        {
          "label": "...",
          "condition": "SQ1=1",
          "reasoning": "..."
        }
      ],
      "warnings": ["optional list of validation warnings"]
    }
  ],
  "validation_summary": "Brief summary of changes made"
}

## Rules
- If a condition references codes that don't exist, correct them to the closest valid codes or remove the value
- If a banner has only 1 valid value after correction, remove the entire banner
- Preserve all valid banners and values as-is
- Only modify what needs fixing
- PRESERVE the "category" field on each banner exactly as provided — do not remove or rename it"""


def _validate_banners(banner_spec: dict,
                      questions: List[SurveyQuestion]) -> dict:
    """Step 3: 생성된 배너를 실제 보기 코드 대비 검증.

    Args:
        banner_spec: Step 2의 배너 스펙 ({"banners": [...]})
        questions: 전체 문항 리스트 (코드 맵 생성용)

    Returns:
        검증/수정된 배너 스펙 dict
    """
    code_map = _build_code_map(questions)


    lines = []
    lines.append("## Banners to Validate")
    lines.append(_json.dumps(banner_spec, ensure_ascii=False, indent=2))
    lines.append("")
    lines.append("## Valid Code Map (question -> valid codes)")
    # 관련 문항만 포함 (compact)
    relevant_qns = set()
    for b in banner_spec.get("banners", []):
        for qn in b.get("source_questions", []):
            relevant_qns.add(qn)
        for v in b.get("values", []):
            cond = v.get("condition", "")
            for part in cond.split("&"):
                qn_part = part.split("=")[0].strip()
                if qn_part:
                    relevant_qns.add(qn_part)

    for qn in sorted(relevant_qns):
        codes = code_map.get(qn, [])
        if codes:
            lines.append(f"  {qn}: [{', '.join(codes)}]")
        else:
            lines.append(f"  {qn}: (no codes found)")

    user_prompt = "\n".join(lines)

    try:
        result = _call_llm_json_with_fallback(
            _BANNER_VALIDATION_SYSTEM_PROMPT, user_prompt,
            DEFAULT_MODEL, temperature=0.1, max_tokens=16384,
        )
        summary = result.get("validation_summary", "")
        if summary:
            logger.info(f"Banner validation: {summary}")
        return result
    except Exception as e:
        logger.error(f"Banner validation failed: {e}")
        return banner_spec  # 실패 시 원본 반환


# ── Legacy Banner Prompt (폴백용) ────────────────────────────────────

_BANNER_SYSTEM_PROMPT = """You are a senior DP specialist and industry domain expert designing cross-tabulation banners for marketing research.

## Your Approach
Analyze the questionnaire and research context to determine the most insightful banner categories for this specific study. Think as both a **marketing research expert** and a **domain specialist for the industry being studied**.

Banner categories should be derived from the study's objectives and question content — NOT from generic templates. Any question type (profiling, behavioral, attitudinal, experiential) can become a banner dimension if cross-tabulating by it reveals actionable insights.

Demographics alone are NEVER sufficient. The majority of banners should come from questions that measure behavior, attitudes, experience, or strategic segments.

## Critical Constraints
1. **Generate 8-15 banners** grouped into at least 3 thematic categories that you determine from the questionnaire.
2. **Each banner should have 3-6 values**. Group codes into meaningful NET categories. Avoid binary splits unless naturally binary (e.g., Gender).
3. **Every banner value MUST have an explicit condition** using "QN=code" format.
4. **At least 2 composite banners** combining 2+ questions with "&".
5. **Pure demographic dimensions should be ≤ 30%** of total banners.

## What to EXCLUDE
- **Screening/filter questions** used to terminate respondents
- **Country/market-specific questions** (filter references geography)
- **Open-ended questions** — text responses cannot be cross-tabulated

## How to Group Banner Values (NET logic)
- **Scales/ratings**: Top2 / Mid / Bot2 (3-way split), NOT individual codes
- **Age**: 3-4 meaningful bands (e.g., 18-29, 30-44, 45+)
- **Regions**: 2-4 major geographic clusters
- **Brands**: Strategic segments (Client Brand / Key Competitors / Others / Non-Owner)
- **Behavioral**: 3-way splits (Heavy/Medium/Light, Recent/Occasional/Lapsed)
- **Attitudinal**: 3-way splits (Promoter/Passive/Detractor, Satisfied/Neutral/Dissatisfied)

## Condition Format Rules
- Single code: "SQ1=1"
- Multiple codes (OR): "SQ1=1,2,3"
- Combined questions (AND): "Q3=1&Q5=1"
- **NEVER** use negative conditions ("!=", "NOT", "≠")
- Values within a banner must be **mutually exclusive**

## JSON Output Format
{
  "banners": [
    {
      "category": "Thematic group name you determined (e.g., Demographics, Ownership & Journey, Attitudes, etc.)",
      "name": "Banner name (the analytical dimension)",
      "rationale": "1-2 sentence explanation of WHY this banner is analytically valuable",
      "source_questions": ["SQ1"],
      "values": [
        {"label": "Value label (short)", "condition": "SQ1=1"},
        {"label": "Value label (short)", "condition": "SQ1=2,3"}
      ]
    }
  ]
}

**IMPORTANT**: Every banner MUST have a `category` field. Group banners into 3-5 thematic categories that you determine from the questionnaire content. Use descriptive category names specific to this study (e.g., "Vehicle Ownership & Journey", "EV Attitudes & Readiness", "Brand Relationship").

## Example: Good Banner Design for a Brand Tracking Study
{
  "banners": [
    {
      "category": "Demographics",
      "name": "Gender",
      "rationale": "Standard demographic cut to identify gender-based differences in brand metrics.",
      "source_questions": ["SQ1"],
      "values": [
        {"label": "Male", "condition": "SQ1=1"},
        {"label": "Female", "condition": "SQ1=2"}
      ]
    },
    {
      "category": "Demographics",
      "name": "Age Group",
      "rationale": "Generational segmentation reveals different brand consideration sets.",
      "source_questions": ["SQ2"],
      "values": [
        {"label": "18-29", "condition": "SQ2=1,2"},
        {"label": "30-44", "condition": "SQ2=3,4,5"},
        {"label": "45+", "condition": "SQ2=6,7,8"}
      ]
    },
    {
      "category": "Ownership & Usage",
      "name": "Ownership Segment",
      "rationale": "Client brand owners vs competitors reveals satisfaction drivers and switching barriers.",
      "source_questions": ["SQ5", "SQ6"],
      "values": [
        {"label": "Client Brand Owner", "condition": "SQ6=1"},
        {"label": "Domestic Competitor", "condition": "SQ6=2,3"},
        {"label": "Import Brand", "condition": "SQ6=4,5,6,7,8"},
        {"label": "Non-Owner", "condition": "SQ5=99"}
      ]
    },
    {
      "category": "Attitudes & Evaluation",
      "name": "Overall Satisfaction",
      "rationale": "Satisfaction tiers reveal which segments need retention vs growth strategies.",
      "source_questions": ["Q15"],
      "values": [
        {"label": "Satisfied (Top2)", "condition": "Q15=1,2"},
        {"label": "Neutral", "condition": "Q15=3"},
        {"label": "Dissatisfied (Bot2)", "condition": "Q15=4,5"}
      ]
    },
    {
      "category": "Attitudes & Evaluation",
      "name": "Purchase Intent",
      "rationale": "Intent levels help prioritize acquisition targets and messaging.",
      "source_questions": ["SQ11"],
      "values": [
        {"label": "High Intent (Top2)", "condition": "SQ11=1,2"},
        {"label": "Medium", "condition": "SQ11=3"},
        {"label": "Low Intent (Bot2)", "condition": "SQ11=4,5"}
      ]
    },
    {
      "category": "Composite Segments",
      "name": "Brand Loyalty Segment",
      "rationale": "Combining ownership and intent reveals retention vs conquest priorities.",
      "banner_type": "composite",
      "source_questions": ["SQ6", "SQ11"],
      "values": [
        {"label": "Loyal Owner", "condition": "SQ6=1&SQ11=1,2"},
        {"label": "At-Risk Owner", "condition": "SQ6=1&SQ11=3,4,5"},
        {"label": "Conquest Target", "condition": "SQ6=2,3,4,5,6,7,8&SQ11=1,2"},
        {"label": "Low Potential", "condition": "SQ6=2,3,4,5,6,7,8&SQ11=3,4,5"}
      ]
    }
  ]
}"""


def _banner_id_from_index(i: int) -> str:
    """배너 인덱스 → ID 문자열 (A-Z, 이후 AA, AB, ..., AZ, BA, ...)."""
    if i < 26:
        return chr(65 + i)
    # 26 이상: AA=26, AB=27, ... AZ=51, BA=52, ...
    first = chr(65 + (i // 26) - 1)
    second = chr(65 + (i % 26))
    return first + second


def _assign_categories_from_plan(banner_spec: dict, analysis_plan: dict) -> None:
    """Analysis Plan의 dimension→category 매핑으로 배너에 카테고리 부여.

    source_questions와 candidate_questions의 겹침, 또는
    배너 이름과 dimension 이름의 유사도로 매칭합니다.
    이미 category가 있는 배너는 건드리지 않습니다.
    """
    if not analysis_plan:
        return

    # Build dimension → category mappings
    dim_cat_map: dict[str, str] = {}    # dim_name_lower → category_name
    qs_cat_map: list[tuple[set, str]] = []  # (candidate_qs_set, category_name)
    for cat in analysis_plan.get("categories", []):
        cat_name = cat.get("category_name", "")
        if not cat_name:
            continue
        for dim in cat.get("banner_dimensions", []):
            dim_name = dim.get("dimension_name", "")
            if dim_name:
                dim_cat_map[dim_name.lower()] = cat_name
            cqs = set(dim.get("candidate_questions", []))
            if cqs:
                qs_cat_map.append((cqs, cat_name))

    for banner in banner_spec.get("banners", []):
        if banner.get("category"):
            continue  # Already assigned

        bname = (banner.get("name", "") or "").lower()

        # Try 1: Match by banner name ↔ dimension name (substring)
        for dim_name_lower, cat_name in dim_cat_map.items():
            if dim_name_lower in bname or bname in dim_name_lower:
                banner["category"] = cat_name
                break

        if banner.get("category"):
            continue

        # Try 2: Match by source_questions overlap with candidate_questions
        src_qs = set(banner.get("source_questions", []))
        if src_qs:
            best_cat = ""
            best_overlap = 0
            for cqs, cat_name in qs_cat_map:
                overlap = len(src_qs & cqs)
                if overlap > best_overlap:
                    best_overlap = overlap
                    best_cat = cat_name
            if best_cat:
                banner["category"] = best_cat


def _infer_banner_category(banner_name: str) -> str:
    """배너 이름에서 카테고리를 추정 (category 필드가 없는 경우 폴백).

    매칭 순서가 중요: 더 구체적인 키워드를 먼저 검사.
    범용적으로 다양한 산업·도메인에 대응하도록 넓은 키워드를 포함.
    """
    name_lower = banner_name.lower()

    # 매칭 규칙: (키워드 리스트, 카테고리명) — 순서대로 첫 매칭 반환.
    # 더 구체적인(multi-word) 패턴을 먼저 배치하여 오분류 방지.
    _CATEGORY_RULES = [
        # ── 1. Technology & Innovation (EV, AI, robotics, hydrogen 등) ──
        ([" ev ", "ev ", " ev", "electric", "전기차", "hybrid", "하이브리드",
          "engine type", "엔진 유형", "powertrain", "hydrogen", "수소",
          "autonomous", "자율주행", "robotaxi", "robot", "로봇",
          "ai ", " ai", "artificial intelligence", "인공지능",
          "iot", "smart home", "스마트홈", "connected", "커넥티드",
          "mobility", "모빌리티", "drone", "드론", "technology", "기술"],
         "Technology & Innovation"),

        # ── 2. Health & Wellness (before Brand — "health awareness" → Health) ──
        (["health", "건강", "wellness", "웰빙", "fitness", "운동",
          "nutrition", "영양", "diet", "식이", "medical", "의료",
          "mental health", "stress", "스트레스", "well-being",
          "symptom", "증상", "treatment", "치료", "pharmaceutical", "제약"],
         "Health & Wellness"),

        # ── 3. Media & Communication (before Purchase — "media consumption" → Media) ──
        (["media", "미디어", "social media", "sns", "소셜",
          "advertising", "광고", "digital", "디지털",
          "tv", "radio", "online", "오프라인", "offline",
          "content", "콘텐츠", "information source", "정보원",
          "streaming", "platform", "플랫폼", "influencer"],
         "Media & Communication"),

        # ── 4. Household & Family ──
        (["household", "가구", "family", "가족", "children", "자녀",
          "kids", "아이", "pet", "반려동물", "living arrangement", "거주",
          "housing", "주거", "home ownership", "주택"],
         "Household & Family"),

        # ── 5. Brand & Funnel ──
        (["brand", "funnel", "awareness", "loyalty", "consideration",
          "ownership segment", "브랜드", "보유 세그먼트", "인지", "충성",
          "top of mind", "switching", "전환"],
         "Brand & Ownership"),

        # ── 6. Attitudes & Values ──
        (["attitude", "perception", "sentiment", "interest", "관심",
          "opinion", "의견", "belief", "concern",
          "importance", "중요", "priority", "우선순위",
          "willingness", "의향", "openness", "readiness",
          "value", "가치", "lifestyle", "라이프스타일",
          "culture", "문화", "k-culture", "k-pop", "한류",
          "mindset", "aspiration", "motivation", "동기",
          "environmental", "sustainability", "환경", "지속가능",
          "trust", "신뢰", "confidence", "자신감", "preference", "선호"],
         "Attitudes & Values"),

        # ── 7. Satisfaction & Evaluation ──
        (["satisfaction", "satisfied", "rating", "nps", "recommend",
          "만족", "평가", "추천", "experience", "경험",
          "expectation", "기대", "performance", "성능",
          "quality", "품질", "csat", "ces"],
         "Satisfaction & Evaluation"),

        # ── 8. Purchase & Usage Behavior ──
        (["purchase", "buying", "구매", "usage", "사용", "이용",
          "frequency", "빈도", "spending", "지출", "budget", "예산",
          "channel", "채널", "touchpoint", "접점", "shopping",
          "subscription", "구독", "adoption", "도입",
          "engagement", "참여", "activity", "활동",
          "consumption", "소비", "behavior", "행동",
          "journey", "여정", "decision", "routine", "습관", "habit"],
         "Purchase & Usage"),

        # ── 9. Vehicle & Ownership (automotive specific) ──
        (["car ownership", "vehicle", "차량", "자동차", "car type",
          "fleet", "mileage", "주행거리", "drivetrain", "sedan",
          "suv", "truck", "pickup"],
         "Vehicle Ownership"),

        # ── 10. Demographics (넓은 매칭 — 구체적 카테고리 다 실패 시) ──
        (["gender", "age group", "age band", "age tier", "generation",
          "region", "city", "province", "state", "area",
          "income", "education", "occupation", "marital", "ethnicity",
          "race", "nationality", "employment", "job", "sector",
          "성별", "연령", "지역", "소득", "학력", "직업",
          "세대", "결혼", "인종", "민족", "고용"],
         "Demographics"),

        # ── 11. Segment & Composite ──
        (["segment", "세그먼트", "cluster", "클러스터", "typology",
          "tier", "cohort", "persona",
          "composite", "combined"],
         "Segments"),
    ]

    padded = f" {name_lower} "
    for keywords, category in _CATEGORY_RULES:
        for kw in keywords:
            if kw in name_lower or kw in padded:
                return category

    return ""


def _parse_banner_spec_to_models(raw: dict) -> List[Banner]:
    """LLM JSON 배너 스펙을 Banner 모델 리스트로 변환."""
    banners = []
    for i, b_data in enumerate(raw.get("banners", [])):
        banner_id = _banner_id_from_index(i)
        source_qs = b_data.get("source_questions", [])
        source_str = "&".join(source_qs) if source_qs else ""

        points = []
        for j, v_data in enumerate(b_data.get("values", [])):
            condition = v_data.get("condition", "")
            # source_question: condition에서 문항번호 추출 또는 배너 레벨 source 사용
            if condition:
                # "SQ1=1" → "SQ1", "A1=2&A2=5" → "A1&A2"
                parts = condition.split("&")
                sq = "&".join(p.split("=")[0].strip() for p in parts)
            else:
                sq = source_str

            points.append(BannerPoint(
                point_id=f"BP_{banner_id}_{j + 1}",
                label=v_data.get("label", ""),
                source_question=sq,
                condition=condition,
            ))

        banner_type = b_data.get("banner_type", "simple")
        # "&" 조건이 있으면 composite으로 강제
        has_composite = any("&" in (v.get("condition", "")) for v in b_data.get("values", []))
        if has_composite:
            banner_type = "composite"

        banner_name = b_data.get("name", f"Banner {banner_id}")
        category = b_data.get("category", "")
        if not category:
            category = _infer_banner_category(banner_name)

        banners.append(Banner(
            banner_id=banner_id,
            name=banner_name,
            points=points,
            rationale=b_data.get("rationale", ""),
            banner_type=banner_type,
            category=category,
        ))

    return banners


def _fallback_heuristic_candidates(questions: List[SurveyQuestion],
                                    intelligence: dict | None) -> List[SurveyQuestion]:
    """Step 1 실패 시 폴백: 배너 후보 문항 선정.

    모든 SA/MA 문항 중 OE, 스크리닝 종료, 국가/마켓 필터를 제외한
    나머지를 후보로 선정. 데모뿐 아니라 behavioral/attitudinal 문항도 포함.
    """
    country_keywords = ["country", "market", "국가", "마켓", "나라"]
    oe_keywords = ["oe", "open", "verbatim", "기타 기재", "서술"]

    candidates = []
    seen_qn = set()
    for q in questions:
        if q.question_number in seen_qn:
            continue
        seen_qn.add(q.question_number)

        # 보기가 2개 미만이면 배너 불가
        if len(q.answer_options) < 2:
            continue

        # 국가/마켓 필터 문항 제외
        filt_lower = (q.filter_condition or "").lower()
        if any(kw in filt_lower for kw in country_keywords):
            continue

        # OE 문항 제외
        qtype = (q.question_type or "").lower()
        qtext = (q.question_text or "").lower()
        if any(kw in qtype for kw in oe_keywords):
            continue

        candidates.append(q)

    return candidates


def _fallback_direct_banner(candidates: List[SurveyQuestion],
                             survey_context: str,
                             language: str = "en") -> List[Banner]:
    """분석 계획 없이 직접 배너 설계 (폴백 경로).

    GPT-5에 전체 문항을 전달하여 산업/조사목적에 맞는 배너를 직접 설계.
    """
    if not candidates:
        return []

    lines = []
    if survey_context:
        lines.append(survey_context)
        lines.append("")
    lines.append(f"Design cross-tabulation banner specifications from the following questionnaire "
                 f"({len(candidates)} questions, language: {language}).\n"
                 "Analyze the research objectives and industry context to determine the most "
                 "insightful banner categories. Select the best banner candidates from ALL "
                 "question types — profiling, behavioral, attitudinal, and create composite "
                 "segments where analytically valuable.\n")
    lines.append(_format_questions_compact(candidates, include_options=True))

    user_prompt = "\n".join(lines)

    try:
        raw = call_llm_json(_BANNER_SYSTEM_PROMPT, user_prompt, MODEL_BANNER_SUGGESTER)
    except Exception as e:
        logger.error(f"Fallback banner suggestion failed: {e}")
        return []

    return _parse_banner_spec_to_models(raw)


_MIN_BANNER_COUNT = 10          # 최소 배너 수 (was 8)
_MIN_COMPOSITE_COUNT = 4        # 최소 composite 배너 수 (was 3)
_MIN_DEEP_COMPOSITE = 1         # 3+ 문항 composite 최소 1개 (NEW)
_MIN_CATEGORY_COUNT = 4         # 최소 카테고리 수 (was 3)
_MAX_RETRY = 2                  # 품질 미달 시 재시도 횟수 (총 3회 시도)


_DEMO_KW_PATTERN = re.compile(
    r'\b(?:gender|age\s*(?:group|band)?|region|city|income|education'
    r'|occupation|marital|household|demographic)'
    r'|(?:성별|연령|나이|지역|소득|학력|직업|결혼|가구)',
    re.IGNORECASE,
)

# _assess_plan_quality에서도 사용 — frozenset 폴백 (카테고리명 전체 매칭용)
_DEMO_KEYWORDS = frozenset([
    "demographics", "demographic", "인구통계",
])

_MIN_AVG_VALUES = 3.0           # 배너당 평균 최소 value 수 (was 2.8)
_MAX_DEMO_RATIO = 0.35          # 데모 배너 비율 상한 (was 0.40)


def _is_demo_banner(banner: dict) -> bool:
    """배너가 인구통계(demographics) 배너인지 판별.

    \b 워드 바운더리로 'age'가 'usage'/'engaged' 등에 오매칭되지 않도록 함.
    """
    cat = (banner.get("category", "") or "").lower()
    name = (banner.get("name", "") or "").lower()
    text = f"{cat} {name}"
    return bool(_DEMO_KW_PATTERN.search(text))


def _assess_banner_quality(banner_spec: dict) -> dict:
    """배너 스펙의 품질 지표를 평가.

    Returns:
        dict with total_banners, composite_count, category_count, categories,
        avg_values, demo_ratio, deep_composite_count, total_values, issues, pass.
    """
    banners = banner_spec.get("banners", [])
    total = len(banners)
    composite_count = 0
    deep_composite_count = 0  # 3+ source questions
    categories = set()
    total_values = 0
    demo_count = 0

    for b in banners:
        cat = b.get("category", "")
        if cat:
            categories.add(cat)
        values = b.get("values", [])
        total_values += len(values)

        btype = b.get("banner_type", "simple")
        has_and = any("&" in v.get("condition", "") for v in values)
        if btype == "composite" or has_and:
            composite_count += 1
            # 3+ source questions 체크
            src_qs = set(b.get("source_questions", []))
            # condition에서도 문항 수 추출
            for v in values:
                cond = v.get("condition", "")
                for part in cond.split("&"):
                    qn = part.split("=")[0].strip()
                    if qn:
                        src_qs.add(qn)
            if len(src_qs) >= 3:
                deep_composite_count += 1

        if _is_demo_banner(b):
            demo_count += 1

    avg_values = total_values / total if total > 0 else 0
    demo_ratio = demo_count / total if total > 0 else 0

    issues = []
    if total < _MIN_BANNER_COUNT:
        issues.append(f"Only {total} banners (minimum: {_MIN_BANNER_COUNT})")
    if composite_count < _MIN_COMPOSITE_COUNT:
        issues.append(f"Only {composite_count} composite banners (minimum: {_MIN_COMPOSITE_COUNT})")
    if len(categories) < _MIN_CATEGORY_COUNT:
        issues.append(f"Only {len(categories)} categories (minimum: {_MIN_CATEGORY_COUNT})")
    if avg_values < _MIN_AVG_VALUES:
        issues.append(f"Avg {avg_values:.1f} values/banner (minimum: {_MIN_AVG_VALUES})")
    if deep_composite_count < _MIN_DEEP_COMPOSITE:
        issues.append(f"Only {deep_composite_count} deep composite (3+ questions) banners (minimum: {_MIN_DEEP_COMPOSITE})")
    if total >= 6 and demo_ratio > _MAX_DEMO_RATIO:
        issues.append(f"Demographics {demo_count}/{total} ({demo_ratio:.0%}) exceeds {_MAX_DEMO_RATIO:.0%} cap — add behavioral/attitudinal banners")

    return {
        "total_banners": total,
        "composite_count": composite_count,
        "deep_composite_count": deep_composite_count,
        "category_count": len(categories),
        "categories": sorted(categories),
        "total_values": total_values,
        "avg_values": round(avg_values, 1),
        "demo_count": demo_count,
        "demo_ratio": round(demo_ratio, 2),
        "issues": issues,
        "pass": len(issues) == 0,
    }


def _assess_plan_quality(plan: dict) -> dict:
    """분석 계획의 품질 지표를 평가.

    Returns:
        dict with total_dims, composite_dims, composite_ratio, category_count,
        has_non_demo_category, issues, pass.
    """
    dims = plan.get("banner_dimensions", [])
    total = len(dims)
    cats = plan.get("categories", [])
    cat_count = len(cats)
    composite_count = sum(1 for d in dims if d.get("is_composite"))
    composite_ratio = composite_count / total if total > 0 else 0

    # 데모 외 카테고리 존재 여부 (behavioral/attitudinal/composite)
    has_non_demo_category = False
    demo_dim_count = 0
    for cat in cats:
        cat_name = (cat.get("category_name", "") or "").lower()
        is_demo_cat = bool(_DEMO_KW_PATTERN.search(cat_name))
        if not is_demo_cat:
            has_non_demo_category = True
        else:
            demo_dim_count += len(cat.get("banner_dimensions", []))

    demo_dim_ratio = demo_dim_count / total if total > 0 else 0

    issues = []
    warnings = []
    if total < 8:
        issues.append(f"Only {total} dimensions (minimum: 8)")
    if composite_ratio < 0.30:
        issues.append(f"Only {composite_count}/{total} composite ({composite_ratio:.0%}, minimum: 30%)")
    if cat_count < 3:
        issues.append(f"Only {cat_count} categories (minimum: 3)")
    if cat_count > 8:
        issues.append(f"{cat_count} categories exceeds maximum (8)")
    if not has_non_demo_category:
        issues.append("All categories are demographics — must include behavioral/attitudinal/composite categories")
    if total >= 6 and demo_dim_ratio > 0.40:
        issues.append(f"Demographic dimensions {demo_dim_count}/{total} ({demo_dim_ratio:.0%}) — add behavioral/attitudinal dimensions")

    # ── Expert Consensus 관련 품질 체크 ──
    agreement_score = plan.get("agreement_score", None)
    if agreement_score is not None and agreement_score < 0.6:
        warnings.append(f"Low expert agreement score: {agreement_score:.2f}")

    # Research Plan 기반: primary objective 커버리지 체크
    research_plan = plan.get("_research_plan")
    if research_plan:
        primary_objs = [
            obj for obj in research_plan.get("research_objectives", [])
            if obj.get("priority") == "primary"
        ]
        if primary_objs:
            # 각 primary objective의 related_questions가 dims에 포함되는지 확인
            dim_questions = set()
            for d in dims:
                for qn in d.get("candidate_questions", []):
                    dim_questions.add(qn)
            uncovered = []
            for obj in primary_objs:
                related = set(obj.get("related_questions", []))
                if related and not (related & dim_questions):
                    uncovered.append(obj.get("id", obj.get("description", "")[:30]))
            if uncovered:
                warnings.append(f"Uncovered primary objectives: {', '.join(uncovered)}")

    return {
        "total_dims": total,
        "composite_dims": composite_count,
        "composite_ratio": composite_ratio,
        "category_count": cat_count,
        "has_non_demo_category": has_non_demo_category,
        "demo_dim_ratio": round(demo_dim_ratio, 2),
        "issues": issues,
        "warnings": warnings,
        "pass": len(issues) == 0,
    }


def suggest_banner_points(
    questions: List[SurveyQuestion],
    language: str = "ko",
    survey_context: str = "",
    intelligence: dict | None = None,
    progress_callback: Callable | None = None,
) -> tuple[List[Banner], dict | None]:
    """Expert Consensus 파이프라인으로 배너 후보 제안.

    Research Plan → Expert Panel (3명 병렬) → Synthesis → Banner Design → Validation
    각 단계에서 품질 미달 시 재시도. 실패 시 graceful degradation.

    Args:
        questions: 전체 문항 리스트
        language: 설문지 언어
        survey_context: Study Brief + Intelligence + Question Flow
        intelligence: Survey Intelligence 결과 dict
        progress_callback: (event, data) 진행 상태 콜백

    Returns:
        tuple: (배너 리스트, 분석 계획 dict 또는 None)
    """
    if not questions:
        return [], None

    def _cb(event: str, data: dict):
        if progress_callback:
            progress_callback(event, data)

    # ── Step 0.5: Research Plan (with retry — Change 3) ──
    _cb("phase", {"name": "research_plan", "status": "start"})
    research_plan = None
    for rp_attempt in range(_MAX_RETRY + 1):
        tag = f" (retry {rp_attempt})" if rp_attempt > 0 else ""
        logger.info(f"Banner pipeline Step 0.5: Creating research plan...{tag}")
        research_plan = _create_research_plan(questions, language, survey_context, intelligence)
        if not research_plan:
            break

        objectives = research_plan.get("research_objectives", [])
        dim_map = research_plan.get("objective_dimension_map", [])
        all_dims = [d for m in dim_map for d in m.get("dimensions", [])]

        if len(objectives) >= 3 and len(all_dims) >= 6:
            break  # 품질 충분

        if rp_attempt < _MAX_RETRY:
            logger.warning(f"Research plan quality low (objectives={len(objectives)}, "
                           f"dimensions={len(all_dims)}) — retrying")
        else:
            logger.warning(f"Research plan quality low after retries — proceeding")
    _cb("phase", {"name": "research_plan", "status": "done"})

    if not research_plan or not research_plan.get("research_objectives"):
        logger.warning("Research plan failed — falling back to legacy analysis plan")
        return _suggest_banner_points_legacy(
            questions, language, survey_context, intelligence, progress_callback,
        )

    # ── Step 1: Expert Panel (3명 병렬) ──
    _cb("phase", {"name": "expert_panel", "status": "start", "count": 3})
    logger.info("Banner pipeline Step 1: Running expert panel (3 experts in parallel)...")
    expert_outputs = _run_expert_panel(
        research_plan, questions, language, survey_context,
        progress_callback=progress_callback,
        intelligence=intelligence,
    )

    if len(expert_outputs) < 2:
        logger.warning(f"Only {len(expert_outputs)} expert(s) succeeded (need >=2 for consensus) "
                       "— falling back to legacy analysis plan")
        return _suggest_banner_points_legacy(
            questions, language, survey_context, intelligence, progress_callback,
        )

    # ── Step 1.5: Synthesis ──
    _cb("phase", {"name": "synthesis", "status": "start"})
    logger.info("Banner pipeline Step 1.5: Synthesizing expert consensus...")
    analysis_plan = _synthesize_expert_consensus(
        expert_outputs, research_plan, questions, language,
        survey_context=survey_context, intelligence=intelligence,
    )

    if not analysis_plan or not analysis_plan.get("banner_dimensions"):
        logger.warning("Synthesis failed — falling back to legacy analysis plan")
        return _suggest_banner_points_legacy(
            questions, language, survey_context, intelligence, progress_callback,
        )

    try:
        agreement = float(analysis_plan.get("agreement_score", 0))
    except (TypeError, ValueError):
        agreement = 0.0
    analysis_plan["agreement_score"] = agreement
    _cb("phase", {"name": "synthesis", "status": "done", "agreement_score": agreement})

    # Store expert outputs and research plan in analysis_plan for UI access
    analysis_plan["_expert_outputs"] = expert_outputs
    if "_research_plan" not in analysis_plan:
        analysis_plan["_research_plan"] = research_plan

    # ── Quality Gate ──
    plan_quality = _assess_plan_quality(analysis_plan)
    if not plan_quality["pass"]:
        logger.warning(f"Consensus plan quality below threshold: {plan_quality['issues']} — proceeding anyway")

    # ── Step 2: Banner Design (with quality gate) ──
    _cb("phase", {"name": "banner_design", "status": "start"})
    banner_spec = None
    for attempt in range(_MAX_RETRY + 1):
        tag = f" (retry {attempt})" if attempt > 0 else ""
        logger.info(f"Banner pipeline Step 2: Designing banners from consensus plan...{tag}")
        banner_spec = _design_banners_from_plan(
            analysis_plan, questions, language, survey_context,
            intelligence=intelligence,
        )

        if not banner_spec or not banner_spec.get("banners"):
            logger.warning("Step 2 failed — returning empty banners")
            return [], analysis_plan

        banner_quality = _assess_banner_quality(banner_spec)
        if banner_quality["pass"]:
            logger.info(f"Step 2 quality OK: {banner_quality['total_banners']} banners, "
                        f"{banner_quality['composite_count']} composite "
                        f"({banner_quality['deep_composite_count']} deep), "
                        f"{banner_quality['category_count']} categories, "
                        f"avg {banner_quality['avg_values']} values/banner, "
                        f"demo {banner_quality['demo_ratio']:.0%}")
            break

        if attempt < _MAX_RETRY:
            logger.warning(f"Step 2 quality below threshold: {banner_quality['issues']} — retrying")
        else:
            logger.warning(f"Step 2 quality below threshold after retries: {banner_quality['issues']} — proceeding anyway")
    _cb("phase", {"name": "banner_design", "status": "done"})

    # ── Step 2.5: Assign categories from analysis plan (robust fallback) ──
    _assign_categories_from_plan(banner_spec, analysis_plan)

    # ── Step 3: Validation ──
    _cb("phase", {"name": "validation", "status": "start"})
    logger.info("Banner pipeline Step 3: Validating banners...")
    validated_spec = _validate_banners(banner_spec, questions)

    # Validation LLM이 category 필드를 드랍하는 경우 원본에서 복원
    orig_banners = banner_spec.get("banners", [])
    orig_cat_map = {ob.get("name", ""): ob.get("category", "")
                    for ob in orig_banners if ob.get("name")}
    for i, vb in enumerate(validated_spec.get("banners", [])):
        if not vb.get("category"):
            if i < len(orig_banners):
                vb["category"] = orig_banners[i].get("category", "")
            if not vb.get("category"):
                vb["category"] = orig_cat_map.get(vb.get("name", ""), "")

    # 검증 결과 파싱 (실패 시 Step 2 결과 사용)
    banners = _parse_banner_spec_to_models(validated_spec)

    if not banners:
        logger.warning("Validation removed all banners — using pre-validation results")
        banners = _parse_banner_spec_to_models(banner_spec)
    _cb("phase", {"name": "validation", "status": "done"})

    # ── Final quality log ──
    composite_final = sum(1 for b in banners if b.banner_type == "composite")
    cat_final = len(set(b.category for b in banners if b.category))
    total_pts = sum(len(b.points) for b in banners)
    avg_pts = total_pts / len(banners) if banners else 0
    logger.info(f"Banner pipeline complete: {len(banners)} banners "
                f"({composite_final} composite, {cat_final} categories, "
                f"{total_pts} total values, avg {avg_pts:.1f}/banner) "
                f"[agreement={agreement:.2f}]")
    return banners, analysis_plan


def _suggest_banner_points_legacy(
    questions: List[SurveyQuestion],
    language: str = "ko",
    survey_context: str = "",
    intelligence: dict | None = None,
    progress_callback: Callable | None = None,
) -> tuple[List[Banner], dict | None]:
    """Legacy 파이프라인 (폴백): Analysis Plan → Banner Design → Validation.

    Expert consensus 실패 시 사용되는 기존 3-step CoT 파이프라인.
    """
    def _cb(event: str, data: dict):
        if progress_callback:
            progress_callback(event, data)

    analysis_plan = None

    # ── Step 1: Analysis Plan (with quality gate) ──
    _cb("phase", {"name": "banner_design", "status": "start"})
    for attempt in range(_MAX_RETRY + 1):
        tag = f" (retry {attempt})" if attempt > 0 else ""
        logger.info(f"Legacy pipeline Step 1: Creating analysis plan...{tag}")
        analysis_plan = _create_analysis_plan(questions, language, survey_context, intelligence)

        if not analysis_plan or not analysis_plan.get("banner_dimensions"):
            logger.warning("Step 1 failed or empty — falling back to heuristic")
            candidates = _fallback_heuristic_candidates(questions, intelligence)
            if not candidates:
                return [], None
            banners = _fallback_direct_banner(candidates, survey_context, language)
            return banners, None

        plan_quality = _assess_plan_quality(analysis_plan)
        if plan_quality["pass"]:
            break
        if attempt < _MAX_RETRY:
            logger.warning(f"Step 1 quality: {plan_quality['issues']} — retrying")
        else:
            logger.warning(f"Step 1 quality: {plan_quality['issues']} — proceeding")

    # ── Step 2: Banner Design (with quality gate) ──
    banner_spec = None
    for attempt in range(_MAX_RETRY + 1):
        tag = f" (retry {attempt})" if attempt > 0 else ""
        logger.info(f"Legacy pipeline Step 2: Designing banners...{tag}")
        banner_spec = _design_banners_from_plan(analysis_plan, questions, language, survey_context)

        if not banner_spec or not banner_spec.get("banners"):
            return [], analysis_plan

        banner_quality = _assess_banner_quality(banner_spec)
        if banner_quality["pass"]:
            break
        if attempt < _MAX_RETRY:
            logger.warning(f"Step 2 quality: {banner_quality['issues']} — retrying")
        else:
            logger.warning(f"Step 2 quality: {banner_quality['issues']} — proceeding")
    _cb("phase", {"name": "banner_design", "status": "done"})

    _assign_categories_from_plan(banner_spec, analysis_plan)

    # ── Step 3: Validation ──
    _cb("phase", {"name": "validation", "status": "start"})
    logger.info("Legacy pipeline Step 3: Validating banners...")
    validated_spec = _validate_banners(banner_spec, questions)

    orig_banners = banner_spec.get("banners", [])
    orig_cat_map = {ob.get("name", ""): ob.get("category", "")
                    for ob in orig_banners if ob.get("name")}
    for i, vb in enumerate(validated_spec.get("banners", [])):
        if not vb.get("category"):
            if i < len(orig_banners):
                vb["category"] = orig_banners[i].get("category", "")
            if not vb.get("category"):
                vb["category"] = orig_cat_map.get(vb.get("name", ""), "")

    banners = _parse_banner_spec_to_models(validated_spec)
    if not banners:
        banners = _parse_banner_spec_to_models(banner_spec)
    _cb("phase", {"name": "validation", "status": "done"})

    composite_final = sum(1 for b in banners if b.banner_type == "composite")
    cat_final = len(set(b.category for b in banners if b.category))
    total_pts = sum(len(b.points) for b in banners)
    avg_pts = total_pts / len(banners) if banners else 0
    logger.info(f"Legacy pipeline complete: {len(banners)} banners "
                f"({composite_final} composite, {cat_final} categories, "
                f"{total_pts} total values, avg {avg_pts:.1f}/banner)")
    return banners, analysis_plan


# ======================================================================
# Banner-to-Question Assignment — helpers
# ======================================================================

_QN_RE = re.compile(r'([A-Za-z]+\d+[a-zA-Z]*)')


def _extract_all_banner_qns(banner: Banner) -> set[str]:
    """배너의 모든 조건에서 참조하는 문항번호를 대문자로 추출."""
    qns: set[str] = set()
    for pt in banner.points:
        if pt.source_question:
            for sq in pt.source_question.split("&"):
                sq = sq.strip()
                if sq:
                    qns.add(sq.upper())
        if pt.condition:
            for m in _QN_RE.findall(pt.condition):
                qns.add(m.upper())
    return qns


def _extract_filter_qns(filter_condition: str) -> set[str]:
    """필터 조건에서 참조하는 문항번호를 대문자로 추출."""
    if not filter_condition:
        return set()
    return {m.upper() for m in _QN_RE.findall(filter_condition)}


# ======================================================================
# Banner-to-Question Assignment
# ======================================================================

def assign_banners_to_questions(questions: List[SurveyQuestion],
                                 banners: List[Banner]) -> dict:
    """문항 role/유형 기반 배너 자동 할당 (semantic fitness rules).

    할당 규칙:
    1. screening → Total only (배너 없음)
    2. demographics → Total only (배너 소스 문항이므로 자기 자신에게 배너 불필요)
    3. OE 문항 → Total only (주관식은 교차분석 불가)
    4. 배너 조건(source_question + condition)에서 참조하는 문항 → 해당 배너 제외 (자기참조 방지)
    5. 문항 필터가 배너 소스와 겹치면 → 해당 배너 제외 (filter overlap)
    6. 나머지 본조사 문항 → All applicable banners

    Returns:
        dict: {question_number: "A,B,C" 형태의 배너 ID 문자열}
    """
    if not banners:
        return {q.question_number: "" for q in questions}

    # 배너별 참조 문항 수집 (source_question + condition 모두)
    banner_ref_map: dict[str, set[str]] = {
        b.banner_id: _extract_all_banner_qns(b) for b in banners
    }
    banner_map: dict[str, Banner] = {b.banner_id: b for b in banners}

    all_banner_ids = [b.banner_id for b in banners]
    result = {}
    seen: set[str] = set()

    for q in questions:
        if q.question_number in seen:
            continue
        seen.add(q.question_number)

        role = (q.role or "").lower()
        qtype = (q.question_type or "").upper()

        # Fallback: role이 비어있으면 문항번호 prefix로 추정
        if not role:
            qn_upper = q.question_number.upper()
            if re.match(r'^S\d', qn_upper) or re.match(r'^SQ\d', qn_upper) or re.match(r'^SC\d', qn_upper):
                role = "screening"
            elif re.match(r'^D\d', qn_upper) or re.match(r'^DQ\d', qn_upper) or re.match(r'^F\d', qn_upper):
                role = "demographics"

        # Rule 1-2: screening/demographics → Total only
        if role in ("screening", "demographics"):
            result[q.question_number] = ""
            continue

        # Rule 3: OE → Total only
        if "OE" in qtype or "OPEN" in qtype:
            result[q.question_number] = ""
            continue

        # Rule 4+5+6: All banners except self-referencing / filter overlap
        qn_upper = q.question_number.upper()
        filter_qns = _extract_filter_qns(q.filter_condition or "")

        applicable = []
        for bid in all_banner_ids:
            banner_qns = banner_ref_map.get(bid, set())
            # Rule 4: self-reference — 문항이 배너 조건에서 참조됨
            if qn_upper in banner_qns:
                continue
            # Rule 5: filter overlap — 문항 필터가 배너 소스와 겹침
            if filter_qns and filter_qns & banner_qns:
                continue
            applicable.append(bid)

        # Rule 6: Sort applicable banners by semantic relevance score (desc)
        if applicable and len(applicable) > 1:
            applicable.sort(
                key=lambda bid: _score_banner_relevance(q, banner_map[bid]),
                reverse=True,
            )
            # Optional threshold filtering (currently 0.0 = no filtering)
            if _MIN_RELEVANCE_THRESHOLD > 0:
                applicable = [
                    bid for bid in applicable
                    if _score_banner_relevance(q, banner_map[bid]) >= _MIN_RELEVANCE_THRESHOLD
                ]

        result[q.question_number] = ",".join(applicable)

    return result


def expand_banner_ids(banner_ids_str: str, banners: List[Banner]) -> str:
    """'A,B,C' → 'A(Gender), B(Age), C(Ownership)' 변환 (서비스 레이어용).

    Args:
        banner_ids_str: 쉼표 구분 배너 ID 문자열
        banners: Banner 객체 리스트
    """
    if not banner_ids_str or not banner_ids_str.strip():
        return ""
    if not banners:
        return banner_ids_str

    bid_to_name = {b.banner_id: b.name for b in banners}
    parts = []
    for bid in banner_ids_str.split(","):
        bid = bid.strip()
        name = bid_to_name.get(bid, "")
        if name:
            parts.append(f"{bid}({name})")
        else:
            parts.append(bid)
    return ", ".join(parts)


# ======================================================================
# Phase 3: Sort Order
# ======================================================================

def generate_sort_orders(questions: List[SurveyQuestion]) -> dict:
    """문항 유형별 기본 정렬 규칙 생성 (순수 알고리즘).

    Returns:
        dict: {question_number: sort_order_string}
    """
    result = {}
    seen_qn = set()

    for q in questions:
        if q.question_number in seen_qn:
            continue
        seen_qn.add(q.question_number)

        qtype = (q.question_type or "").strip().upper()

        if "SCALE" in qtype or re.match(r'\d+\s*PT\s*X\s*\d+', qtype):
            result[q.question_number] = "by code"
        elif re.match(r'(?i)(TOP|RANK)\s*\d+', qtype):
            result[q.question_number] = "by % desc"
        elif "MA" in qtype:
            result[q.question_number] = "by % desc"
        elif "OE" in qtype or "OPEN" in qtype:
            result[q.question_number] = "by code"
        else:
            # SA 및 기타
            result[q.question_number] = "by code"

    return result


# ======================================================================
# Phase 3: SubBanner
# ======================================================================

_SUBBANNER_SYSTEM_PROMPT = """You are a DP specialist for marketing research cross-tabulation.
Identify questions that need a SubBanner (secondary analysis dimension), typically for grid/matrix questions.

## Rules
1. Grid/matrix questions (e.g., "5PT X 10", "GRID", "MATRIX", "5PT SCALE X 8") need sub-banners
   when they evaluate multiple items on a common scale.
2. The sub-banner = the list of row stubs (items being rated), NOT the scale points.
3. Extract item names from the question text, answer options, or the Instructions field.
4. **Piping detection**: If the text says "Q3에서 선택한 항목", "pipe from Q3", "items selected at Q3",
   reference the source question instead of listing items (e.g., "Items selected at Q3").
5. Single-scale questions with NO item list → return empty string.
6. List 3-5 representative items followed by "등" or "etc." to keep it concise.
7. Only suggest sub-banners for matrix/grid questions. Return empty for SA, MA, OE, and others.
8. Write in the same language as the question text.

## Examples

### Example 1 — Standard matrix (Korean)
Question: "다음 각 브랜드에 대해 얼마나 만족하십니까?"
Type: 5PT X 8
→ sub_banner: "평가 항목 (브랜드 A, 브랜드 B, 브랜드 C 등)"

### Example 2 — Standard matrix (English)
Question: "How satisfied are you with each of the following aspects?"
Type: 7PT SCALE X 5
→ sub_banner: "Rated items (Price, Quality, Service, Variety, etc.)"

### Example 3 — Piping (Korean)
Question: "Q3에서 선택한 브랜드 각각에 대해 평가해 주세요"
Type: 5PT X GRID
→ sub_banner: "Q3에서 선택한 항목"

### Example 4 — Piping (English)
Question: "Please rate each brand you selected in Q5"
Type: MATRIX
→ sub_banner: "Items selected at Q5"

### Example 5 — Single scale, no items
Question: "Overall, how satisfied are you?"
Type: 5PT SCALE
→ sub_banner: ""

## JSON Output Format
{
  "results": [
    {"question_number": "Q5", "sub_banner": "Rated items (Taste, Price, Quality, etc.)"},
    {"question_number": "Q6", "sub_banner": ""}
  ]
}"""


def suggest_sub_banners(questions: List[SurveyQuestion],
                        language: str = "ko",
                        survey_context: str = "") -> dict:
    """Grid/매트릭스 문항의 SubBanner 제안.

    Returns:
        dict: {question_number: sub_banner_string}
    """
    result = {}
    matrix_qs = []
    seen_qn = set()

    for q in questions:
        if q.question_number in seen_qn:
            continue
        seen_qn.add(q.question_number)

        qtype = (q.question_type or "").strip().upper()
        if (re.match(r'\d+\s*PT\s*X\s*\d+', qtype)
                or re.match(r'\d+\s*PT\s+SCALE\s*X\s*\d+', qtype)
                or "GRID" in qtype or "MATRIX" in qtype):
            matrix_qs.append(q)
        else:
            result[q.question_number] = ""

    if not matrix_qs:
        return result

    lines = []
    if survey_context:
        lines.append(survey_context)
        lines.append("")
    lines.append("Suggest SubBanner dimensions for these grid/matrix questions:\n")
    for q in matrix_qs:
        lines.append(f"[{q.question_number}] {q.question_text}")
        lines.append(f"  Type: {q.question_type or ''}")
        lines.append(f"  Options: {q.answer_options_compact()}")
        if q.instructions:
            lines.append(f"  Instructions: {q.instructions}")
        if q.filter_condition:
            lines.append(f"  Filter: {q.filter_condition}")
        if q.skip_logic:
            skip_text = q.skip_logic_display()[:200]
            lines.append(f"  Skip: {skip_text}")
        lines.append("")

    user_prompt = "\n".join(lines)

    try:
        raw = call_llm_json(_SUBBANNER_SYSTEM_PROMPT, user_prompt, MODEL_SUBBANNER_SUGGESTER)
        for r in raw.get("results", []):
            qn = str(r.get("question_number", "")).strip()
            sb = str(r.get("sub_banner", "")).strip()
            if qn:
                result[qn] = sb
    except Exception as e:
        logger.error(f"SubBanner suggestion failed: {e}")

    for q in matrix_qs:
        if q.question_number not in result:
            result[q.question_number] = ""

    return result


# ======================================================================
# Phase 4: Special Instructions
# ======================================================================

_SPECIAL_INSTR_SYSTEM_PROMPT = """You are a DP specialist for marketing research.
Generate Special Instructions (programming notes) for survey cross-tabulation tables.

## Pattern Categories (detect and generate instructions for these)
1. **Rotation/Randomization**: "보기 로테이션", "randomize", "rotate", "shuffle"
   → "Randomize option order" / "보기 순서 로테이션"
2. **Piping**: "pipe from Q3", "Q3에서 가져오기", "items selected at Q3"
   → "Pipe from Q3" / "Q3에서 파이핑"
3. **Open-end Coding**: OE/open-ended questions
   → "Open-end coding required" / "주관식 코딩 필요"
4. **Exclusive Answer**: "단독응답", "exclusive", "none of the above", "해당 없음"
   → "Exclusive code handling required" / "단독응답 코드 처리 필요"
5. **Scale Anchoring**: Labeled endpoints on scales (e.g., "1=전혀 아님, 5=매우 그러함")
   → "Scale anchoring: 1=Not at all, 5=Very much" / "척도 양 끝단: 1=전혀 아님, 5=매우 그러함"
6. **Rank Limit**: "rank top 3", "상위 3개 선택", "최대 N개"
   → "Rank limit: top 3" / "순위 제한: 상위 3개"
7. **Quota Reference**: quota, weighting, 쿼터, 가중치
   → "Quota/weight reference" / "쿼터/가중치 참조"
8. **NET Grouping**: Top2Box, Bottom2Box, NET calculation
   → "Top2/Bottom2 NET" / "Top2/Bottom2 NET 계산"
9. **Multiple Response**: MA questions, multiple punch
   → "Multiple response" / "복수응답 처리"
10. **Filter/Skip Instruction**: conditional display, skip patterns
    → Include the filter/skip condition in the instruction
11. **Show Card**: "SHOW CARD", "보기 카드", "show list", "카드 제시"
    → "Show Card" / "보기 카드 제시"

## Rules
- If no special instruction is needed, return empty string.
- Combine multiple applicable patterns with " / " separator.
- Keep instructions concise and actionable.
- Write in the same language as the question text.
- When Filter or Skip fields are present, incorporate them into the instruction.

## Examples

### Korean
- Q1 (로테이션+단독응답): "보기 순서 로테이션 / 단독응답 코드 처리 필요"
- Q5 (파이핑+순위): "Q3에서 파이핑 / 순위 제한: 상위 3개"
- Q8 (주관식): "주관식 코딩 필요"

### English
- Q1 (rotation+exclusive): "Randomize option order / Exclusive code handling required"
- Q5 (piping+rank): "Pipe from Q3 / Rank limit: top 3"
- Q8 (OE): "Open-end coding required"

## JSON Output Format
{
  "results": [
    {"question_number": "Q1", "instruction": "Randomize option order / Exclusive code handling required"},
    {"question_number": "Q5", "instruction": "Pipe selected brands from Q3"},
    {"question_number": "Q10", "instruction": ""}
  ]
}"""


def generate_special_instructions(questions: List[SurveyQuestion],
                                  language: str = "ko",
                                  progress_callback=None,
                                  survey_context: str = "") -> dict:
    """Special Instructions 생성 — 패턴 매칭 + LLM.

    Returns:
        dict: {question_number: instruction_string}
    """
    result = {}
    needs_llm = []
    seen_qn = set()

    # 패턴 매칭 키워드
    rotation_patterns = re.compile(
        r'(로테이션|randomiz|rotat|shuffle|무작위)', re.IGNORECASE
    )
    piping_patterns = re.compile(
        r'(pipe|파이핑|가져오기|from\s+[A-Z]+\d+|에서\s+선택)', re.IGNORECASE
    )
    exclusive_patterns = re.compile(
        r'(exclusive|단독\s*응답|배타적|single[\-\s]*punch|해당\s*없음|none\s+of\s+the\s+above)',
        re.IGNORECASE,
    )
    rank_patterns = re.compile(
        r'(rank\s+top\s*\d+|상위\s*\d+\s*개|순위|최대\s*\d+\s*개|select\s+up\s+to\s*\d+)',
        re.IGNORECASE,
    )
    show_card_patterns = re.compile(
        r'(SHOW\s+CARD|보기\s*카드|show\s+list|카드\s*제시)', re.IGNORECASE
    )

    for q in questions:
        if q.question_number in seen_qn:
            continue
        seen_qn.add(q.question_number)

        instructions_text = (q.instructions or "") + " " + (q.question_text or "")
        options_text = " ".join(
            opt.label for opt in (q.answer_options or []) if opt.label
        )
        search_text = instructions_text + " " + options_text
        auto_parts = []

        if rotation_patterns.search(search_text):
            auto_parts.append("Randomize option order" if language == "en"
                              else "보기 순서 로테이션")

        if piping_patterns.search(search_text):
            auto_parts.append("Pipe from previous question" if language == "en"
                              else "이전 문항에서 파이핑")

        qtype = (q.question_type or "").strip().upper()
        if "OE" in qtype or "OPEN" in qtype:
            auto_parts.append("Open-end coding required" if language == "en"
                              else "주관식 코딩 필요")

        if exclusive_patterns.search(search_text):
            auto_parts.append("Exclusive code handling required" if language == "en"
                              else "단독응답 코드 처리 필요")

        if rank_patterns.search(search_text):
            auto_parts.append("Rank limit applies" if language == "en"
                              else "순위 제한 적용")

        if show_card_patterns.search(search_text):
            auto_parts.append("Show Card" if language == "en"
                              else "보기 카드 제시")

        if not auto_parts and ("MA" in qtype or "MULTI" in qtype):
            auto_parts.append("Multiple response" if language == "en"
                              else "복수응답 처리")

        if auto_parts:
            result[q.question_number] = " / ".join(auto_parts)
        else:
            needs_llm.append(q)

    if not needs_llm:
        return result

    # LLM으로 복잡한 패턴 감지
    batches = [needs_llm[i:i + BATCH_SIZE] for i in range(0, len(needs_llm), BATCH_SIZE)]
    total_batches = len(batches)

    for batch_idx, batch in enumerate(batches):
        if progress_callback:
            progress_callback("si_batch_start", {
                "batch_index": batch_idx, "total_batches": total_batches,
                "question_count": len(batch),
            })

        lines = []
        if survey_context:
            lines.append(survey_context)
            lines.append("")
        lines.append("Generate special instructions for these questions if needed:\n")
        for q in batch:
            lines.append(f"[{q.question_number}] {q.question_text}")
            lines.append(f"  Type: {q.question_type or ''}")
            if q.instructions:
                lines.append(f"  Instructions: {q.instructions}")
            if q.answer_options:
                lines.append(f"  Options: {q.answer_options_compact()}")
            if q.filter_condition:
                lines.append(f"  Filter: {q.filter_condition}")
            if q.skip_logic:
                skip_text = q.skip_logic_display()[:200]
                lines.append(f"  Skip: {skip_text}")
            lines.append("")

        user_prompt = "\n".join(lines)

        try:
            raw = call_llm_json(
                _SPECIAL_INSTR_SYSTEM_PROMPT, user_prompt, MODEL_SPECIAL_INSTRUCTIONS
            )
            for r in raw.get("results", []):
                qn = str(r.get("question_number", "")).strip()
                instr = str(r.get("instruction", "")).strip()
                if qn:
                    result[qn] = instr
        except Exception as e:
            logger.error(f"Special instructions batch {batch_idx} failed: {e}")

        for q in batch:
            if q.question_number not in result:
                result[q.question_number] = ""

        if progress_callback:
            progress_callback("si_batch_done", {
                "batch_index": batch_idx, "total_batches": total_batches,
            })

    return result


# ======================================================================
# Phase 4: Compile & Export
# ======================================================================

def compile_table_guide(doc: SurveyDocument, project_name: str = "",
                        language: str = "ko") -> TableGuideDocument:
    """전 필드 통합 Table Guide 조립.

    Returns:
        TableGuideDocument
    """
    rows = []
    for q in doc.questions:
        rows.append({
            "QuestionNumber": q.question_number,
            "TableNumber": q.table_number,
            "QuestionText": q.question_text,
            "TableTitle": q.table_title,
            "QuestionType": q.question_type or "",
            "SummaryType": q.summary_type,
            "Base": q.base,
            "NetRecode": q.net_recode,
            "Sort": q.sort_order,
            "SubBanner": q.sub_banner,
            "BannerIDs": q.banner_ids,
            "SpecialInstructions": q.special_instructions,
            "AnswerOptions": q.answer_options_compact(),
            "Filter": q.filter_condition or "",
        })

    return TableGuideDocument(
        project_name=project_name or doc.filename,
        filename=doc.filename,
        generated_at=datetime.now().isoformat(),
        banners=doc.banners,
        rows=rows,
        language=language,
    )


def export_table_guide_excel(tg_doc: TableGuideDocument,
                             survey_doc: SurveyDocument,
                             intelligence: dict | None = None) -> bytes:
    """다중시트 전문 Table Guide Excel 출력.

    Sheets:
        1. Cover — 프로젝트 정보 + Survey Intelligence
        2. Table Guide — 메인 테이블 (Filter 포함)
        3. Banner Spec — 배너 정의
        4. Net/Recode Spec — Net/Recode 상세
        5. Answer Options — 응답 보기
    """
    wb = Workbook()
    header_fill = PatternFill(start_color="0033A0", end_color="0033A0", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    wrap_align = Alignment(wrap_text=True, vertical='top')
    center_align = Alignment(horizontal='center', vertical='center')

    def _style_header(ws):
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

    # ── Sheet 1: Cover ──
    ws_cover = wb.active
    ws_cover.title = "Cover"
    ws_cover.append(["Table Guide"])
    ws_cover['A1'].font = Font(size=18, bold=True)
    ws_cover.append([])
    ws_cover.append(["Project", tg_doc.project_name])
    ws_cover.append(["Source File", tg_doc.filename])
    ws_cover.append(["Generated", tg_doc.generated_at])
    ws_cover.append(["Language", tg_doc.language])
    ws_cover.append(["Total Questions", len(tg_doc.rows)])
    ws_cover.append(["Banners", len(tg_doc.banners)])

    # Survey Intelligence 요약
    intel = intelligence or (survey_doc.survey_intelligence if survey_doc else None)
    if intel:
        ws_cover.append([])
        ws_cover.append(["Survey Intelligence"])
        ws_cover[ws_cover.max_row][0].font = Font(size=14, bold=True)
        if intel.get("client_name"):
            ws_cover.append(["Client", intel["client_name"]])
        if intel.get("study_type"):
            ws_cover.append(["Study Type", intel["study_type"]])
        objectives = intel.get("research_objectives", [])
        if objectives:
            ws_cover.append(["Research Objectives", " | ".join(objectives[:5])])
        segments = intel.get("key_segments", [])
        if segments:
            seg_strs = [f"{s.get('name', '')} ({s.get('type', '')})" for s in segments]
            ws_cover.append(["Key Segments", " | ".join(seg_strs)])
        # Banner 카테고리 요약
        if tg_doc.banners:
            cats = {}
            for b in tg_doc.banners:
                cat = b.category or "Other"
                if cat not in cats:
                    cats[cat] = []
                cats[cat].append(f"{b.banner_id}({b.name})")
            cat_lines = [f"{cat}: {', '.join(names)}" for cat, names in cats.items()]
            ws_cover.append(["Banner Structure", " | ".join(cat_lines)])

    ws_cover.column_dimensions['A'].width = 22
    ws_cover.column_dimensions['B'].width = 80

    # ── Sheet 2: Table Guide ──
    ws_tg = wb.create_sheet("Table Guide")
    tg_headers = [
        "Base", "Sort", "QuestionNumber", "TableNumber", "QuestionText",
        "TableTitle", "SubBanner", "QuestionType", "SummaryType",
        "NetRecode", "BannerIDs", "BannerNames", "SpecialInstructions",
        "Filter", "GrammarChecker",
    ]
    ws_tg.append(tg_headers)
    _style_header(ws_tg)

    # GrammarChecker 매핑
    qn_grammar = {}
    for q in survey_doc.questions:
        qn_grammar[q.question_number + "_" + q.table_number] = q.grammar_checked

    for row in tg_doc.rows:
        key = row["QuestionNumber"] + "_" + row["TableNumber"]
        gc = qn_grammar.get(key, "")
        banner_ids_val = row.get("BannerIDs", "")
        ws_tg.append([
            row.get("Base", ""),
            row.get("Sort", ""),
            row.get("QuestionNumber", ""),
            row.get("TableNumber", ""),
            row.get("QuestionText", ""),
            row.get("TableTitle", ""),
            row.get("SubBanner", ""),
            row.get("QuestionType", ""),
            row.get("SummaryType", ""),
            row.get("NetRecode", ""),
            banner_ids_val,
            expand_banner_ids(banner_ids_val, tg_doc.banners),
            row.get("SpecialInstructions", ""),
            row.get("Filter", ""),
            gc,
        ])

    for row in ws_tg.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_align

    tg_col_widths = [20, 12, 15, 12, 50, 35, 20, 12, 25, 30, 12, 40, 35, 30, 35]
    for i, w in enumerate(tg_col_widths, 1):
        ws_tg.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Banner Spec (데이터 있을 때만) ──
    if tg_doc.banners:
        ws_banner = wb.create_sheet("Banner Spec")
        ws_banner.append(["Category", "BannerID", "BannerName", "PointLabel",
                           "SourceQuestion", "Condition", "IsNet", "NetDefinition"])
        _style_header(ws_banner)

        for banner in tg_doc.banners:
            for pt in banner.points:
                ws_banner.append([
                    banner.category or "",
                    banner.banner_id,
                    banner.name,
                    pt.label,
                    pt.source_question,
                    pt.condition,
                    "Yes" if pt.is_net else "No",
                    pt.net_definition,
                ])

        banner_col_widths = {'A': 20, 'B': 12, 'C': 22, 'D': 25, 'E': 18, 'F': 30, 'G': 8, 'H': 25}
        for col_letter, w in banner_col_widths.items():
            ws_banner.column_dimensions[col_letter].width = w

    # ── Sheet 4: Net/Recode Spec (데이터 있을 때만) ──
    has_net = any(q.net_recode for q in survey_doc.questions)
    if has_net:
        ws_net = wb.create_sheet("Net Recode Spec")
        ws_net.append(["QuestionNumber", "QuestionType", "NetRecode"])
        _style_header(ws_net)

        seen_qn = set()
        for q in survey_doc.questions:
            if q.question_number in seen_qn:
                continue
            seen_qn.add(q.question_number)
            if q.net_recode:
                ws_net.append([q.question_number, q.question_type or "", q.net_recode])

        ws_net.column_dimensions['A'].width = 18
        ws_net.column_dimensions['B'].width = 15
        ws_net.column_dimensions['C'].width = 50

    # ── Sheet 5: Answer Options ──
    ws_opts = wb.create_sheet("Answer Options")
    ws_opts.append(["QuestionNumber", "OptionCode", "OptionLabel"])
    _style_header(ws_opts)

    for q in survey_doc.questions:
        for opt in q.answer_options:
            ws_opts.append([q.question_number, opt.code, opt.label])

    ws_opts.column_dimensions['A'].width = 18
    ws_opts.column_dimensions['B'].width = 12
    ws_opts.column_dimensions['C'].width = 50

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
