"""Path Simulator UI 페이지.

SurveyDocument의 스킵 로직 그래프를 분석하여
테스트 경로 시뮬레이션 + 테스트 시나리오를 표시한다.
LLM 불필요 — 버튼 클릭 시 즉시 계산.
"""

import io
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from services.path_simulator import (
    parse_condition,
    simulate_paths,
    trace_path,
    generate_persona_scenarios,
    SimulationResult,
    SimulatedPath,
    PersonaScenario,
)
from services.skip_logic_service import build_skip_logic_graph
from services.checklist_generator import generate_checklist


def page_path_simulator():
    st.title("Path Simulator")

    # Guard clause
    if "survey_document" not in st.session_state or st.session_state["survey_document"] is None:
        st.warning(
            '먼저 Questionnaire Analyzer에서 문서를 처리해주세요.',
        )
        return

    survey_doc = st.session_state["survey_document"]
    questions = survey_doc.questions
    if not questions:
        st.warning("문서에서 문항을 찾을 수 없습니다.")
        return

    st.info(
        f"**{survey_doc.filename}**에서 **{len(questions)}**개 문항을 발견했습니다. "
        "**경로 분석** 버튼을 클릭하여 모든 설문 경로를 시뮬레이션하세요.",
    )

    # Analyze button
    if st.button("경로 분석", type="primary"):
        st.session_state.pop("traced_path", None)
        with st.spinner("경로 분석 중..."):
            result = simulate_paths(questions)
            st.session_state["path_simulator_result"] = result

    # Results
    if "path_simulator_result" not in st.session_state:
        return

    result: SimulationResult = st.session_state["path_simulator_result"]

    st.divider()

    # Dashboard metrics
    _render_dashboard(result)

    # Graph analysis warnings
    _render_graph_warnings(result)

    st.divider()

    # Tabs
    tab_personas, tab_scenarios, tab_tracer, tab_paths = st.tabs(
        ["페르소나 시나리오", "분기 테스트", "인터랙티브 추적기", "전체 경로"]
    )

    with tab_personas:
        _render_persona_scenarios(questions, result)

    with tab_scenarios:
        _render_test_scenarios(result)

    with tab_tracer:
        _render_interactive_tracer(questions)

    with tab_paths:
        _render_all_paths(result)


def _render_dashboard(result: SimulationResult):
    """요약 메트릭 4칸."""
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("전체 경로", result.total_paths)
    with col2:
        st.metric("최장 경로", result.max_path_length)
    with col3:
        st.metric("최단 경로", result.min_path_length)
    with col4:
        st.metric("분기 커버리지", f"{result.branch_coverage_percent:.0f}%")


def _render_graph_warnings(result: SimulationResult):
    """그래프 분석 경고 표시."""
    analysis = result.graph_analysis

    if analysis.unreachable_questions:
        qns = ", ".join(analysis.unreachable_questions)
        st.warning(
            f"**도달 불가 문항 감지:** {qns}\n\n"
            "이 문항들은 첫 번째 문항에서 어떤 경로로도 도달할 수 없습니다.",
        )

    if analysis.loop_detected:
        for loop in analysis.loop_details[:3]:
            cycle = " -> ".join(loop)
            st.warning(f"**루프 감지:** {cycle}")

    if result.unparsed_conditions:
        items = [f"- **{qn}**: `{cond}`" for qn, cond in result.unparsed_conditions[:10]]
        st.info(
            f"**{len(result.unparsed_conditions)}**건의 스킵 조건을 파싱할 수 없습니다:\n\n"
            + "\n".join(items),
        )


def _render_persona_scenarios(questions, result: SimulationResult):
    """페르소나 기반 테스트 시나리오."""
    personas = generate_persona_scenarios(questions)

    if not personas:
        st.info("스크리닝/인구통계 문항이 없어 페르소나를 생성할 수 없습니다.")
        return

    st.subheader(f"페르소나 시나리오 ({len(personas)}건)")
    st.caption("인구통계/스크리닝 문항의 보기 조합으로 자동 생성된 대표 응답자 경로입니다.")

    for p in personas:
        icon = "🚫" if p.is_termination else "👤"
        with st.expander(
            f"{icon} 시나리오 {p.persona_id}: {p.persona_label} ({p.path_length}문항)",
            expanded=(p.persona_id <= 2),
        ):
            # 응답 선택
            sel_parts = [f"**{qn}**: {p.answer_labels.get(qn, '')} ({code})"
                         for qn, code in p.answer_selections.items()]
            st.markdown("**응답 선택:** " + " | ".join(sel_parts))

            # 경로
            path_str = " → ".join(p.expected_path[:15])
            if len(p.expected_path) > 15:
                path_str += f" ... (총 {len(p.expected_path)}문항)"
            st.markdown(f"**경로:** {path_str}")

            if p.is_termination:
                st.warning("이 페르소나는 스크리닝에서 탈락합니다. 탈락 처리가 올바른지 확인하세요.")

    # 엑셀에 포함
    st.session_state["persona_scenarios"] = personas


def _render_test_scenarios(result: SimulationResult):
    """테스트 시나리오 테이블 + Excel 다운로드."""
    scenarios = result.test_scenarios

    if not scenarios:
        st.info("테스트 시나리오가 생성되지 않았습니다 (스킵 로직 없음).")
        return

    st.subheader(f"테스트 시나리오 ({len(scenarios)}건)")

    rows = []
    for ts in scenarios:
        # 보기 라벨 포함: "Q1: 남성(1), Q3: 20대(2)"
        answer_parts = []
        for k, v in ts.answer_selections.items():
            label = ts.answer_labels.get(k, "")
            if label and not label.startswith("코드"):
                answer_parts.append(f"{k}: {label}({v})")
            else:
                answer_parts.append(f"{k}={v}")
        answers_str = ", ".join(answer_parts)

        path_str = " → ".join(ts.expected_path[:10])
        if len(ts.expected_path) > 10:
            path_str += f" ... (총 {len(ts.expected_path)}개)"
        branches_str = ", ".join(ts.verified_branches[:5])
        if len(ts.verified_branches) > 5:
            branches_str += f" ... (총 {len(ts.verified_branches)}개)"

        rows.append({
            "#": ts.scenario_id,
            "우선순위": ts.priority,
            "설명": ts.description,
            "응답 선택": answers_str,
            "예상 경로": path_str,
            "검증 분기": branches_str,
        })

    df = pd.DataFrame(rows)
    st.dataframe(
        df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "#": st.column_config.NumberColumn("#", width="small"),
            "우선순위": st.column_config.TextColumn("우선순위", width="small"),
            "설명": st.column_config.TextColumn("설명", width="medium"),
            "응답 선택": st.column_config.TextColumn("응답 선택", width="medium"),
            "예상 경로": st.column_config.TextColumn("예상 경로", width="large"),
            "검증 분기": st.column_config.TextColumn("검증 분기", width="medium"),
        },
    )

    # ── 통합 링크테스트 가이드 엑셀 ──
    survey_doc = st.session_state.get("survey_document")
    questions = survey_doc.questions if survey_doc else []

    # 체크리스트 생성 (알고리즘만, LLM 없이 즉시)
    checklist = None
    if questions:
        try:
            checklist = generate_checklist(questions, language="ko", use_llm=False)
        except Exception:
            pass

    excel_data = _build_link_test_excel(scenarios, checklist)
    st.download_button(
        label="링크테스트 가이드 다운로드 (Excel)",
        data=excel_data,
        file_name="link_test_guide.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


def _render_interactive_tracer(questions):
    """인터랙티브 경로 추적기."""
    st.subheader("인터랙티브 추적기")
    st.caption("스킵 로직이 있는 문항에 대해 응답을 선택한 후 '경로 추적' 버튼을 클릭하세요.")

    graph = build_skip_logic_graph(questions)

    # 스킵 로직이 있는 문항만 selectbox 표시
    questions_with_skip = [q for q in questions if q.skip_logic]

    if not questions_with_skip:
        st.info("스킵 로직이 있는 문항이 없습니다. 경로가 순차적입니다.")
        if st.button("순차 경로 표시", key="trace_sequential"):
            path = trace_path(questions, graph, {})
            _render_traced_path(path)
        return

    answer_selections: dict = {}

    for q in questions_with_skip:
        options = ["(선택 안 함)"]
        if q.answer_options:
            options += [f"{o.code}. {o.label}" for o in q.answer_options]
        else:
            # 스킵 조건에서 코드 추출
            codes = set()
            for sl in q.skip_logic:
                cond = parse_condition(sl.condition)
                if cond.is_parsed:
                    codes.update(cond.answer_codes)
            if codes:
                options += sorted(codes)

        q_label = f"{q.question_number}: {q.question_text[:60]}"
        if len(q.question_text) > 60:
            q_label += "..."

        selected = st.selectbox(
            q_label,
            options=options,
            key=f"tracer_{q.question_number}",
        )

        if selected and selected != "(선택 안 함)":
            # 코드만 추출 ("1. 매우 그렇다" → "1")
            code = selected.split(".")[0].strip()
            answer_selections[q.question_number] = code

    if st.button("경로 추적", type="primary", key="trace_btn"):
        path = trace_path(questions, graph, answer_selections)
        st.session_state["traced_path"] = path

    if "traced_path" in st.session_state:
        _render_traced_path(st.session_state["traced_path"])


def _render_traced_path(path: SimulatedPath):
    """추적 결과 경로 표시."""
    st.divider()
    st.subheader(f"추적 경로 ({path.length}단계)")

    # 경로 요약
    qn_display = " -> ".join(path.question_numbers[:15])
    if len(path.question_numbers) > 15:
        qn_display += f" ... ({path.length} total)"
    st.code(qn_display, language=None)

    # 상세 스텝
    for step in path.steps:
        skip_info = f" **SKIP -> {step.skip_triggered}**" if step.skip_triggered else ""
        answer_info = f" [Answer: {step.selected_answer}]" if step.selected_answer else ""
        terminal_info = " (TERMINAL)" if step.is_terminal and not step.skip_triggered else ""

        st.markdown(
            f"`{step.question_number}` ({step.question_type}) "
            f"{step.question_text[:80]}{answer_info}{skip_info}{terminal_info}"
        )


def _render_all_paths(result: SimulationResult):
    """모든 경로 expander 표시."""
    paths = result.all_paths

    if not paths:
        st.info("경로를 찾을 수 없습니다.")
        return

    st.subheader(f"전체 경로 ({len(paths)}건)")

    if len(paths) > 50:
        st.caption(f"{len(paths)}개 경로 중 처음 50개를 표시합니다.")
        paths_to_show = paths[:50]
    else:
        paths_to_show = paths

    for path in paths_to_show:
        qn_summary = " -> ".join(path.question_numbers[:10])
        if len(path.question_numbers) > 10:
            qn_summary += " ..."

        label = f"Path #{path.path_id} ({path.length} steps): {qn_summary}"

        with st.expander(label):
            for step in path.steps:
                skip_info = f" **-> {step.skip_triggered}**" if step.skip_triggered else ""
                terminal = " (END)" if step.is_terminal else ""
                st.markdown(
                    f"- `{step.question_number}` ({step.question_type}) "
                    f"{step.question_text[:80]}{skip_info}{terminal}"
                )


def _build_link_test_excel(scenarios, checklist) -> bytes:
    """시나리오 + 체크리스트 + 페르소나 통합 엑셀 생성.

    Sheet 1: 페르소나 시나리오 (인구통계 기반 경로)
    Sheet 2: 분기 테스트 시나리오 (스킵 로직 커버리지)
    Sheet 3: 체크리스트 (확인란 포함)
    Sheet 4: 시나리오별 체크 가이드 (프린트용)
    """
    wb = Workbook()
    hdr_fill = PatternFill(start_color="0033A0", end_color="0033A0", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="top")

    def _style_header(ws):
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center

    # ── Sheet 1: 페르소나 시나리오 ──
    ws_persona = wb.active
    ws_persona.title = "페르소나 시나리오"
    ws_persona.append(["#", "페르소나", "응답 선택", "예상 경로", "경로 길이", "탈락 여부"])
    _style_header(ws_persona)

    persona_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    personas = st.session_state.get("persona_scenarios", [])
    for p in personas:
        sel_parts = [f"{qn}: {p.answer_labels.get(qn, '')}({code})"
                     for qn, code in p.answer_selections.items()]
        row_num = ws_persona.max_row + 1
        ws_persona.append([
            p.persona_id,
            p.persona_label,
            ", ".join(sel_parts),
            " → ".join(p.expected_path[:15]),
            p.path_length,
            "탈락" if p.is_termination else "",
        ])
        if p.is_termination:
            for cell in ws_persona[row_num]:
                cell.fill = persona_fill

    for row in ws_persona.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    ws_persona.column_dimensions['A'].width = 5
    ws_persona.column_dimensions['B'].width = 35
    ws_persona.column_dimensions['C'].width = 45
    ws_persona.column_dimensions['D'].width = 50
    ws_persona.column_dimensions['E'].width = 10
    ws_persona.column_dimensions['F'].width = 10

    # ── Sheet 2: 분기 테스트 시나리오 ──
    ws1 = wb.create_sheet("분기 테스트")
    ws1.append(["#", "우선순위", "응답 선택", "예상 경로", "검증 분기"])
    _style_header(ws1)

    for ts in (scenarios or []):
        answer_parts = []
        for k, v in ts.answer_selections.items():
            label = ts.answer_labels.get(k, "")
            if label and not label.startswith("코드"):
                answer_parts.append(f"{k}: {label}({v})")
            else:
                answer_parts.append(f"{k}={v}")

        ws1.append([
            ts.scenario_id,
            ts.priority,
            ", ".join(answer_parts),
            " → ".join(ts.expected_path[:15]),
            ", ".join(ts.verified_branches[:8]),
        ])

    for row in ws1.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 45
    ws1.column_dimensions['D'].width = 50
    ws1.column_dimensions['E'].width = 35

    # ── Sheet 2: 체크리스트 ──
    ws2 = wb.create_sheet("체크리스트")
    ws2.append(["확인", "카테고리", "우선순위", "문항", "제목", "상세", "예상 동작"])
    _style_header(ws2)

    critical_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    warning_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

    checklist_items = checklist.items if checklist else []
    pri_kr = {"HIGH": "높음", "MEDIUM": "중간", "LOW": "낮음"}

    for item in checklist_items:
        row_num = ws2.max_row + 1
        ws2.append([
            "☐",
            item.category,
            pri_kr.get(item.priority, item.priority),
            item.question_number,
            item.title,
            item.detail,
            item.expected_behavior,
        ])
        if item.priority == "HIGH":
            for cell in ws2[row_num]:
                cell.fill = critical_fill
        elif item.priority == "MEDIUM":
            for cell in ws2[row_num]:
                cell.fill = warning_fill

    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 8
    ws2.column_dimensions['D'].width = 10
    ws2.column_dimensions['E'].width = 35
    ws2.column_dimensions['F'].width = 50
    ws2.column_dimensions['G'].width = 40

    # ── Sheet 3: 시나리오별 체크 가이드 (프린트용) ──
    ws3 = wb.create_sheet("시나리오별 체크 가이드")
    bold = Font(bold=True, size=11)
    scenario_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

    current_row = 1
    for ts in (scenarios or []):
        # 시나리오 헤더
        answer_parts = []
        for k, v in ts.answer_selections.items():
            label = ts.answer_labels.get(k, "")
            if label and not label.startswith("코드"):
                answer_parts.append(f"{k}: {label}({v})")
            else:
                answer_parts.append(f"{k}={v}")

        ws3.cell(row=current_row, column=1,
                 value=f"시나리오 {ts.scenario_id}: {', '.join(answer_parts)}")
        ws3.cell(row=current_row, column=1).font = bold
        ws3.cell(row=current_row, column=1).fill = scenario_fill
        current_row += 1

        ws3.cell(row=current_row, column=1, value=f"경로: {' → '.join(ts.expected_path[:12])}")
        ws3.cell(row=current_row, column=1).font = Font(size=9, color="666666")
        current_row += 1

        # 이 시나리오의 경로에 해당하는 체크항목
        path_qns = set(ts.expected_path)
        relevant_items = [
            item for item in checklist_items
            if item.question_number in path_qns or item.question_number == "GLOBAL"
        ]

        if relevant_items:
            ws3.cell(row=current_row, column=1, value="확인")
            ws3.cell(row=current_row, column=2, value="항목")
            ws3.cell(row=current_row, column=3, value="예상 동작")
            for cell in ws3[current_row]:
                if cell.value:
                    cell.font = Font(bold=True, size=9)
            current_row += 1

            for item in relevant_items[:15]:
                ws3.cell(row=current_row, column=1, value="☐")
                ws3.cell(row=current_row, column=2, value=f"[{item.question_number}] {item.title}")
                ws3.cell(row=current_row, column=3, value=item.expected_behavior)
                for cell in ws3[current_row]:
                    if cell.value:
                        cell.alignment = wrap
                        cell.font = Font(size=9)
                current_row += 1
        else:
            ws3.cell(row=current_row, column=1, value="(이 경로에 해당하는 체크항목 없음)")
            ws3.cell(row=current_row, column=1).font = Font(size=9, color="999999")
            current_row += 1

        current_row += 1  # 시나리오 간 빈 행

    ws3.column_dimensions['A'].width = 6
    ws3.column_dimensions['B'].width = 50
    ws3.column_dimensions['C'].width = 45

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
