"""Survey QA — 설문 로직 시각화 + 테스트 시나리오 + 체크리스트 통합.

3개 메뉴(Skip Logic, Path Simulator, Checklist)와
Quality Checker의 알고리즘 검출을 하나로 통합한다.
"""

import io
import re
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from services.path_simulator import (
    generate_persona_scenarios,
    PersonaScenario,
    simulate_paths,
    SimulationResult,
)
from services.skip_logic_service import build_skip_logic_graph, parse_target
from services.checklist_generator import generate_checklist
from services.prescripting_checker import (
    run_algorithmic_checks,
    ReviewReport,
    CATEGORY_LABELS,
    SEVERITY_LABELS,
)


WORKSTREAMS = [
    "설문지 수정 필요",
    "Script 구현 확인",
    "링크 테스트 확인",
]

WORKSTREAM_OWNERS = {
    "설문지 수정 필요": "Researcher",
    "Script 구현 확인": "Researcher / Script",
    "링크 테스트 확인": "Link Test",
}

CHECKLIST_CATEGORY_LABELS = {
    "SKIP_LOGIC": "스킵 로직",
    "PIPING": "파이핑",
    "ROTATION": "보기 로테이션",
    "EXCLUSIVE_OPTION": "배타적 보기",
    "SCALE_CONSISTENCY": "척도 일관성",
    "FILTER_VALIDITY": "필터 유효성",
    "DEAD_END": "도달 불가",
}


def _collect_qa_items(result: SimulationResult | None, review: ReviewReport, checklist) -> list[dict]:
    """Collect QA findings into researcher-facing workstreams."""
    items = []

    if result and result.unparsed_conditions:
        for qn, cond in result.unparsed_conditions:
            workstream = "링크 테스트 확인"
            items.append({
                "severity": "warning",
                "workstream": workstream,
                "owner": WORKSTREAM_OWNERS[workstream],
                "status": "미확인",
                "category": "수동 확인 필요",
                "question": qn,
                "title": f"{qn} 스킵 조건 수동 확인 필요",
                "detail": (
                    f"조건 '{cond}'를 자동 파싱할 수 없습니다.\n"
                    f"1. 원본 설문지에서 {qn}의 스킵 조건을 직접 확인\n"
                    f"2. 해당 조건 충족 시 올바른 문항으로 이동하는지 수동 테스트"
                ),
                "source": "파싱 실패",
            })

    for item in review.items:
        workstream = _workstream_for_review_item(item)
        items.append({
            "severity": item.severity,
            "workstream": workstream,
            "owner": WORKSTREAM_OWNERS[workstream],
            "status": "미확인",
            "category": CATEGORY_LABELS.get(item.category, item.category),
            "question": item.question_number,
            "title": item.title,
            "detail": item.detail,
            "source": "구조 검증",
        })

    pri_to_sev = {"HIGH": "critical", "MEDIUM": "warning", "LOW": "info"}
    if checklist:
        for item in checklist.items:
            workstream = _workstream_for_checklist_item(item.category, item.priority)
            items.append({
                "severity": pri_to_sev.get(item.priority, "info"),
                "workstream": workstream,
                "owner": WORKSTREAM_OWNERS[workstream],
                "status": "미확인",
                "category": CHECKLIST_CATEGORY_LABELS.get(item.category, item.category),
                "question": item.question_number,
                "title": item.title,
                "detail": item.detail,
                "expected": item.expected_behavior,
                "source": item.source,
            })

    return items


def _workstream_for_review_item(item) -> str:
    if item.category in {"duplicate_qn", "option_code", "skip_logic", "filter"}:
        return "설문지 수정 필요"
    if item.category in {"type_mismatch", "mece", "logic"}:
        return "설문지 수정 필요" if item.severity != "info" else "Script 구현 확인"
    return "Script 구현 확인"


def _workstream_for_checklist_item(category: str, priority: str) -> str:
    if category in {"DEAD_END", "SCALE_CONSISTENCY"}:
        return "설문지 수정 필요"
    if category in {"ROTATION", "EXCLUSIVE_OPTION", "PIPING"}:
        return "Script 구현 확인"
    if category in {"SKIP_LOGIC", "FILTER_VALIDITY"}:
        return "링크 테스트 확인"
    return "링크 테스트 확인" if priority == "HIGH" else "Script 구현 확인"


def _qa_decision(result: SimulationResult, qa_items: list[dict]) -> dict:
    """Return researcher-facing overall QA decision."""
    critical_count = sum(
        1 for item in qa_items
        if item["severity"] == "critical" and item["workstream"] == "설문지 수정 필요"
    )
    warning_count = sum(
        1 for item in qa_items
        if item["severity"] == "warning"
        or (item["severity"] == "critical" and item["workstream"] != "설문지 수정 필요")
    )
    survey_fix_count = sum(1 for item in qa_items if item["workstream"] == "설문지 수정 필요")
    unparsed_count = len(result.unparsed_conditions) if result else 0
    unreachable_count = len(result.graph_analysis.unreachable_questions) if result else 0
    loop_detected = bool(result and result.graph_analysis.loop_detected)

    if critical_count or unreachable_count or loop_detected:
        return {
            "label": "로직 수정 또는 수동 확인 필요",
            "tone": "error",
            "message": (
                "스크립팅 전달 전에 설문지 구조 또는 로직을 먼저 확인해야 합니다. "
                "특히 존재하지 않는 이동 대상, 도달 불가 문항, 루프, 중복 코드 같은 항목은 "
                "링크 테스트 전에 수정하는 것이 안전합니다."
            ),
            "critical": critical_count,
            "warning": warning_count,
            "survey_fix": survey_fix_count,
            "unparsed": unparsed_count,
        }

    if warning_count or unparsed_count:
        return {
            "label": "확인 후 스크립팅 전달 권장",
            "tone": "warning",
            "message": (
                "큰 구조 오류는 감지되지 않았지만, Script 구현 또는 링크 테스트에서 확인해야 할 항목이 있습니다. "
                "아래의 우선 확인 항목을 검토한 뒤 스크립팅 전달 또는 링크 테스트를 진행해주세요."
            ),
            "critical": critical_count,
            "warning": warning_count,
            "survey_fix": survey_fix_count,
            "unparsed": unparsed_count,
        }

    return {
        "label": "스크립팅 전달 가능",
        "tone": "ok",
        "message": (
            "자동 점검에서 큰 리스크가 감지되지 않았습니다. "
            "생성된 분기 테스트 시나리오를 링크 테스트 체크리스트로 활용하면 됩니다."
        ),
        "critical": critical_count,
        "warning": warning_count,
        "survey_fix": survey_fix_count,
        "unparsed": unparsed_count,
    }


def _workstream_counts(qa_items: list[dict]) -> dict:
    return {name: sum(1 for item in qa_items if item["workstream"] == name) for name in WORKSTREAMS}


def _extract_qn_refs(text: str) -> list[str]:
    if not text:
        return []
    refs = re.findall(r"\b([A-Za-z]+\d+[a-z]?(?:[-_]\d+)*)\b", text, flags=re.IGNORECASE)
    seen = set()
    result = []
    for ref in refs:
        norm = ref.upper()
        if norm not in seen:
            seen.add(norm)
            result.append(norm)
    return result


def _format_answer_selection(qn: str, code: str, label: str = "") -> str:
    if label and not label.startswith("코드"):
        return f"{qn}: {label} ({code})"
    return f"{qn}={code}"


def _build_logic_map_rows(questions) -> list[dict]:
    """Build a human-readable logic table for researchers."""
    qn_lookup = {q.question_number.upper(): q.question_number for q in questions}
    rows = []

    for q in questions:
        for sl in q.skip_logic:
            parsed = parse_target(sl.target)
            if parsed == "END":
                target = "TERMINATE / END"
                status = "종료"
            elif parsed:
                target = qn_lookup.get(parsed.upper(), parsed)
                status = "확인됨" if parsed.upper() in qn_lookup else "대상 확인 필요"
            else:
                target = sl.target
                status = "수동 확인 필요"

            rows.append({
                "유형": "스킵/종료",
                "기준 문항": q.question_number,
                "조건": sl.condition,
                "대상": target,
                "확인 포인트": "조건 충족 시 대상 문항으로 이동하고 중간 문항이 표시되지 않는지 확인",
                "상태": status,
            })

        filt = q.filter_condition or ""
        if filt and filt.lower() not in ("all respondents", "모두", "전원", "모두에게"):
            refs = _extract_qn_refs(filt)
            source = ", ".join(qn_lookup.get(ref, ref) for ref in refs) if refs else "조건 원문 확인"
            status = "확인됨" if refs and all(ref in qn_lookup for ref in refs) else "수동 확인 필요"
            rows.append({
                "유형": "필터/대상 조건",
                "기준 문항": source,
                "조건": filt,
                "대상": q.question_number,
                "확인 포인트": "조건 충족 시 대상 문항이 표시되고 미충족 시 표시되지 않는지 확인",
                "상태": status,
            })

    return rows


def _branch_context_map(questions) -> dict:
    rows = _build_logic_map_rows(questions)
    context = {}
    for row in rows:
        if row["유형"] != "스킵/종료":
            continue
        source = row["기준 문항"]
        target = str(row["대상"]).replace("TERMINATE / END", "END")
        context[f"{source}->{target}"] = row
    return context


def _scenario_rows(questions, result: SimulationResult) -> list[dict]:
    context = _branch_context_map(questions)
    rows = []

    for ts in result.test_scenarios:
        answer_parts = [
            _format_answer_selection(k, v, ts.answer_labels.get(k, ""))
            for k, v in ts.answer_selections.items()
        ]
        verified = ts.verified_branches[:5]
        first_branch = verified[0] if verified else ""
        branch_info = context.get(first_branch, {})
        if first_branch:
            purpose = (
                f"{branch_info.get('기준 문항', first_branch.split('->')[0])}에서 "
                f"{branch_info.get('대상', first_branch.split('->')[-1])} 이동 확인"
            )
        else:
            purpose = "순차 진행 경로 확인"

        rows.append({
            "#": ts.scenario_id,
            "우선순위": "필수" if ts.priority == "REQUIRED" else "권장",
            "테스트 목적": purpose,
            "응답 선택": ", ".join(answer_parts) if answer_parts else "특정 선택 없음",
            "예상 경로": " → ".join(ts.expected_path[:12]),
            "확인 포인트": ", ".join(verified) if verified else "전체 순차 흐름",
        })

    return rows


def page_survey_qa():
    st.title("Logic Checker")

    # Guard
    if "survey_document" not in st.session_state or st.session_state["survey_document"] is None:
        st.warning("먼저 Questionnaire Analyzer에서 문서를 처리해주세요.", icon="⚠️")
        return

    survey_doc = st.session_state["survey_document"]
    questions = survey_doc.questions
    if not questions:
        st.warning("문서에서 문항을 찾을 수 없습니다.", icon="⚠️")
        return

    st.info(f"**{survey_doc.filename}** — **{len(questions)}**개 문항", icon="🔍")

    # ── 분석 실행 버튼 ──
    run_clicked = st.button("QA 분석 실행", type="primary", use_container_width=True)

    if run_clicked:
        with st.status("QA 분석 중...", expanded=True) as status:
            # 1. 경로 시뮬레이션
            status.write("경로 시뮬레이션 중...")
            result = simulate_paths(questions)
            st.session_state["qa_simulation"] = result
            graph = build_skip_logic_graph(questions)

            status.write("대표 응답자 경로 생성 중...")
            personas = generate_persona_scenarios(questions, graph)
            st.session_state["qa_personas"] = personas

            # 2. 알고리즘 검출 (Quality Checker 통합)
            status.write("설문 구조 검증 중...")
            review = run_algorithmic_checks(questions)
            st.session_state["qa_review"] = review

            # 3. 체크리스트 생성 (알고리즘만, LLM 없이)
            status.write("체크리스트 생성 중...")
            checklist = generate_checklist(questions, language="ko", use_llm=False)
            st.session_state["qa_checklist"] = checklist

            total_issues = review.critical_count + review.warning_count + len(checklist.items)
            status.update(
                label=f"QA 분석 완료! 시나리오 {len(result.test_scenarios)}건, "
                      f"응답자 경로 {len(personas)}건, 체크항목 {total_issues}건",
                state="complete",
            )

    # ── 결과 표시 ──
    if "qa_simulation" not in st.session_state:
        st.caption("**QA 분석 실행** 버튼을 눌러 시작하세요.")
        return

    result = st.session_state["qa_simulation"]
    review = st.session_state.get("qa_review", ReviewReport())
    checklist = st.session_state.get("qa_checklist")
    personas = st.session_state.get("qa_personas", [])
    qa_items = _collect_qa_items(result, review, checklist)
    decision = _qa_decision(result, qa_items)

    _render_qa_summary(result, personas, qa_items, decision)

    st.divider()

    # ── 탭 ──
    tab_logic, tab_branch, tab_persona, tab_checklist = st.tabs([
        "로직 시각화", "분기 테스트", "응답자 경로", "체크리스트"
    ])

    with tab_logic:
        _render_logic_visualization(questions, result)

    with tab_branch:
        _render_branch_tests(questions, result)

    with tab_persona:
        _render_persona_scenarios(personas)

    with tab_checklist:
        _render_checklist(qa_items)

    # ── 통합 다운로드 ──
    st.divider()
    excel_data = _build_qa_excel(result, review, checklist, questions, personas, qa_items, decision)
    st.download_button(
        label="Logic Checker 결과 다운로드 (Excel)",
        data=excel_data,
        file_name="logic_checker.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


def _render_qa_summary(
    result: SimulationResult,
    personas: list[PersonaScenario],
    qa_items: list[dict],
    decision: dict,
):
    """Render the researcher-facing QA decision summary."""
    with st.container(border=True):
        st.markdown("### Logic Checker 결과")
        body = f"**상태: {decision['label']}**\n\n{decision['message']}"
        if decision["tone"] == "error":
            st.error(body, icon="🚨")
        elif decision["tone"] == "warning":
            st.warning(body, icon="⚠️")
        else:
            st.success(body, icon="✅")

        counts = _workstream_counts(qa_items)
        required_scenarios = sum(1 for s in result.test_scenarios if s.priority == "REQUIRED")
        termination_personas = sum(1 for p in personas if p.is_termination)

        col1, col2, col3, col4 = st.columns(4)
        col1.metric("필수 테스트 시나리오", required_scenarios)
        col2.metric("분기 커버리지", f"{result.branch_coverage_percent:.0f}%")
        col3.metric("설문지 수정 필요", counts["설문지 수정 필요"])
        col4.metric("Script/링크 확인", counts["Script 구현 확인"] + counts["링크 테스트 확인"])

        st.caption(
            f"전체 경로 {result.total_paths}개 · 대표 응답자 경로 {len(personas)}개 "
            f"(탈락 경로 {termination_personas}개) · "
            f"파싱 불가 조건 {decision['unparsed']}개"
        )

        priority_items = [
            item for item in qa_items
            if item["severity"] in ("critical", "warning")
        ][:5]
        if priority_items:
            st.markdown("**우선 확인 항목**")
            for idx, item in enumerate(priority_items, start=1):
                st.markdown(
                    f"{idx}. **{item['title']}** "
                    f"({item['workstream']} · {item['owner']})"
                )


# ══════════════════════════════════════════════════════════════
# 탭 1: 로직 시각화
# ══════════════════════════════════════════════════════════════

def _render_logic_visualization(questions, result: SimulationResult):
    """스킵 로직 Graphviz 시각화."""
    from services.skip_logic_service import build_skip_logic_graph, generate_dot

    graph = build_skip_logic_graph(questions)

    if not graph.edges:
        st.info("스킵 로직이 없습니다. 모든 문항이 순차 진행됩니다.")
        return

    col1, col2 = st.columns(2)
    with col1:
        view_mode = st.radio("보기 모드", ["스킵만", "전체 흐름"],
                             horizontal=True, key="qa_view_mode")
    with col2:
        orientation = st.radio("방향", ["위 → 아래", "왼쪽 → 오른쪽"],
                               horizontal=True, key="qa_orientation")

    mode = "skip_only" if view_mode == "스킵만" else "full_flow"
    rankdir = "TB" if orientation == "위 → 아래" else "LR"

    dot_str = generate_dot(graph, view_mode=mode, orientation=rankdir)
    st.graphviz_chart(dot_str)

    # 범례
    st.caption(
        "━━ 파란 실선: 스킵 로직 (조건부 이동)  |  "
        "┅┅ 주황 점선: 필터 조건 (조건부 표시)  |  "
        "── 회색 실선: 순차 진행"
    )

    # 경고
    analysis = result.graph_analysis
    if analysis.unreachable_questions:
        st.warning(f"**도달 불가 문항:** {', '.join(analysis.unreachable_questions)}")
    if analysis.loop_detected:
        for loop in analysis.loop_details[:3]:
            st.warning(f"**루프 감지:** {' → '.join(loop)}")
    if result.unparsed_conditions:
        with st.expander(f"파싱 불가 조건 ({len(result.unparsed_conditions)}건)"):
            for qn, cond in result.unparsed_conditions[:10]:
                st.markdown(f"- **{qn}**: `{cond}`")

    logic_rows = _build_logic_map_rows(questions)
    if logic_rows:
        st.subheader("로직 테이블")
        st.caption(
            "그래프를 표 형태로 풀어 쓴 목록입니다. Researcher와 Script가 From/조건/대상을 함께 확인할 때 사용하세요."
        )
        st.dataframe(pd.DataFrame(logic_rows), use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════
# 탭 2: 분기 테스트
# ══════════════════════════════════════════════════════════════

def _render_branch_tests(questions, result: SimulationResult):
    """분기 커버리지 테스트 시나리오."""
    scenarios = result.test_scenarios
    if not scenarios:
        st.info("스킵 로직이 없어 분기 테스트가 필요하지 않습니다.")
        return

    st.subheader(f"분기 테스트 ({len(scenarios)}건)")
    st.caption(
        f"자동 생성된 시나리오가 스킵 분기의 **{result.branch_coverage_percent:.0f}%**를 커버합니다. "
        "필수 시나리오는 링크 테스트에서 먼저 확인하세요."
    )

    rows = _scenario_rows(questions, result)
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    required = [s for s in scenarios if s.priority == "REQUIRED"]
    if required:
        st.markdown("**필수 시나리오 상세**")
        for ts in required[:8]:
            row = next((r for r in rows if r["#"] == ts.scenario_id), None)
            if not row:
                continue
            with st.expander(f"시나리오 {ts.scenario_id}: {row['테스트 목적']}", expanded=False):
                st.markdown(f"1. 응답 선택: {row['응답 선택']}")
                st.markdown(f"2. 예상 경로: {row['예상 경로']}")
                st.markdown(f"3. 확인 포인트: {row['확인 포인트']}")


def _render_persona_scenarios(personas: list[PersonaScenario]):
    """Representative respondent paths for researcher review."""
    if not personas:
        st.info(
            "대표 응답자 경로를 만들 수 없습니다. 스크리닝/인구통계 문항에 SA 보기와 명확한 문항번호가 있으면 자동 생성됩니다."
        )
        return

    st.subheader(f"대표 응답자 경로 ({len(personas)}건)")
    st.caption(
        "스크리닝/인구통계 문항의 대표 보기 조합으로 만든 응답자 경로입니다. "
        "주요 타깃, 비타깃, 탈락 응답자의 실제 흐름을 링크 테스트 전에 확인할 때 사용하세요."
    )

    rows = []
    for p in personas:
        selections = [
            _format_answer_selection(qn, code, p.answer_labels.get(qn, ""))
            for qn, code in p.answer_selections.items()
        ]
        rows.append({
            "#": p.persona_id,
            "응답자 유형": p.persona_label,
            "선택값": ", ".join(selections),
            "예상 경로": " → ".join(p.expected_path[:12]),
            "경로 길이": p.path_length,
            "탈락 경로": "예" if p.is_termination else "아니오",
        })

    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════
# 탭 3: 체크리스트
# ══════════════════════════════════════════════════════════════

def _render_checklist(qa_items: list[dict]):
    """Render QA items grouped by researcher workflow."""
    if not qa_items:
        st.success("이슈가 발견되지 않았습니다! 스크립팅을 진행해도 좋습니다.", icon="✅")
        return

    sev_count = {"critical": 0, "warning": 0, "info": 0}
    for item in qa_items:
        sev_count[item["severity"]] = sev_count.get(item["severity"], 0) + 1

    counts = _workstream_counts(qa_items)
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("설문지 수정", counts["설문지 수정 필요"])
    with col2:
        st.metric("Script 구현 확인", counts["Script 구현 확인"])
    with col3:
        st.metric("링크 테스트 확인", counts["링크 테스트 확인"])
    with col4:
        st.metric("심각/경고", sev_count["critical"] + sev_count["warning"])

    severity_filter = st.radio(
        "심각도 필터",
        options=["all", "critical", "warning", "info"],
        format_func=lambda x: {"all": "전체", "critical": "심각만", "warning": "경고만", "info": "참고만"}[x],
        horizontal=True,
        key="qa_checklist_filter",
    )

    stream_tabs = st.tabs(WORKSTREAMS + ["전체"])
    for idx, stream_name in enumerate(WORKSTREAMS + ["전체"]):
        with stream_tabs[idx]:
            if stream_name == "전체":
                stream_items = qa_items
            else:
                stream_items = [i for i in qa_items if i["workstream"] == stream_name]

            if severity_filter != "all":
                stream_items = [i for i in stream_items if i["severity"] == severity_filter]

            if not stream_items:
                st.info("해당 항목이 없습니다.")
                continue

            _render_checklist_items(stream_items)


def _render_checklist_items(items: list[dict]):
    severity_icon = {"critical": "🔴", "warning": "⚠️", "info": "ℹ️"}
    for item in items:
        icon = severity_icon.get(item["severity"], "")
        sev_label = SEVERITY_LABELS.get(item["severity"], item["severity"])
        with st.expander(f"{icon} [{sev_label}] {item['title']}", expanded=(item["severity"] == "critical")):
            st.markdown(
                f"**문항:** {item['question']}  |  **카테고리:** {item['category']}  |  "
                f"**담당:** {item['owner']}  |  **상태:** {item['status']}"
            )
            st.markdown(item["detail"])
            if item.get("expected"):
                st.caption(f"기대 동작: {item['expected']}")
            st.caption(f"출처: {item['source']}")


# ══════════════════════════════════════════════════════════════
# 통합 엑셀
# ══════════════════════════════════════════════════════════════

def _build_qa_excel(
    result,
    review,
    checklist,
    questions,
    personas,
    qa_items,
    decision,
) -> bytes:
    """Survey QA workbook for research, script implementation, and link-test handoff."""
    wb = Workbook()
    hdr_fill = PatternFill(start_color="0033A0", end_color="0033A0", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="top")

    def _style_header(ws):
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center

    def _finish_sheet(ws, widths):
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    critical_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    warning_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    ws.append(["항목", "값"])
    _style_header(ws)
    counts = _workstream_counts(qa_items)
    ws.append(["상태", decision["label"]])
    ws.append(["판단", decision["message"]])
    ws.append(["전체 문항", result.total_questions if result else 0])
    ws.append(["전체 경로", result.total_paths if result else 0])
    ws.append(["분기 커버리지", f"{result.branch_coverage_percent:.0f}%" if result else ""])
    ws.append(["테스트 시나리오", len(result.test_scenarios) if result else 0])
    ws.append(["대표 응답자 경로", len(personas or [])])
    ws.append(["설문지 수정 필요", counts["설문지 수정 필요"]])
    ws.append(["Script 구현 확인", counts["Script 구현 확인"]])
    ws.append(["링크 테스트 확인", counts["링크 테스트 확인"]])
    ws.append(["파싱 불가 조건", len(result.unparsed_conditions) if result else 0])
    _finish_sheet(ws, {"A": 24, "B": 90})

    # ── Sheet 2: Logic Map ──
    ws_logic = wb.create_sheet("Logic Map")
    logic_headers = ["유형", "기준 문항", "조건", "대상", "확인 포인트", "상태"]
    ws_logic.append(logic_headers)
    _style_header(ws_logic)
    for row in _build_logic_map_rows(questions or []):
        ws_logic.append([row.get(h, "") for h in logic_headers])
    _finish_sheet(ws_logic, {"A": 16, "B": 16, "C": 45, "D": 18, "E": 55, "F": 16})

    # ── Sheet 3: Branch Test ──
    ws_branch = wb.create_sheet("Branch Test")
    branch_headers = ["#", "우선순위", "테스트 목적", "응답 선택", "예상 경로", "확인 포인트"]
    ws_branch.append(branch_headers)
    _style_header(ws_branch)
    for row in _scenario_rows(questions or [], result) if result else []:
        ws_branch.append([row.get(h, "") for h in branch_headers])
    _finish_sheet(ws_branch, {"A": 5, "B": 10, "C": 35, "D": 45, "E": 55, "F": 35})

    # ── Sheet 4: Respondent Paths ──
    ws_persona = wb.create_sheet("Respondent Paths")
    persona_headers = ["#", "응답자 유형", "선택값", "예상 경로", "경로 길이", "탈락 경로"]
    ws_persona.append(persona_headers)
    _style_header(ws_persona)
    for p in personas or []:
        selections = [
            _format_answer_selection(qn, code, p.answer_labels.get(qn, ""))
            for qn, code in p.answer_selections.items()
        ]
        ws_persona.append([
            p.persona_id,
            p.persona_label,
            ", ".join(selections),
            " → ".join(p.expected_path[:20]),
            p.path_length,
            "예" if p.is_termination else "아니오",
        ])
    _finish_sheet(ws_persona, {"A": 5, "B": 35, "C": 45, "D": 60, "E": 10, "F": 12})

    # ── Sheet 5: Checklist ──
    ws_check = wb.create_sheet("Checklist")
    check_headers = [
        "순서", "상태", "담당", "업무 구분", "심각도", "카테고리",
        "문항", "제목", "상세", "기대 동작", "메모",
    ]
    ws_check.append(check_headers)
    _style_header(ws_check)

    for seq, item in enumerate(qa_items, start=1):
        sev_label = SEVERITY_LABELS.get(item["severity"], item["severity"])
        row_num = ws_check.max_row + 1
        ws_check.append([
            seq,
            item.get("status", "미확인"),
            item.get("owner", ""),
            item.get("workstream", ""),
            sev_label,
            item.get("category", ""),
            item.get("question", ""),
            item.get("title", ""),
            item.get("detail", ""),
            item.get("expected", ""),
            "",
        ])
        fill = critical_fill if item["severity"] == "critical" else (
            warning_fill if item["severity"] == "warning" else None
        )
        if fill:
            for cell in ws_check[row_num]:
                cell.fill = fill

    _finish_sheet(
        ws_check,
        {"A": 6, "B": 12, "C": 16, "D": 18, "E": 10, "F": 16,
         "G": 12, "H": 40, "I": 60, "J": 45, "K": 30},
    )

    # ── Sheet 6: Unparsed Conditions ──
    ws_unparsed = wb.create_sheet("Unparsed")
    ws_unparsed.append(["문항", "조건", "확인 방법"])
    _style_header(ws_unparsed)
    for qn, cond in (result.unparsed_conditions if result else []):
        ws_unparsed.append([
            qn,
            cond,
            "원본 설문지의 조건을 확인하고 링크 테스트에서 해당 조건 충족/미충족 케이스를 모두 확인",
        ])
    _finish_sheet(ws_unparsed, {"A": 12, "B": 55, "C": 70})

    for ws_any in wb.worksheets:
        for row in ws_any.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
