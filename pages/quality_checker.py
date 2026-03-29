"""Pre-Scripting Review — 스크립팅 전 설문지 오류 검토.

알고리즘 검출 (즉시): 보기 코드, 스킵 로직, 필터, 데드엔드, 문항번호 중복, 유형 불일치
AI 검출 (선택): 오타, MECE 검증, 논리 충돌
"""

import io
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from services.prescripting_checker import (
    run_algorithmic_checks,
    run_ai_checks,
    ReviewReport,
    ReviewItem,
    CATEGORY_LABELS,
    SEVERITY_LABELS,
)


def page_quality_checker():
    st.title("Pre-Scripting Review")

    # Guard
    if "survey_document" not in st.session_state or st.session_state["survey_document"] is None:
        st.warning("먼저 Questionnaire Analyzer에서 문서를 처리해주세요.", icon="⚠️")
        return

    survey_doc = st.session_state["survey_document"]
    questions = survey_doc.questions
    if not questions:
        st.warning("문서에서 문항을 찾을 수 없습니다.", icon="⚠️")
        return

    st.info(
        f"**{survey_doc.filename}**에서 **{len(questions)}**개 문항을 검토합니다.",
        icon="🔍",
    )

    # ── 알고리즘 검출 (항상 즉시 실행) ──
    report = run_algorithmic_checks(questions)
    st.session_state["review_report"] = report

    # ── AI 검출 옵션 ──
    with st.expander("AI 검출 (선택사항 — 오타, MECE, 논리 충돌)", expanded=False):
        st.caption("LLM을 사용하여 추가 검출을 실행합니다. 시간이 소요될 수 있습니다.")
        ai_clicked = st.button("AI 검출 실행", key="ai_check_btn")

        if ai_clicked:
            with st.status("AI 검출 중...", expanded=True) as status:
                progress_bar = st.progress(0)
                log_area = st.empty()
                batch_done = [0]
                total_batches = [1]

                def _cb(event, data):
                    if event == "batch_start":
                        total_batches[0] = data["total_batches"]
                        log_area.text(f"배치 {data['batch_index'] + 1}/{data['total_batches']} 처리 중...")
                    elif event == "batch_done":
                        batch_done[0] += 1
                        progress_bar.progress(batch_done[0] / total_batches[0])
                        log_area.text(f"배치 {data['batch_index'] + 1}/{data['total_batches']} 완료 ({data['issues_found']}건)")

                ai_items = run_ai_checks(questions, language="ko", progress_callback=_cb)
                report.items.extend(ai_items)
                # 재정렬
                sev_order = {"critical": 0, "warning": 1, "info": 2}
                report.items.sort(key=lambda x: (sev_order.get(x.severity, 9), x.question_number))
                st.session_state["review_report"] = report
                status.update(label=f"AI 검출 완료! {len(ai_items)}건 추가 발견", state="complete")

    # ── 결과 표시 ──
    _render_report(report)


def _render_report(report: ReviewReport):
    """검토 결과 리포트 렌더링."""
    st.divider()

    # ── 메트릭 ──
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("전체 문항", report.total_questions)
    with col2:
        st.metric("심각", report.critical_count,
                  delta=None if report.critical_count == 0 else f"{report.critical_count}건",
                  delta_color="inverse")
    with col3:
        st.metric("경고", report.warning_count)
    with col4:
        st.metric("참고", report.info_count)

    if not report.has_issues:
        st.success("검토 결과 이슈가 발견되지 않았습니다! 스크립팅을 진행해도 좋습니다.", icon="✅")
        return

    st.caption(f"총 {len(report.items)}건의 이슈가 발견되었습니다.")

    # ── 카테고리별 분포 ──
    cat_counts = {}
    for item in report.items:
        cat = CATEGORY_LABELS.get(item.category, item.category)
        cat_counts[cat] = cat_counts.get(cat, 0) + 1

    if cat_counts:
        st.subheader("카테고리별 분포")
        max_count = max(cat_counts.values())
        for cat, count in sorted(cat_counts.items(), key=lambda x: -x[1]):
            col_label, col_bar, col_count = st.columns([2, 6, 1])
            with col_label:
                st.text(cat)
            with col_bar:
                st.progress(count / max_count if max_count > 0 else 0)
            with col_count:
                st.text(str(count))

    # ── 심각도 필터 ──
    st.divider()
    severity_filter = st.radio(
        "필터",
        options=["all", "critical", "warning", "info"],
        format_func=lambda x: {"all": "전체", "critical": "심각만", "warning": "경고만", "info": "참고만"}[x],
        horizontal=True,
        key="review_severity_filter",
    )

    filtered = report.items
    if severity_filter != "all":
        filtered = [i for i in report.items if i.severity == severity_filter]

    if not filtered:
        st.info("선택한 필터에 해당하는 이슈가 없습니다.")
        return

    # ── 이슈 리스트 ──
    st.subheader(f"이슈 목록 ({len(filtered)}건)")

    severity_icon = {"critical": "🔴", "warning": "⚠️", "info": "ℹ️"}

    for item in filtered:
        icon = severity_icon.get(item.severity, "")
        sev_label = SEVERITY_LABELS.get(item.severity, item.severity)
        cat_label = CATEGORY_LABELS.get(item.category, item.category)

        with st.expander(f"{icon} [{sev_label}] {item.title}", expanded=(item.severity == "critical")):
            st.markdown(f"**카테고리:** {cat_label}  |  **문항:** {item.question_number}")
            st.markdown(item.detail)

    # ── 엑셀 다운로드 ──
    st.divider()
    excel_data = _export_review_excel(report)
    st.download_button(
        label="검토 결과 다운로드 (Excel)",
        data=excel_data,
        file_name="pre_scripting_review.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


def _export_review_excel(report: ReviewReport) -> bytes:
    """검토 결과를 엑셀 체크리스트로 출력."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pre-Scripting Review"

    header_fill = PatternFill(start_color="0033A0", end_color="0033A0", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center")

    headers = ["확인", "심각도", "카테고리", "문항번호", "이슈 제목", "상세 설명"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    critical_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    warning_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

    for item in report.items:
        sev_label = SEVERITY_LABELS.get(item.severity, item.severity)
        cat_label = CATEGORY_LABELS.get(item.category, item.category)
        row_num = ws.max_row + 1
        ws.append(["☐", sev_label, cat_label, item.question_number, item.title, item.detail])

        # 심각도별 배경색
        if item.severity == "critical":
            for cell in ws[row_num]:
                cell.fill = critical_fill
        elif item.severity == "warning":
            for cell in ws[row_num]:
                cell.fill = warning_fill

    # 컬럼 너비
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 60

    # wrap text
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
