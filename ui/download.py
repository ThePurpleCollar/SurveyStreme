import io
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def df_for_download(processed_df):
    """다운로드용 DataFrame 준비 (컬럼 순서 정리 및 빈 컬럼 추가)"""
    processed_df = processed_df.copy()
    for col in ['Sort', 'TableTitle', 'SubBanner', 'NetRecode',
                 'BannerIDs', 'SpecialInstructions', 'GrammarChecker', 'Other']:
        if col not in processed_df.columns:
            processed_df[col] = ''

    # 컬럼 순서: Sort, QN, TN, QText, Title, SubBanner, QType, SType, NetRecode, BannerIDs, SpecialInstructions, Other, GrammarChecker
    base_columns = ['Sort', 'QuestionNumber', 'TableNumber', 'QuestionText',
                    'TableTitle', 'SubBanner', 'QuestionType', 'SummaryType',
                    'NetRecode', 'BannerIDs', 'SpecialInstructions', 'Other', 'GrammarChecker']

    # DOCX 추출에서 추가되는 컬럼들
    extra_columns = ['AnswerOptions', 'SkipLogic', 'Filter', 'Instructions']
    all_columns = base_columns + [c for c in extra_columns if c in processed_df.columns]

    return processed_df.reindex(columns=all_columns)


def prepare_excel_download(survey_doc) -> bytes:
    """SurveyDocument를 서식이 적용된 Excel 파일로 변환.

    Sheet 1: 메인 문항 테이블 (모든 필드 포함)
    Sheet 2: 응답 보기 flat 테이블
    Sheet 3: Banner Spec (배너 정의, 있을 경우)
    """
    wb = Workbook()
    header_fill = PatternFill(start_color="0033A0", end_color="0033A0", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    center_align = Alignment(horizontal='center', vertical='center')

    # ── Sheet 1: Questions ──
    ws_main = wb.active
    ws_main.title = "Questions"

    headers = [
        "Sort", "QuestionNumber", "TableNumber", "QuestionText",
        "TableTitle", "SubBanner", "QuestionType", "SummaryType",
        "NetRecode", "BannerIDs", "SpecialInstructions",
        "AnswerOptions", "SkipLogic", "Filter",
        "Instructions", "GrammarChecker",
    ]
    ws_main.append(headers)

    for cell in ws_main[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for q in survey_doc.questions:
        ws_main.append([
            q.sort_order,
            q.question_number,
            q.table_number,
            q.question_text,
            q.table_title,
            q.sub_banner,
            q.question_type or "",
            q.summary_type,
            q.net_recode,
            q.banner_ids,
            q.special_instructions,
            q.answer_options_display(),
            q.skip_logic_display(),
            q.filter_condition or "",
            q.instructions or "",
            q.grammar_checked,
        ])

    for row in ws_main.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    col_widths = [12, 15, 12, 50, 35, 20, 12, 25, 30, 12, 35, 40, 30, 25, 20, 25]
    for i, width in enumerate(col_widths, 1):
        ws_main.column_dimensions[get_column_letter(i)].width = width

    # ── Sheet 2: Answer Options ──
    ws_opts = wb.create_sheet("AnswerOptions")
    ws_opts.append(["QuestionNumber", "OptionCode", "OptionLabel"])

    for cell in ws_opts[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for q in survey_doc.questions:
        for opt in q.answer_options:
            ws_opts.append([q.question_number, opt.code, opt.label])

    ws_opts.column_dimensions['A'].width = 18
    ws_opts.column_dimensions['B'].width = 12
    ws_opts.column_dimensions['C'].width = 50

    # ── Sheet 3: Banner Spec (있을 경우) ──
    if hasattr(survey_doc, 'banners') and survey_doc.banners:
        ws_banner = wb.create_sheet("Banner Spec")
        ws_banner.append(["BannerID", "BannerName", "PointID", "PointLabel",
                          "SourceQuestion", "Condition", "IsNet", "NetDefinition"])

        for cell in ws_banner[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        for banner in survey_doc.banners:
            for pt in banner.points:
                ws_banner.append([
                    banner.banner_id,
                    banner.name,
                    pt.point_id,
                    pt.label,
                    pt.source_question,
                    pt.condition,
                    "Yes" if pt.is_net else "No",
                    pt.net_definition,
                ])

        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws_banner.column_dimensions[col_letter].width = 18

    # ── Sheet 3b: Banner Layout (가로 Cross-Tab, DP용) ──
    if hasattr(survey_doc, 'banners') and survey_doc.banners:
        ws_xtab = wb.create_sheet("Banner Layout")
        cat_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
        banner_fill = PatternFill(start_color="C5CAE9", end_color="C5CAE9", fill_type="solid")
        cond_font = Font(size=9, color="666666", italic=True)

        cat_order = []
        cat_map = {}
        for b in survey_doc.banners:
            cat = getattr(b, 'category', '') or "기타"
            if cat not in cat_map:
                cat_map[cat] = []
                cat_order.append(cat)
            cat_map[cat].append(b)

        col = 1
        for cat_name in cat_order:
            cat_start_col = col
            for banner in cat_map[cat_name]:
                banner_start_col = col
                for pt in banner.points:
                    ws_xtab.cell(row=3, column=col, value=pt.label)
                    ws_xtab.cell(row=3, column=col).font = Font(bold=True)
                    ws_xtab.cell(row=3, column=col).alignment = center_align
                    ws_xtab.cell(row=4, column=col, value=pt.condition or "")
                    ws_xtab.cell(row=4, column=col).font = cond_font
                    col += 1
                ws_xtab.cell(row=3, column=col, value="Total")
                ws_xtab.cell(row=3, column=col).font = Font(bold=True)
                btype = " [Composite]" if getattr(banner, 'banner_type', '') == "composite" else ""
                ws_xtab.cell(row=2, column=banner_start_col,
                             value=f"Banner {banner.banner_id}: {banner.name}{btype}")
                ws_xtab.cell(row=2, column=banner_start_col).font = Font(bold=True)
                ws_xtab.cell(row=2, column=banner_start_col).fill = banner_fill
                if col > banner_start_col:
                    ws_xtab.merge_cells(start_row=2, start_column=banner_start_col,
                                        end_row=2, end_column=col)
                col += 1

            cat_end_col = col - 1
            if cat_end_col >= cat_start_col:
                ws_xtab.cell(row=1, column=cat_start_col, value=cat_name)
                ws_xtab.cell(row=1, column=cat_start_col).font = Font(bold=True, size=11)
                ws_xtab.cell(row=1, column=cat_start_col).fill = cat_fill
                ws_xtab.cell(row=1, column=cat_start_col).alignment = center_align
                if cat_end_col > cat_start_col:
                    ws_xtab.merge_cells(start_row=1, start_column=cat_start_col,
                                        end_row=1, end_column=cat_end_col)

        for i in range(1, col):
            ws_xtab.column_dimensions[get_column_letter(i)].width = 15

    # ── Sheet 4: Net/Recode Spec (있을 경우) ──
    seen_qn = set()
    has_net = False
    for q in survey_doc.questions:
        if q.question_number not in seen_qn and q.net_recode:
            has_net = True
            break

    if has_net:
        ws_net = wb.create_sheet("Net Recode Spec")
        ws_net.append(["QuestionNumber", "QuestionType", "NetRecode"])

        for cell in ws_net[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        seen_qn = set()
        for q in survey_doc.questions:
            if q.question_number in seen_qn:
                continue
            seen_qn.add(q.question_number)
            if q.net_recode:
                ws_net.append([q.question_number, q.question_type or "", q.net_recode])

        ws_net.column_dimensions['A'].width = 18
        ws_net.column_dimensions['B'].width = 15
        ws_net.column_dimensions['C'].width = 50

    # 바이트로 변환
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def render_download_buttons(page_name: str, include_excel: bool = False):
    """CSV (+ 선택적 Excel) 다운로드 버튼 렌더링."""
    if 'edited_df' not in st.session_state:
        return

    prepared_df = df_for_download(st.session_state['edited_df'])
    base_name = st.session_state.get('uploaded_file_name', 'data')
    csv_bytes = prepared_df.to_csv(index=False).encode('utf-8-sig')
    csv_filename = f"{base_name}_{page_name}.csv"

    if include_excel and 'survey_document' in st.session_state:
        col1, col2 = st.columns(2)
        with col1:
            st.download_button(
                label="CSV 다운로드",
                data=csv_bytes,
                file_name=csv_filename,
                mime='text/csv',
            )
        with col2:
            excel_data = prepare_excel_download(st.session_state['survey_document'])
            excel_filename = f"{base_name}_{page_name}.xlsx"
            st.download_button(
                label="Excel 다운로드",
                data=excel_data,
                file_name=excel_filename,
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
    else:
        st.download_button(
            label="CSV 다운로드",
            data=csv_bytes,
            file_name=csv_filename,
            mime='text/csv',
        )
