import io
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook

from models.survey import AnswerOption, SkipLogic, SurveyQuestion
from pages.survey_qa import (
    _build_logic_map_rows,
    _build_qa_excel,
    _collect_qa_items,
    _qa_decision,
    _scenario_rows,
    _workstream_counts,
)
from services.checklist_generator import generate_checklist
from services.path_simulator import generate_persona_scenarios, simulate_paths
from services.prescripting_checker import run_algorithmic_checks
from services.skip_logic_service import build_skip_logic_graph
from services.skip_logic_service import parse_target


questions = [
    SurveyQuestion(
        question_number="SC1",
        question_text="What is your age?",
        question_type="SA",
        answer_options=[
            AnswerOption("1", "17 years old or younger"),
            AnswerOption("2", "18 years old or older"),
        ],
        skip_logic=[SkipLogic(condition="SC1=1", target="TERMINATE")],
        role="screening",
    ),
    SurveyQuestion(
        question_number="SC2",
        question_text="What is your gender?",
        question_type="SA",
        answer_options=[AnswerOption("1", "Male"), AnswerOption("2", "Female")],
        role="demographics",
    ),
    SurveyQuestion(
        question_number="Q1",
        question_text="Do you own a TV?",
        question_type="SA",
        answer_options=[AnswerOption("1", "Yes"), AnswerOption("2", "No")],
        filter_condition="SC1=2",
    ),
    SurveyQuestion(
        question_number="Q2",
        question_text="Which brands are you aware of?",
        question_type="MA",
        answer_options=[
            AnswerOption("1", "Samsung"),
            AnswerOption("2", "LG"),
            AnswerOption("99", "None of the above"),
        ],
        instructions="ROTATE options",
    ),
]

result = simulate_paths(questions)
review = run_algorithmic_checks(questions)
checklist = generate_checklist(questions, language="ko", use_llm=False)
personas = generate_persona_scenarios(questions, build_skip_logic_graph(questions))
qa_items = _collect_qa_items(result, review, checklist)
decision = _qa_decision(result, qa_items)
counts = _workstream_counts(qa_items)

assert "Script 구현 확인" in counts
assert counts["링크 테스트 확인"] >= 1
assert any(i["workstream"] == "Script 구현 확인" for i in qa_items)
assert decision["label"] == "확인 후 스크립팅 전달 권장"

logic_rows = _build_logic_map_rows(questions)
assert any(r["유형"] == "스킵/종료" for r in logic_rows)
assert any(r["유형"] == "필터/대상 조건" for r in logic_rows)

scenario_rows = _scenario_rows(questions, result)
assert scenario_rows
assert "테스트 목적" in scenario_rows[0]
assert "응답 선택" in scenario_rows[0]

excel_bytes = _build_qa_excel(result, review, checklist, questions, personas, qa_items, decision)
wb = load_workbook(io.BytesIO(excel_bytes))
assert wb.sheetnames == [
    "Summary",
    "Logic Map",
    "Branch Test",
    "Branch Diagnostics",
    "Respondent Paths",
    "Checklist",
    "Unparsed",
]
assert "업무 구분" in [cell.value for cell in wb["Checklist"][1]]
assert "분기" in [cell.value for cell in wb["Branch Diagnostics"][1]]
assert "Script 구현 확인" in [row[0] for row in wb["Summary"].iter_rows(min_row=2, max_col=1, values_only=True)]

assert parse_target("END") == "END"
assert parse_target("Go to END") == "END"
assert parse_target("End of survey") == "END"
assert parse_target("Go to end of section 5") is None
assert parse_target("Read until the end") is None
assert parse_target("by the end of the day") is None
branch_questions = [
    SurveyQuestion(
        question_number="S1",
        question_text="Screening age",
        question_type="SA",
        answer_options=[AnswerOption("1", "Under 18"), AnswerOption("2", "18+")],
        skip_logic=[SkipLogic(condition="S1=1", target="END")],
    ),
    SurveyQuestion(
        question_number="Q1",
        question_text="Own a TV?",
        question_type="SA",
        answer_options=[AnswerOption("1", "Yes"), AnswerOption("2", "No")],
        skip_logic=[SkipLogic(condition="Q1=2", target="Q3")],
    ),
    SurveyQuestion(
        question_number="Q2",
        question_text="Main TV brand",
        question_type="SA",
        answer_options=[AnswerOption("1", "Samsung"), AnswerOption("2", "LG")],
    ),
    SurveyQuestion(
        question_number="Q3",
        question_text="Purchase intent",
        question_type="SA",
        answer_options=[AnswerOption("1", "High"), AnswerOption("2", "Low")],
    ),
]
branch_result = simulate_paths(branch_questions)
assert branch_result.total_skip_rules == 2
assert len(branch_result.test_scenarios) == 2
assert branch_result.branch_coverage_percent == 100.0

print("ALL LOGIC CHECKER RESEARCHER UI TESTS PASSED")
