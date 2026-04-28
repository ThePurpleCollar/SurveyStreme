"""Smoke tests for DP handoff two-sheet Excel export."""

import io
import os
import sys

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from models.survey import AnswerOption, Banner, BannerPoint, SurveyDocument, SurveyQuestion
from services.table_guide_service import (
    build_dp_handoff_validation_summary,
    compile_table_guide,
    export_dp_handoff_excel,
    export_table_guide_excel,
)
from ui.download import prepare_excel_download


def _opts(*items):
    return [AnswerOption(code=str(code), label=label) for code, label in items]


questions = [
    SurveyQuestion(
        question_number="S1",
        source_variable="SCR_GENDER",
        table_number="T001",
        question_text="What is your gender?",
        question_type="SA",
        answer_options=_opts((1, "Male"), (2, "Female")),
        table_title="Gender",
        sort_order="by code",
        role="demographics",
        variable_type="demographic",
    ),
    SurveyQuestion(
        question_number="Q1",
        source_variable="PI_Q1",
        table_number="T002",
        question_text="How likely are you to purchase?",
        question_type="5PT SCALE",
        answer_options=_opts(
            (1, "Definitely would"),
            (2, "Probably would"),
            (3, "Might or might not"),
            (4, "Probably would not"),
            (5, "Definitely would not"),
        ),
        table_title="Purchase Intent",
        sort_order="by code",
        banner_ids="A",
        role="intent_loyalty",
    ),
]

banners = [
    Banner(
        banner_id="A",
        name="Gender",
        category="Demographics",
        points=[
            BannerPoint(point_id="BP_A_1", label="Male", source_question="S1", condition="S1=1"),
            BannerPoint(point_id="BP_A_2", label="Female", source_question="S1", condition="S1=2"),
        ],
    ),
    Banner(
        banner_id="B",
        name="Invalid Gender Code",
        category="Demographics",
        points=[
            BannerPoint(point_id="BP_B_1", label="Invalid", source_question="S1", condition="S1=9"),
        ],
    ),
]

doc = SurveyDocument(filename="sample.docx", questions=questions, banners=banners)
tg_doc = compile_table_guide(doc, project_name="Sample", language="en")
summary = build_dp_handoff_validation_summary(tg_doc, doc)
assert summary["table_total"] == 2
assert summary["banner_total"] == 3
assert summary["banner_review"] == 1
data = export_dp_handoff_excel(tg_doc, doc)
wb = load_workbook(io.BytesIO(data), read_only=True)

assert wb.sheetnames == ["Table Guide", "Banner Spec"], wb.sheetnames

tg = wb["Table Guide"]
tg_headers = [cell.value for cell in next(tg.iter_rows(min_row=1, max_row=1))]
assert "DP Review Status" in tg_headers
assert "AnswerOptions" in tg_headers
assert "BannerNames" in tg_headers

tg_rows = list(tg.iter_rows(min_row=2, values_only=True))
q1_row = next(row for row in tg_rows if row[tg_headers.index("QuestionNumber")] == "Q1")
assert q1_row[tg_headers.index("DP Review Status")] == "Ready for DP"
assert q1_row[tg_headers.index("SourceVariable")] == "PI_Q1"
assert "1. Definitely would" in q1_row[tg_headers.index("AnswerOptions")]
assert q1_row[tg_headers.index("BannerNames")] == "A(Gender)"

banner = wb["Banner Spec"]
banner_headers = [cell.value for cell in next(banner.iter_rows(min_row=1, max_row=1))]
assert "HumanCondition" in banner_headers
assert "SPSSCondition" in banner_headers
assert "CodeLabels" in banner_headers

banner_rows = list(banner.iter_rows(min_row=2, values_only=True))
male_row = next(row for row in banner_rows if row[banner_headers.index("BannerValueLabel")] == "Male")
assert male_row[banner_headers.index("DP Review Status")] == "Ready for DP"
assert male_row[banner_headers.index("SourceVariable")] == "SCR_GENDER"
assert male_row[banner_headers.index("SPSSCondition")] == "SCR_GENDER = 1"
assert male_row[banner_headers.index("CodeLabels")] == "S1 1=Male"

invalid_row = next(row for row in banner_rows if row[banner_headers.index("BannerValueLabel")] == "Invalid")
assert invalid_row[banner_headers.index("DP Review Status")] == "Needs Researcher Review"
assert "Invalid code(s) for S1: 9" in invalid_row[banner_headers.index("QA Warning")]

internal_data = export_table_guide_excel(tg_doc, doc)
internal_wb = load_workbook(io.BytesIO(internal_data), read_only=True)
internal_tg_headers = [
    cell.value
    for cell in next(internal_wb["Table Guide"].iter_rows(min_row=1, max_row=1))
]
assert "ReviewStatus" in internal_tg_headers
assert "QAWarning" in internal_tg_headers
assert "SourceVariable" in internal_tg_headers
assert "AnswerOptions" in internal_tg_headers

internal_banner_headers = [
    cell.value
    for cell in next(internal_wb["Banner Spec"].iter_rows(min_row=1, max_row=1))
]
assert "SPSSCondition" in internal_banner_headers
assert "CodeLabels" in internal_banner_headers
assert "Rationale(KO)" in internal_banner_headers

analyzer_data = prepare_excel_download(doc)
analyzer_wb = load_workbook(io.BytesIO(analyzer_data), read_only=True)
analyzer_headers = [
    cell.value
    for cell in next(analyzer_wb["Questions"].iter_rows(min_row=1, max_row=1))
]
assert "SourceVariable" in analyzer_headers
analyzer_banner_headers = [
    cell.value
    for cell in next(analyzer_wb["Banner Spec"].iter_rows(min_row=1, max_row=1))
]
assert "Category" in analyzer_banner_headers
assert "SourceVariable" in analyzer_banner_headers
assert "SPSSCondition" in analyzer_banner_headers
assert "CodeLabels" in analyzer_banner_headers
assert "Rationale(KO)" in analyzer_banner_headers
analyzer_banner_rows = list(analyzer_wb["Banner Spec"].iter_rows(min_row=2, values_only=True))
analyzer_male_row = next(
    row for row in analyzer_banner_rows
    if row[analyzer_banner_headers.index("PointLabel")] == "Male"
)
assert analyzer_male_row[analyzer_banner_headers.index("SourceVariable")] == "SCR_GENDER"
assert analyzer_male_row[analyzer_banner_headers.index("SPSSCondition")] == "SCR_GENDER = 1"
assert analyzer_male_row[analyzer_banner_headers.index("CodeLabels")] == "S1 1=Male"

answer_option_headers = [
    cell.value
    for cell in next(analyzer_wb["AnswerOptions"].iter_rows(min_row=1, max_row=1))
]
assert "SourceVariable" in answer_option_headers

print("ALL DP HANDOFF EXPORT TESTS PASSED")
